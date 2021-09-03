import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal
import os
from pathlib import Path
import numpy as np
import matplotlib.gridspec as gridspec
import csv
import openpyxl as excel
from openpyxl.utils import get_column_letter
from scipy.stats import tstd


def oscillograms():
    test = ProcessSignal('200925')
    csv_types = test.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = test.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_dicts['magnetron']
    for signal in csv_signals:
        file = test.open_file(signal, reduced=True)
        u = file['voltage']
        t = file['time']
        dt = file['time_resolution']
        u_filt = test.bandpass_filter(t, u, 2.725e9, 2.755e9)
        num = signal[3:6]
        if int(num) < 66:
            absorbers = '№2, 3'
        elif 66 <= int(num) < 125 and num in magnetron_nums:
            absorbers = 0
        elif 125 <= int(num) <= 184 and num in magnetron_nums:
            absorbers = '№2, 3, A'
        else:
            absorbers = 0
        pl_d_nums = test.read_excel(csv_signal_nums)['dicts'].keys()
        if num in pl_d_nums:
            pl_density = test.read_excel(csv_signal_nums)['dicts'][num]['Ток плазмы, А']
            test.oscill_picture(num, t, u, u_filt, pl_density, absorbers, save=True)
#oscillograms()


def one_signal_oscillogramm(signal_num, voltage_num, exp_num, central_freq=2.71e9):
    test = ProcessSignal(str(exp_num))
    signal_file = test.open_file(signal_num, reduced=False)
    print('Signal {} data obtained'.format(signal_num[3:6]))
    pl_density = test.read_excel(signal_num)['dicts'][signal_num[3:6]]['Ток плазмы, А']
    voltage_file = test.open_file(voltage_num, reduced=False)
    print('Voltage {} data obtained'.format(voltage_num[3:6]))
    signal_u = signal_file['voltage']
    signal_t = signal_file['time'] + 50e-9
    voltage_u = voltage_file['voltage']
    voltage_t = voltage_file['time']

    ind_y_shift = signal_t < 20e-9
    y_shift = np.mean(signal_u[ind_y_shift])
    u_shift = signal_u - y_shift
    try:
        filt_freq_min, filt_freq_max = central_freq - 15e6, central_freq + 15e6
        u_filt = test.fft_filter(signal_t, signal_u, filt_freq_min, filt_freq_max)
        #u_filt_1 = test.bandpass_filter(signal_t, signal_u, 2.695e9, 2.725e9)
        print('Filtering done')
    except:
        pass
    #pl_density = test.read_excel(signal_num)['dicts'][signal_num[3:6]]['Ток плазмы, А']
    #magnetron_abs = test.read_excel(signal_num)['dicts'][signal_num[3:6]]['Поглотители в тракте магнетрона']
    fig = plt.figure(num=1, dpi=150)
    ax = fig.add_subplot(111)
    print('Creating a picture...')
    line1, = ax.plot(signal_t / 1e-9, u_shift, linewidth=0.7, color='dodgerblue')
    #line1, = ax.plot(signal_t / 1e-9, signal_u, linewidth=0.7, color='dodgerblue')
    line1.set_label('СВЧ сигнал')

    try:
        line2, = ax.plot(signal_t / 1e-9, u_filt, linewidth=0.7, color='red')
        line2.set_label('Фильтрованный сигнал (2,71 +/- 0.015) ГГц')
        
        #line4, = ax.plot(signal_t / 1e-9, u_filt_1, linewidth=0.7, color='green')
        #line4.set_label('Bandpass (2,74 +/- 0.03) ГГц')
    except:
        pass
    line3, = ax.plot(voltage_t / 1e-9, voltage_u, linewidth=0.7, color='blue')
    line3.set_label('Импульс напряжения')
    ax.set_xlabel(r'$Время, нс$', fontsize=14, fontweight='black')
    ax.set_ylabel(r'$Напряжение, В$', fontsize=14, fontweight='black')
    ax.set_ylim(bottom=-2.5, top=2.5)
    #ax.legend()
    ax.grid(which='both', axis='both')
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    '''
    if magnetron_abs is not None:
        plt.title('№ {}, n = {}, Поглотители = {}'.format(signal_num[3:6], pl_density, magnetron_abs))
    else:
        plt.title('№ {}, n = {}'.format(signal_num[3:6], pl_density))
    '''
    plt.title(f'№ {signal_num[3:6]}-{voltage_num[3:6]},n = {pl_density}')
    png_name = test.signal_pics_path / f'all_ocs_{pl_density}_{signal_num[3:6]}.png'
    fig.savefig(png_name)
    plt.close(fig)


def exp_oscillogramms(exp_num, last_num=0, first_num=0):
    test = ProcessSignal(str(exp_num))
    signal_nums_0 = test.read_type_file()['signal_files']
    voltage_nums_0 = test.read_type_file()['voltage_files']
    if last_num == 0:
        signal_nums, voltage_nums = signal_nums_0, voltage_nums_0
    else:
        signal_nums = [signal_nums_0[i] for i in range(len(signal_nums_0)) if first_num < int(signal_nums_0[i][3:6]) < last_num]
        voltage_nums = [voltage_nums_0[i] for i in range(len(voltage_nums_0)) if first_num < int(voltage_nums_0[i][3:6]) < last_num]
    print('Signals:', signal_nums)
    for element in zip(signal_nums, voltage_nums):
        one_signal_oscillogramm(element[0], element[1], exp_num)


#exp_oscillogramms(210708, first_num=0)


def magnetron_osc(exp_num):
    test = ProcessSignal(str(exp_num))
    path = r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}'.format(str(exp_num))
    file_list = os.listdir(path)
    for file in file_list:
        if 'str' in file:
            data = test.open_file(file, reduced=False)
            t, u = data['time'], data['voltage']
            fig = plt.figure(num=1, dpi=150)
            ax = fig.add_subplot(111)
            print(f'Creating a picture {file}...')
            #line1, = ax.plot(t / 1e-9, u, linewidth=0.7)
            line1, = ax.plot(t / 1e-9, u, linewidth=1.2)
            ax.set_xlabel(r'$Время, нс$', fontsize=12, fontweight='black')
            ax.set_ylabel(r'$Напряжение, В$', fontsize=12, fontweight='black')
            ax.grid(which='both', axis='both')
            min_y, max_y = min(u) - 0.05, max(u) + 0.05
            ax.set_ylim(bottom=min_y, top=max_y)
            #ax.set_ylim(bottom=-4, top=4)
            plt.xticks(fontsize=12)
            plt.yticks(fontsize=12)
            plt.title(f'№ {file}')
            plt.show()
            png_name = test.signal_pics_path / 'ocs_{}'.format(file[3:6])
            #fig.savefig(png_name)
            plt.close(fig)

#magnetron_osc(171222)

def double_magnetron_osc(exp_num):
    test = ProcessSignal(str(exp_num))
    path = r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}'.format(str(exp_num))
    file_list = os.listdir(path)
    print(file_list)
    file_list_1 = [file_list[i] for i in range(len(file_list)) if 'csv' in file_list[i] and int(file_list[i][3:6]) % 2 != 0]
    file_list_2 = [file_list[i] for i in range(len(file_list)) if 'csv' in file_list[i] and int(file_list[i][3:6]) % 2 == 0]
    print(file_list_1)
    print(file_list_2)
    for i, file in enumerate(file_list_1):
        if 'str' in file:
            data = test.open_file(file, reduced=False)
            t, u = data['time'], data['voltage']
            data_2 = test.open_file(file_list_2[i], reduced=False)
            t_2, u_2 = data_2['time'], data_2['voltage']
            fig = plt.figure(num=1, dpi=150)
            ax = fig.add_subplot(111)
            print(f'Creating a picture {file}...')
            #line1, = ax.plot(t / 1e-9, u, linewidth=0.7)
            line1, = ax.plot(t / 1e-9, u, linewidth=1.2, color='orange')
            line2, = ax.plot(t_2 / 1e-9, u_2, linewidth=1.2, color='darkmagenta')
            line1.set_label('Ch_1')
            line2.set_label('Ch_3')
            ax.set_xlabel(r'$Время, нс$', fontsize=12, fontweight='black')
            ax.set_ylabel(r'$Напряжение, В$', fontsize=12, fontweight='black')
            ax.grid(which='both', axis='both')
            min_y, max_y = min(u) - 0.05, max(u) + 0.05
            ax.set_ylim(bottom=min_y, top=max_y)
            ax.legend()
            #ax.set_ylim(bottom=-4, top=4)
            plt.xticks(fontsize=12)
            plt.yticks(fontsize=12)
            plt.title(f'№ {file[3:6]}_{file_list_2[i][3:6]}')
            png_name = test.signal_pics_path / 'ocs_{}_{}'.format(file[3:6], file_list_2[i][3:6])
            fig.savefig(png_name)
            plt.close(fig)

#double_magnetron_osc(210803)


def shot_series_nums(exp_num):
    excel_fpath = r"C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\{}.xlsx".format(exp_num, exp_num)
    wb = excel.load_workbook(excel_fpath)
    sheet = wb['Лист1']
    row_max = sheet.max_row
    row_min = 1
    for i in range(1, row_max):
        cell = sheet.cell(row=i, column=1)
        cell_row_val = cell.value
        if cell_row_val == 'Номер файла' and row_min == 1:
            row_min = i
    first_row = [sheet.cell(row=i, column=1).value for i in range(row_min, row_max)]
    second_row = [sheet.cell(row=i, column=2).value for i in range(row_min, row_max)]
    str_inds = [i for i in range(len(second_row)) if type(second_row[i]) == 'str']


def file_nums_oscillogramms(exp_num, file_nums):
    print(file_nums)
    test = ProcessSignal(str(exp_num))
    for file_num in file_nums:
        #file_name = f'str{file_num:03d}.csv'
        file_name = f'str{file_num}.csv'
        data = test.open_file(file_name, reduced=False)
        t, u = data['time'], data['voltage']
        fig = plt.figure(num=1, dpi=150)
        ax = fig.add_subplot(111)
        print(f'Creating a picture {file_num}...')
        line1, = ax.plot(t / 1e-9, u, linewidth=0.7)
        ax.set_xlabel(r'$Время, нс$', fontsize=14, fontweight='black')
        ax.set_ylabel(r'$Напряжение, В$', fontsize=14, fontweight='black')
        ax.grid(which='both', axis='both')
        min_y, max_y = min(u) - 0.05, max(u) + 0.05
        ax.set_ylim(bottom=min_y, top=max_y)
        plt.xticks(fontsize=12)
        plt.yticks(fontsize=12)
        plt.title(f'№ {file_num}')
        png_name = test.signal_pics_path / f'ocs_{file_num}.png'
        fig.savefig(png_name)
        plt.close(fig)


#file_nums_oscillogramms(210707, [f'{i:03d}' for i in range(133, 188)])


def pl_densities_from_excel(exp_num, file_nums_list):
    proc = ProcessSignal(f'{exp_num}')
    all_files_list = os.listdir(proc.exp_file_path)
    pl_d_vals = []
    for file_num in file_nums_list:
        file_name = f'str{file_num}.csv'
        try:
            pl_density = proc.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
        except:
            pl_density = 0
        pl_d_vals.append(pl_density)
    print(f'd_vals:{pl_d_vals}')
    return(pl_d_vals)

#pl_densities_from_excel(210414, [f'{i:03d}' for i in range(121, 151)])

def comments_from_excel(exp_num, file_nums_list):
    proc = ProcessSignal(f'{exp_num}')
    all_files_list = os.listdir(proc.exp_file_path)
    comments = []
    for file_num in file_nums_list:
        file_name = f'str{file_num}.csv'
        try:
            comment = proc.read_excel(file_name)['dicts'][file_num]['Комментарий']
        except:
            comment = ' '
        comments.append(comment)
    return(comments)

#comments_from_excel(210414, [f'{i:03d}' for i in range(121, 151)])


def file_nums_oscillogramms_density_vals(exp_num, file_nums):
    print(file_nums)
    test = ProcessSignal(str(exp_num))
    density_vals = pl_densities_from_excel(exp_num, file_nums)
    comments = comments_from_excel(exp_num, file_nums)
    comment_mtrx = np.reshape(np.asarray(comments), (int(len(comments)/2), 2))
    density_mtrx = np.reshape(np.asarray(density_vals), (int(len(density_vals)/2), 2))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums)/2), 2))
    print(density_mtrx[0, ])
    central_freq = 2.714E9
    t_start = -85E-9
    for j in range(nums_mtrx.shape[0]):
        gs = gridspec.GridSpec(2, 1)
        fig = plt.figure(num=1, dpi=200, figsize=[11.69, 12.27])
        for i in range(nums_mtrx.shape[1]):
            file_num = f'{nums_mtrx[j, i]}'
            file_name = f'str{file_num}.csv'
            data = test.open_file(file_name, reduced=False)
            t, u = data['time'], (data['voltage'] - np.mean(data['voltage']))
            if t[0] > t_start:
                dif_t = t[0] - t_start
                t = t - dif_t
            elif t[0] < t_start:
                dif_t = t_start - t[0]
                t = t + dif_t
            filt_freq_min, filt_freq_max = central_freq - 30e6, central_freq + 30e6
            u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)

            ax = fig.add_subplot(gs[i, 0])
            print(f'Creating a picture {file_num}...')
            line1, = ax.plot(t / 1e-9, u, linewidth=0.7)
            #line2, = ax.plot(t / 1e-9, u_filt, linewidth=0.7)
            ax.set_xlabel(r'$Время, нс$', fontsize=14, fontweight='black')
            ax.set_ylabel(r'$Напряжение, В$', fontsize=14, fontweight='black')
            min_y, max_y = min(u) - 0.25, max(u) + 0.25
            min_y, max_y = -2.5, 2.5
            major_ticks = np.arange(-4, 4, 0.5)
            minor_ticks = np.arange(-4, 4, 0.1)
            ax.set_yticks(major_ticks)
            ax.set_yticks(minor_ticks, minor=True)
            ax.grid(which='both', axis='both')
            ax.set_ylim(bottom=min_y, top=max_y)
            plt.xticks(fontsize=12)
            plt.yticks(fontsize=12)
            ax.set_title(f'{exp_num}, № {file_num}, n={density_mtrx[j, i]}, {comment_mtrx[j, i]}')
            #ax.set_title(f'{exp_num}, № {file_num}')
        #png_name = test.signal_pics_path / f'ocs_{density_mtrx[j,0]}_{nums_mtrx[j, 0]}_{nums_mtrx[j, 1]}.png'
        #plt.show()
        png_name = test.signal_pics_path / f'ocs_{nums_mtrx[j, 0]}_{nums_mtrx[j, 1]}.png'
        fig.savefig(png_name)
        plt.close(fig)

#exception_list = [43, 203]
exception_list = [119, 120]
#file_nums_oscillogramms_density_vals(210622, [f'{i:03d}' for i in range(1, 3) if i not in exception_list])


def write_2_antennas_osc_csv(exp_num, file_nums):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    central_freq = 2.714E9
    for j in range(nums_mtrx.shape[0]):
        antennas_data_dict = {}
        for i in range(nums_mtrx.shape[1]):
            file_num = f'{nums_mtrx[j, i]}'
            file_name = f'str{file_num}.csv'
            data = test.open_file(file_name, reduced=False)
            filt_freq_min, filt_freq_max = central_freq - 15e6, central_freq + 15e6
            if int(nums_mtrx[j, i]) % 2 == 0:
                t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                antennas_data_dict['t_main'], antennas_data_dict['u_main'] = t, u
                antennas_data_dict['u_filt_main'] = u_filt
                plt.plot(t, u)
            else:
                #t, u = data['time'], (data['voltage'] - np.mean(data['voltage'])) / 3.16
                t, u = data['time'], (data['voltage'] - np.mean(data['voltage']))
                u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                antennas_data_dict['t'], antennas_data_dict['u'] = t, u
                antennas_data_dict['u_filt'] = u_filt
                #plt.plot(t + 4.347E-9, u)
                plt.plot(t, u)

        plt.show()
        '''
        print(f'Writing {nums_mtrx[j, i]} csv file...')
        print(antennas_data_dict.keys())

        dif_file_path = test.exp_file_path / '2_antennas_CSV'
        dif_file_path.mkdir(parents=True, exist_ok=True)
        file = open(str(dif_file_path / f'{nums_mtrx[j, 0]}_{nums_mtrx[j, 1]}.csv'), 'w', newline='')
        with file:
            writer = csv.writer(file)
            writer.writerow(['t_main(s)', 'V_main(V)', 'V_main_filt(V)', 't(s)', 'V(V)', 'V_filt(V)'])
            for i in range(1, max(len(antennas_data_dict['t_main']), len(antennas_data_dict['t']))):
                writer.writerow([antennas_data_dict['t_main'][i], antennas_data_dict['u_main'][i], antennas_data_dict['u_filt_main'][i],
                                 antennas_data_dict['t'][i], antennas_data_dict['u'][i], antennas_data_dict['u_filt'][i]])
        '''

#write_2_antennas_osc_csv(210622, ['001', '002'])


def find_zeros(time, voltage):
    time, voltage = np.asarray(time), np.asarray(voltage)
    voltage_signs = np.sign(voltage)
    l = 0
    for i in range(len(voltage_signs) - 1):
        if l < 1:
            if voltage_signs[i] != voltage_signs[i + 1] and voltage_signs[i + 1] > 0:
                start_ind = i
                l += 1
    zero_inds = [start_ind]
    m = 0
    for j in range(start_ind + 1, len(voltage_signs) - 1):
        if m < 2:
            if voltage_signs[j] != voltage_signs[j + 1] and voltage_signs[j + 1] != 0:
                zero_inds.append(j)
                m += 1
    print(zero_inds)
    try:
        time_part_max, volt_part_max = time[zero_inds[0]:zero_inds[1]], voltage[zero_inds[0]:zero_inds[1]]
        time_part_min, volt_part_min = time[zero_inds[1]:zero_inds[2]], voltage[zero_inds[1]:zero_inds[2]]

        ind_max, ind_min = np.argmax(volt_part_max), np.argmin(volt_part_min)
        time_max, volt_max = time_part_max[ind_max], volt_part_max[ind_max]
        time_min, volt_min = time_part_min[ind_min], volt_part_min[ind_min]
        #plt.plot(time, voltage)
        #plt.plot(time_part_min, volt_part_min)
        #plt.plot(time_part_max, volt_part_max)
        #plt.show()
        max_min_dict = {'time_max': time_max,
                        'time_min': time_min,
                        'volt_max': volt_max,
                        'volt_min': volt_min}
        return max_min_dict
    except:
        pass


def two_antennas_max_table(exp_num, file_nums):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    print(nums_mtrx)
    central_freq = 2.714E9
    pts = np.arange(100, 525, 25)
    compare_pts = [pts[i] * 1E-9 for i in range(len(pts))]
    col_nums = [i for i in range(3, 3 + len(compare_pts))]
    print(col_nums)
    vals_for_mean = []

    ex_table = excel.Workbook()
    ex_table.create_sheet(title='Integral', index=0)
    sheet = ex_table['Integral']
    sheet['A1'] = exp_num
    sheet['C1'] = 'Без фильтра'
    sheet['A2'] = 'No (центр/бок)'
    sheet['B2'] = 'n, отн.ед.'
    sheet[f'{get_column_letter(3 + len(compare_pts) + 1)}1'] = 'Фильтрованный'

    for n in range(len(col_nums)):
        letter = get_column_letter(col_nums[n])
        sheet[f'{letter}2'] = f't_{n + 1} = {np.round(compare_pts[n] / 1e-9, 0)}'

        filt_letter = get_column_letter(col_nums[n] + len(col_nums) + 1)
        sheet[f'{filt_letter}2'] = f't_{n + 1} = {np.round(compare_pts[n] / 1e-9, 0)} c'

    for k, pt in enumerate(compare_pts):
        row_ind = 0
        for j in range(nums_mtrx.shape[0]):
            antennas_data_dict = {}
            for i in range(nums_mtrx.shape[1]):
                print(nums_mtrx[j, i], 'pt=', pt)
                file_num = f'{nums_mtrx[j, i]}'
                file_name = f'str{file_num}.csv'
                data = test.open_file(file_name, reduced=False)
                filt_freq_min, filt_freq_max = central_freq - 30e6, central_freq + 30e6
                if int(nums_mtrx[j, i]) % 2 == 0:
                    t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    pl_density = test.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
                    comment = test.read_excel(file_name)['dicts'][file_num]['Комментарий']
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    delta_main = volt_max - volt_min

                    u_filt = u_filt[ind_mask]
                    filt_dict = find_zeros(t, u_filt)
                    time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                    time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                    delta_main_filt = volt_max_filt - volt_min_filt

                    cell = sheet.cell(row=j + 3, column=2)
                    cell.value = f'{pl_density}'
                    cell = sheet.cell(row=1, column=2)
                    cell.value = f'{comment}'

                    plt.plot(t, u)
                    plt.plot(time_min, volt_min,  marker='o')
                    plt.plot(time_max, volt_max,  marker='o')
                else:
                    t_shift = 4.347E-9
                    t_shift = 0
                    t, u = data['time'] + t_shift, (data['voltage'] - np.mean(data['voltage']))
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)

                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    delta_sub = volt_max - volt_min

                    u_filt = u_filt[ind_mask]
                    filt_dict = find_zeros(t, u_filt)
                    time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                    time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                    delta_sub_filt = volt_max_filt - volt_min_filt

                    plt.plot(t, u_filt)
                    plt.plot(time_min_filt, volt_min_filt, marker='o')
                    plt.plot(time_max_filt, volt_max_filt, marker='o')

            if int(file_num) < 156:
                u_relat_filt = delta_main_filt / delta_sub_filt
                u_relat = delta_main / delta_sub
            else:
                u_relat_filt = delta_sub_filt / delta_main_filt
                u_relat = delta_sub / delta_main

            if pt <= 325:
                vals_for_mean.append(u_relat_filt)

            column = col_nums[k]
            column_filt = column + len(compare_pts) + 1

            cell = sheet.cell(row=j + 3, column=column)
            cell.value = f'{np.round(u_relat, 3)}'
            cell = sheet.cell(row=j+3, column=column_filt)
            cell.value = f'{np.round(u_relat_filt, 3)}'

            print(u_relat_filt)
            plt.title(nums_mtrx[j, i])
            #plt.show()
            cell = sheet.cell(row=j + 3, column=1)
            cell.value = f'{int(nums_mtrx[j, i]) -1 } / {nums_mtrx[j, i]}'

    path = test.excel_folder_path / f'{exp_num}_{nums_mtrx[j, i]}_{comment[:6]}_30MHz.xlsx'
    ex_table.save(path)


def two_antennas_max_table_new(exp_num, file_nums, start_time=100, end_time=325, time_step=25):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    print(nums_mtrx)
    central_freq = 2.714E9
    pts = np.arange(start_time, end_time + time_step, time_step)
    compare_pts = [pts[i] * 1E-9 for i in range(len(pts))]
    col_nums = [i for i in range(3, 3 + len(compare_pts))]
    print('Start reading excel for comments')
    excel_inf = test.read_excel()['row_dicts']
    comment = excel_inf[f'{nums_mtrx[0, 0]}']['Комментарий']

    ex_table = excel.Workbook()
    ex_table.create_sheet(title='Integral', index=0)
    sheet = ex_table['Integral']
    sheet['A1'] = exp_num
    sheet['B1'] = comment
    sheet['A2'] = 'Без фильтра'
    sheet['A3'] = 'No (центр/бок)'
    sheet['B3'] = 'n, отн.ед.'
    filt_row = 5 + nums_mtrx.shape[0]
    sheet[f'A{filt_row}'] = 'Фильтрованный'
    sheet[f'A{filt_row + 1}'] = 'No (центр/бок)'
    sheet[f'B{filt_row + 1}'] = 'n, отн.ед.'
    mean_val_letter = get_column_letter(len(pts) + 3)
    std_dev_letter = get_column_letter(len(pts) + 4)
    sheet[f'{mean_val_letter}{filt_row + 1}'] = 'Средн знач'
    sheet[f'{std_dev_letter}{filt_row + 1}'] = 'Станд откл'

    for n in range(len(col_nums)):
        letter = get_column_letter(col_nums[n])
        sheet[f'{letter}3'] = f't_{n + 1} = {np.round(compare_pts[n] / 1e-9, 0)}'
        sheet[f'{letter}{filt_row + 1}'] = f't_{n + 1} = {np.round(compare_pts[n] / 1e-9, 0)} c'
    u_1_vals_list = []
    u_3_vals_list = []
    for k, pt in enumerate(compare_pts):
        row_ind = 0
        for j in range(nums_mtrx.shape[0]):
            antennas_data_dict = {}
            for i in range(nums_mtrx.shape[1]):
                print(nums_mtrx[j, i], 'pt=', pt)
                file_num = f'{nums_mtrx[j, i]}'
                file_name = f'str{file_num}.csv'
                data = test.open_file(file_name, reduced=False)
                filt_freq_min, filt_freq_max = central_freq - 30e6, central_freq + 30e6
                if int(nums_mtrx[j, i]) % 2 == 0:
                    t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    pl_density = excel_inf[file_num]['Ток плазмы, А']
                    comment = excel_inf[file_num]['Комментарий']
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1.3E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    try:
                        peak_dict = find_zeros(t, u)
                        time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                        delta_u_3 = volt_max - volt_min
                    except:
                        delta_u_3 = 0

                    u_filt = u_filt[ind_mask]
                    try:
                        filt_dict = find_zeros(t, u_filt)
                        time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                        time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                        delta_u_3_filt = volt_max_filt - volt_min_filt
                    except:
                        delta_u_3_filt = 0

                    u_3_vals_list.append(delta_u_3_filt)

                    cell = sheet.cell(row=j + 4, column=2)
                    cell.value = f'{pl_density}'

                    cell = sheet.cell(row=j + 2 + filt_row, column=2)
                    cell.value = f'{pl_density}'
                else:
                    t_shift = 0
                    t, u = data['time'] + t_shift, (data['voltage'] - np.mean(data['voltage']))
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    try:
                        peak_dict = find_zeros(t, u)
                        time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                        delta_u_1 = volt_max - volt_min
                    except:
                        delta_u_1 = 0

                    u_filt = u_filt[ind_mask]
                    try:
                        filt_dict = find_zeros(t, u_filt)
                        time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                        time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                        delta_u_1_filt = volt_max_filt - volt_min_filt
                    except:
                        delta_u_1_filt = 0
                    u_1_vals_list.append(delta_u_1_filt)

            cell = sheet.cell(row=j + 4, column=1)
            cell.value = f'{int(nums_mtrx[j, i]) - 1} / {nums_mtrx[j, i]}'

            cell = sheet.cell(row=j + 2 + filt_row, column=1)
            cell.value = f'{int(nums_mtrx[j, i]) - 1} / {nums_mtrx[j, i]}'

            if int(file_num) < 156:
                try:
                    u_relat_filt = delta_u_1_filt / delta_u_3_filt
                    u_relat = delta_u_1 / delta_u_3
                except:
                    u_relat_filt = ' '
                    u_relat = ' '
            else:
                try:
                    u_relat_filt = delta_u_3_filt / delta_u_1_filt
                    u_relat = delta_u_3 / delta_u_1
                except:
                    u_relat_filt = ' '
                    u_relat = ' '

            column = col_nums[k]

            cell = sheet.cell(row=j + 4, column=column)
            if type(u_relat) is not str:
                cell.value = np.round(u_relat, 3)
            else:
                cell.value = ' '
            cell = sheet.cell(row=j + 2 + filt_row, column=column)
            if type(u_relat) is not str:
                cell.value = np.round(u_relat_filt, 3)
            else:
                cell.value = ' '
        #path = test.excel_folder_path / f'{exp_num}_{nums_mtrx[j, i]}_{comment[:2]}_{comment[4:7]}.xlsx'
        path = test.excel_folder_path /'max_table.xlsx'
    for j in range(nums_mtrx.shape[0]):
        vals_list = []
        for i in range(len(pts)):
            num_row = j + 2 + filt_row
            num_column = i + 3
            vals_list.append(float(sheet.cell(row=num_row, column=num_column).value))
        mean_val = np.mean(np.asarray(vals_list))
        tstd_val = tstd(np.asarray(vals_list))
        cell = sheet.cell(row=j + 2 + filt_row, column=len(compare_pts) + 3)
        cell.value = np.round(mean_val, 3)
        cell = sheet.cell(row=j + 2 + filt_row, column=len(compare_pts) + 4)
        cell.value = np.round(tstd_val, 3)

    ex_table.save(path)

#exception_list = [63, 64, 65, 66]
exception_list = []
two_antennas_max_table_new(210304, [f'{i:03d}' for i in range(2, 6) if i not in exception_list])

file_nums = [f'{i:03d}' for i in range(61, 67) if i not in exception_list]


def fnums_list(file_nums):
    exp_num = 210806
    nums_list_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 6), 6))
    for nums_list in nums_list_mtrx:
        two_antennas_max_table_new(exp_num, nums_list)

#fnums_list(file_nums)


def polarization_classification():
    exp_num = 210707
    exception_list = [139, 140, 155]
    nums_list = [f'{i:03d}' for i in range(55, 188) if i not in exception_list]
    test = ProcessSignal(str(exp_num))
    e_x_nums, e_y_nums = [], []
    for num in nums_list:
        file_name = f'str{num}.csv'
        comment = test.read_excel(file_name)['dicts'][num]['Комментарий']
        if 'x' in comment:
            e_x_nums.append(num)
        elif 'y' in comment:
            e_y_nums.append(num)
    polar_dict = {'Ex': e_x_nums,
                  'Ey': e_y_nums}
    return polar_dict


def average_coordinate_plot(exp_num=210707):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    polar_dict = polarization_classification()
    nums_list = [f'{i:03d}' for i in range(55, 189) if i not in exception_list]
    e_y_nums, e_x_nums = polar_dict['Ey'], polar_dict['Ex']
    print('E_x:', e_x_nums, '\n', 'E_y:', e_y_nums)
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    filt_row = 5 + nums_mtrx.shape[0]

    plot_table = excel.Workbook()
    plot_table.create_sheet(title='Plot', index=0)
    table = plot_table['Plot']
    table['A1'] = exp_num
    table['A2'] = 'Ex'
    table['A3'] = 'No'
    table['B3'] = 'x, см'
    table['C3'] = 'Средн знач'
    table['D3'] = 'Станд откл'

    table['F2'] = 'Ey'
    table['F3'] = 'No'
    table['G3'] = 'x, см'
    table['H3'] = 'Средн знач'
    table['I3'] = 'Станд откл'

    docs_folder_path = test.excel_folder_path / '30MHz'
    docs_list = os.listdir(docs_folder_path)
    k_x, k_y = 0, 0
    for doc in docs_list:
        num_in_fname = doc[7:10]
        doc_path = test.excel_folder_path / '30MHz' / doc
        wb = excel.load_workbook(doc_path)
        sheet = wb['Integral']
        coordinate = int(sheet.cell(row=1, column=2).value[3:6]) - 133
        min_row = filt_row + 3
        max_row = sheet.max_row
        max_col = sheet.max_column
        first_nums_row, last_nums_row = sheet.cell(row=filt_row + 3, column=1).value, sheet.cell(max_row, column=1).value
        first_num, last_num = first_nums_row.split('/')[0], last_nums_row.split('/')[-1]
        vals_list = []
        for i in range(min_row, max_row + 1):
            for j in range(3, max_col - 1):
                vals_list.append(sheet.cell(row=i, column=j).value)
        aver_val = np.mean(np.asarray(vals_list))
        dev = tstd(np.asarray(vals_list))
        print('vals_in_cells:', vals_list)
        if num_in_fname in e_x_nums:
            #'No'
            cell = table.cell(row=k_x + 4, column=1)
            cell.value = f'{first_num} - {last_num}'
            #'x, см'
            cell = table.cell(row=k_x + 4, column=2)
            cell.value = coordinate
            #'Средн знач'
            cell = table.cell(row=k_x + 4, column=3)
            cell.value = np.round(aver_val, 3)
            #'Станд откл'
            cell = table.cell(row=k_x + 4, column=4)
            cell.value = np.round(dev, 3)
            k_x = k_x + 1
        if num_in_fname in e_y_nums:
            #'No'
            cell = table.cell(row=k_y + 4, column=6)
            cell.value = f'{first_num} - {last_num}'
            #'x, см'
            cell = table.cell(row=k_y + 4, column=7)
            cell.value = coordinate
            #'Средн знач'
            cell = table.cell(row=k_y + 4, column=8)
            cell.value = np.round(aver_val, 3)
            #'Станд откл'
            cell = table.cell(row=k_y + 4, column=9)
            cell.value = np.round(dev, 3)
            k_y = k_y + 1
    plot_table_path = test.excel_folder_path / f'{exp_num}_Ex_Ey_plot_210728.xlsx'
    plot_table.save(plot_table_path)


#average_coordinate_plot()


def e_abs_int(t, u):
    d_time = np.abs(t[1] - t[0])
    u_abs = np.abs(u)
    e_abs_integral = np.trapz(u_abs, x=t, dx=d_time) / 225E-9
    return e_abs_integral


def two_antennas_int_abs_table(exp_num, file_nums, filt=True):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    polar_dict = polarization_classification()
    e_y_nums, e_x_nums = polar_dict['Ey'], polar_dict['Ex']
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    print(nums_mtrx)
    central_freq = 2.714E9
    comment = test.read_excel(f'str{nums_mtrx[0, 0]}.csv')['dicts'][f'{nums_mtrx[0, 0]}']['Комментарий']

    plot_table = excel.Workbook()
    plot_table.create_sheet(title='Integral', index=0)
    table = plot_table['Integral']
    table['A1'] = exp_num
    table['A2'] = 'Ex'
    if filt:
        b2_val = 'Фильтрованные'
    else:
        b2_val = 'Без фильтра'
    table['B2'] = b2_val
    table['A3'] = 'No'
    table['B3'] = 'x, см'
    table['C3'] = 'int(|U1|)/int(|U3|)'
    table['D3'] = 'x, см'
    table['E3'] = 'Средн знач'
    table['F3'] = 'Станд откл'

    table['H2'] = 'Ey'
    table['I2'] = b2_val
    table['H3'] = 'No'
    table['I3'] = 'x, см'
    table['J3'] = 'int(|U1|)/int(|U3|)'
    table['K3'] = 'x, см'
    table['L3'] = 'Средн знач'
    table['M3'] = 'Станд откл'
    k_x, k_y = 0, 0
    for j in range(nums_mtrx.shape[0]):
        for i in range(nums_mtrx.shape[1]):
            file_num = f'{nums_mtrx[j, i]}'
            file_name = f'str{file_num}.csv'
            data = test.open_file(file_name, reduced=True, red_type=225)
            filt_freq_min, filt_freq_max = central_freq - 30e6, central_freq + 30e6
            if int(nums_mtrx[j, i]) % 2 == 0:
                t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                comment = test.read_excel(file_name)['dicts'][file_num]['Комментарий']
                if file_num in e_y_nums:
                    cell = table.cell(row=k_y + 4, column=9)
                    cell.value = int(comment[3:6]) - 133
                    cell = table.cell(row=k_y + 4, column=8)
                    cell.value = f'{int(nums_mtrx[j, i]) - 1} / {nums_mtrx[j, i]}'
                elif file_num in e_x_nums:
                    cell = table.cell(row=k_x + 4, column=2)
                    cell.value = int(comment[3:6]) - 133
                    cell = table.cell(row=k_x + 4, column=1)
                    cell.value = f'{int(nums_mtrx[j, i]) - 1} / {nums_mtrx[j, i]}'
                if filt:
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    time_inds = t <= 325e-9
                    abs_3 = np.abs(u_filt)[time_inds]
                    t_3 = t[time_inds]
                    abs_int_3 = np.trapz(abs_3, t_3)
                else:
                    time_inds = t <= 325e-9
                    abs_3 = np.abs(u)[time_inds]
                    t_3 = t[time_inds]
                    abs_int_3 = np.trapz(abs_3, t_3)
            else:
                t_shift = 0
                t, u = data['time'] + t_shift, (data['voltage'] - np.mean(data['voltage']))
                if filt:
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    time_inds = t <= 325e-9
                    abs_1 = np.abs(u_filt)[time_inds]
                    t_1 = t[time_inds]
                    abs_int_1 = np.trapz(abs_1, t_1)
                else:
                    time_inds = t <= 325e-9
                    abs_1 = np.abs(u)[time_inds]
                    t_1 = t[time_inds]
                    abs_int_1 = np.trapz(abs_1, t_1)

        if file_num in e_x_nums:
            row = 4 + k_x
            column = 3
        else:
            row = 4 + k_y
            column = 10

        if int(file_num) < 156:
            if filt:
                abs_int_relat = abs_int_1 / abs_int_3
                cell = table.cell(row=row, column=column)
                cell.value = np.round(abs_int_relat, 3)
                if file_num in e_x_nums:
                    k_x = k_x + 1
                else:
                    k_y = k_y + 1
            else:
                abs_int_relat = abs_int_1 / abs_int_3
                cell = table.cell(row=row, column=column)
                cell.value = np.round(abs_int_relat, 3)
                if file_num in e_x_nums:
                    k_x = k_x + 1
                else:
                    k_y = k_y + 1
        else:
            if filt:
                abs_int_relat = abs_int_3 / abs_int_1
                cell = table.cell(row=row, column=column)
                cell.value = np.round(abs_int_relat, 3)
                if file_num in e_x_nums:
                    k_x = k_x + 1
                else:
                    k_y = k_y + 1
            else:
                abs_int_relat = abs_int_3 / abs_int_1
                cell = table.cell(row=row, column=column)
                cell.value = np.round(abs_int_relat, 3)
                if file_num in e_x_nums:
                    k_x = k_x + 1
                else:
                    k_y = k_y + 1

    max_row = table.max_row
    start_coord = table.cell(row=4, column=2)
    k_x = 4
    k = 0
    vals_list = []
    for i in range(4, max_row):
        coord = table.cell(row=i, column=2).value
        if coord == start_coord or len(vals_list) < 3:
            print(f'{coord}')
            if k == 0:
                table.cell(row=k_x, column=4).value = coord
            vals_list.append(table.cell(row=i, column=3).value)
            k = k + 1
        else:
            av_value, stand_dev = np.mean(np.asarray(vals_list)), tstd(vals_list)
            table.cell(row=k_x, column=5).value = np.round(av_value, 3)
            table.cell(row=k_x, column=6).value = np.round(stand_dev, 3)
            print(f'vals_list = {vals_list}')
            k = 0
            vals_list = [table.cell(row=i, column=3).value]
            start_coord = table.cell(row=i, column=2).value
            print(f'i={i}, new_start={start_coord}')
            k_x = k_x + 1
    if filt:
        path = test.excel_folder_path / f'{exp_num}_integrals_filt.xlsx'
    else:
        path = test.excel_folder_path / f'{exp_num}_integrals.xlsx'
    plot_table.save(path)

exception_list = [139, 140, 155]
#exception_list = []
#two_antennas_int_abs_table(210707, [f'{i:03d}' for i in range(55, 188) if i not in exception_list], filt=False)

file_nums = [f'{i:03d}' for i in range(153, 158) if i not in exception_list]


def fnums_list_ints(file_nums):
    exp_num = 210707
    nums_list_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 4), 4))
    for nums_list in nums_list_mtrx:
        two_antennas_int_abs_table(exp_num, nums_list)

#fnums_list_ints(file_nums)

def two_antennas_integral_table(exp_num, file_nums):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    polar_dict = polarization_classification()
    nums_list = [f'{i:03d}' for i in range(55, 189) if i not in exception_list]
    e_y_nums, e_x_nums = polar_dict['Ex'], polar_dict['Ey']
    central_freq = 2.714E9

    ex_table = excel.Workbook()
    ex_table.create_sheet(title='Integral', index=0)
    sheet = ex_table['Integral']
    sheet['A1'] = exp_num
    sheet['B1'] = 'E_y'
    sheet['A2'] = 'No (центр/опорн)'
    sheet['B2'] = 'n, отн.ед.'
    sheet['C2'] = 'Координата центра'
    sheet['D2'] = 'W'
    sheet['E2'] = 'W_опорн'

    sheet['G1'] = 'E_x'
    sheet['G2'] = 'No (центр/опорн)'
    sheet['H2'] = 'n, отн.ед.'
    sheet['I2'] = 'Координата центра'
    sheet['J2'] = 'W'
    sheet['K2'] = 'W_опорн'


    for j in range(nums_mtrx.shape[0]):
        for i in range(nums_mtrx.shape[1]):
            file_num = f'{nums_mtrx[j, i]}'
            file_name = f'str{file_num}.csv'
            data = test.open_file(file_name, reduced=True, red_type=325)
            filt_freq_min, filt_freq_max = central_freq - 30e6, central_freq + 30e6

            t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
            u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
            pl_density = test.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
            comment = test.read_excel(file_name)['dicts'][file_num]['Комментарий']
            full_int = np.round(e_abs_int(t, u_filt) / 1e-8, 3)
            if 'y' in comment:
                cell = sheet.cell(row=j + 3, column=1)
                cell.value = f'{nums_mtrx[j, i]}'
                cell = sheet.cell(row=j + 3, column=2)
                cell.value = f'{pl_density}'
                cell = sheet.cell(row=j + 3, column=3)
                cell.value = int(comment[3:6]) - 133
                if int(nums_mtrx[j, i]) < 156:
                    if int(nums_mtrx[j, i]) % 2 == 0:
                        cell = sheet.cell(row=j + 3, column=5)
                        cell.value = f'{full_int}'
                    else:
                        cell = sheet.cell(row=j + 3, column=4)
                        cell.value = f'{full_int}'
                else:
                    if int(nums_mtrx[j, i]) % 2 == 0:
                        cell = sheet.cell(row=j + 3, column=4)
                        cell.value = f'{full_int}'
                    else:
                        cell = sheet.cell(row=j + 3, column=5)
                        cell.value = f'{full_int}'
                row_y = row_y + 1
            if 'x' in comment:
                cell = sheet.cell(row=j + 3, column=7)
                cell.value = f'{nums_mtrx[j, i]}'
                cell = sheet.cell(row=j + 3, column=8)
                cell.value = f'{pl_density}'
                cell = sheet.cell(row=j + 3, column=9)
                cell.value = int(comment[3:6]) - 133
                if int(nums_mtrx[j, i]) < 156:
                    if int(nums_mtrx[j, i]) % 2 == 0:
                        cell = sheet.cell(row=j + 3, column=11)
                        cell.value = f'{full_int}'
                    else:
                        cell = sheet.cell(row=j + 3, column=10)
                        cell.value = f'{full_int}'
                else:
                    if int(nums_mtrx[j, i]) % 2 == 0:
                        cell = sheet.cell(row=j + 3, column=10)
                        cell.value = f'{full_int}'
                    else:
                        cell = sheet.cell(row=j + 3, column=11)
                        cell.value = f'{full_int}'
                row_x = row_x + 1
    path = test.excel_folder_path / f'E_x_E_y_ integrals_210707(2).xlsx'
    ex_table.save(path)


exception_list = [139, 140, 155]
#exception_list = []
#two_antennas_integral_table(210707, [f'{i:03d}' for i in range(55, 188) if i not in exception_list])

file_nums = [f'{i:03d}' for i in range(153, 158) if i not in exception_list]


def two_antennas_ampl_relate(exp_num, file_nums, pts):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    print(nums_mtrx)
    central_freq = 2.714E9
    compare_pts = [pts[i] * 1E-9for i in range(len(pts))]
    for pt in compare_pts:
        for j in range(nums_mtrx.shape[0]):
            for i in range(nums_mtrx.shape[1]):
                print(nums_mtrx[j, i], 'pt=', np.round(pt/1e-9, 1))
                file_num = f'{nums_mtrx[j, i]}'
                file_name = f'str{file_num}.csv'
                data = test.open_file(file_name, reduced=False)
                filt_freq_min, filt_freq_max = central_freq - 15e6, central_freq + 15e6
                if int(nums_mtrx[j, i]) % 2 == 0:
                    t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                    #plt.plot(t, u)
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    plt.plot(t, u_filt)
                    #pl_density = test.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    delta_main = volt_max - volt_min
                    print('volt_max =', volt_max, 'time_max', time_max)

                    u_filt = u_filt[ind_mask]
                    filt_dict = find_zeros(t, u_filt)
                    time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                    time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                    delta_main_filt = volt_max_filt - volt_min_filt

                    #plt.plot(t, u)
                    #plt.plot(time_min, volt_min,  marker='o')
                    #plt.plot(time_max, volt_max,  marker='o')
                else:
                    #t, u = data['time'] + 4.347E-9, (data['voltage'] - np.mean(data['voltage']))
                    t, u = data['time'] + 32E-12, (data['voltage'] - np.mean(data['voltage']))
                    #plt.plot(t, u)
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    plt.plot(t, u_filt)
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    delta_sub = volt_max - volt_min
                    print('volt_max =', volt_max, 'time_max', time_max)

                    u_filt = u_filt[ind_mask]
                    filt_dict = find_zeros(t, u_filt)
                    time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                    time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                    delta_sub_filt = volt_max_filt - volt_min_filt

                    #plt.plot(t, u_filt)
                    #plt.plot(time_min_filt, volt_min_filt, marker='o')
                    #plt.plot(time_max_filt, volt_max_filt, marker='o')
            plt.show()
            plt.close()
            u_relat_filt = delta_sub_filt / delta_main_filt
            u_relat = delta_sub / delta_main
            print('Отн_без_фильтра:', np.round(u_relat, 3))
            print('Oтн_фильтр:', np.round(u_relat_filt, 3))


#two_antennas_ampl_relate(210622, ['001', '002'], [100, 200, 300, 400])


def two_antennas_time_max_dif(exp_num, file_nums, pts):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    print(nums_mtrx)
    central_freq = 2.714E9
    compare_pts = [pts[i] * 1E-9for i in range(len(pts))]
    for pt in compare_pts:
        for j in range(nums_mtrx.shape[0]):
            for i in range(nums_mtrx.shape[1]):
                print(nums_mtrx[j, i], 'pt=', np.round(pt/1e-9, 1))
                file_num = f'{nums_mtrx[j, i]}'
                file_name = f'str{file_num}.csv'
                data = test.open_file(file_name, reduced=False)
                filt_freq_min, filt_freq_max = central_freq - 15e6, central_freq + 15e6
                if int(nums_mtrx[j, i]) % 2 == 0:
                    t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                    #plt.plot(t, u)
                    #plt.show()
                    #plt.close()
                    #pl_density = test.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    time_max_ch_1 = time_max
                    print('volt_max =', volt_max, 'time_max', np.round(time_max / 1e-9, 3))

                    plt.plot(t, u)
                    plt.plot(time_max, volt_max,  marker='o')
                else:
                    t, u = data['time'], (data['voltage'] - np.mean(data['voltage']))
                    #plt.plot(t, u)
                    #plt.show()
                    #plt.close()
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    time_max_ch_2 = time_max
                    print('volt_max =', volt_max, 'time_max', np.round(time_max / 1e-9, 3))

                    plt.plot(t, u)
                    plt.plot(time_max, volt_max, marker='o')

                    #plt.plot(t, u_filt)
                    #plt.plot(time_min_filt, volt_min_filt, marker='o')
                    #plt.plot(time_max_filt, volt_max_filt, marker='o')
            plt.show()
            u_dif = time_max_ch_2 - time_max_ch_1
            print('raw_u_dif', u_dif / 1e-12)
            print('time_dif =', np.round(u_dif / 1e-12))

#file_list = [f'{i:03d}' for i in range(31, 33)]
#two_antennas_time_max_dif(210530, file_list, [700, 700.4])


def read_excel_average_vals(exp_num, exclusios_list):
    excel_folder_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\Excel'.format(exp_num))
    excel_path = excel_folder_path / f'Ampl_{exp_num}_2_antennas.xlsx'
    wb = excel.load_workbook(excel_path)
    sheet = wb['Integral']

    max_row = sheet.max_row
    max_col = sheet.max_column

    for c in range(1, max_col + 1):
        cell_name = str(sheet.cell(row=1, column=c).value)
        if 'Средн' in cell_name:
            av_column = get_column_letter(c)
        elif 'откл' in cell_name:
            dev_column = get_column_letter(c)
    n_vals = np.asarray([float(sheet[f'B{i}'].value) for i in range(3, max_row) if i not in exclusios_list])
    average_vals = np.asarray([float(sheet[f'{av_column}{i}'].value) for i in range(3, max_row) if i not in exclusios_list])
    stand_dev_vals = np.asarray([float(sheet[f'{dev_column}{i}'].value) for i in range(3, max_row) if i not in exclusios_list])
    sorted_inds = np.argsort(n_vals)

    n_vals = n_vals[sorted_inds]
    average_vals = average_vals[sorted_inds]
    stand_dev_vals = stand_dev_vals[sorted_inds]

    new_table = excel.Workbook()
    new_table.create_sheet(title='Average_vals', index=0)
    sheet = new_table['Average_vals']
    sheet['A1'] = 'n, отн.ед.'
    sheet['B1'] = 'Среднее'
    sheet['C1'] = 'Станд откл'

    for z in range(len(n_vals)):
        cell = sheet.cell(row=z + 2, column=1)
        cell.value = n_vals[z]
        cell = sheet.cell(row=z + 2, column=2)
        cell.value = average_vals[z]
        cell = sheet.cell(row=z + 2, column=3)
        cell.value = stand_dev_vals[z]

    new_table_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\Excel\average_vals_210622(1).xlsx'.format(exp_num))
    new_table.save(new_table_path)

#read_excel_average_vals(210607, [])


def file_nums_oscillogramms_10_ns(exp_num, file_nums, start_time, density_vals=False, filt=False):
    print(file_nums)
    start_time = start_time * 1e-9
    end_time = start_time + 10e-9
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    if density_vals:
        density_vals = pl_densities_from_excel(exp_num, file_nums)
        density_mtrx = np.reshape(np.asarray(density_vals), (int(len(density_vals) / 2), 2))
    for j in range(nums_mtrx.shape[0]):
        fig = plt.figure(num=1, dpi=200)
        for i in range(nums_mtrx.shape[1]):
            file_num = f'{nums_mtrx[j, i]}'
            file_name = f'str{file_num}.csv'
            data = test.open_file(file_name, reduced=False)
            if int(file_num) % 2 == 0:
                t, u = data['time'], (data['voltage'] - np.mean(data['voltage'])) / 3
                '''
                filt_freq_min, filt_freq_max = 2.714e9 - 15e6, 2.714e9 + 15e6
                u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                '''
                time_inds = np.logical_and(t >= start_time, t <= end_time)
            else:
                time_shift = 4.35 * 1e-9
                t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                '''
                filt_freq_min, filt_freq_max = 2.714e9 - 15e6, 2.714e9 + 15e6
                u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                '''
                time_inds = np.logical_and(t >= start_time - time_shift, t <= end_time - time_shift)
            time_part_centr, u_part_centr = t[time_inds], u[time_inds]
            ax = fig.add_subplot(111)
            print(f'Creating a picture {file_num}...')
            if int(file_num) % 2 == 0:
                line1, = ax.plot(time_part_centr / 1e-9, u_part_centr, linewidth=1.2, color='red')
                line1.set_label('Сигнал на оси')
                #min_y, max_y = min(u_part_centr) - 0.25, max(u_part_centr) + 0.25
            else:
                line1, = ax.plot((time_part_centr + time_shift) / 1e-9, u_part_centr, linewidth=1.2)
                line1.set_label('Сигнал сбоку')
                min_y, max_y = min(u_part_centr) - 0.25, max(u_part_centr) + 0.25
        ax.set_xlabel(r'$Время, нс$', fontsize=14, fontweight='black')
        ax.set_ylabel(r'$Напряжение, В$', fontsize=14, fontweight='black')
        ax.grid(which='both', axis='both')
        ax.set_ylim(bottom=min_y, top=max_y)
        ax.legend()
        plt.xticks(fontsize=12)
        plt.yticks(fontsize=12)
        if density_vals:
            ax.set_title(f'№ {nums_mtrx[j, 0]} - {nums_mtrx[j, 1]} (магнетрон), n={density_mtrx[j, i]}, {int(start_time / 1e-9)} - {(start_time /1e-9) + 10}ns')
        else:
            ax.set_title(f'№ {nums_mtrx[j, 0]} - {nums_mtrx[j, 1]}, {int(start_time / 1e-9)} - {(start_time / 1e-9) + 10} ns')
        png_name = test.signal_pics_path / f'{exp_num}_{nums_mtrx[j, 0]}_{nums_mtrx[j, 1]}_магнетрон_{int(start_time / 1e-9)}ns.png'
        fig.savefig(png_name)
        plt.close(fig)


#file_nums_oscillogramms_10_ns(210423, [f'{i:03d}' for i in range(165, 167)], start_time=120, density_vals=True)


def rename_osc_files(exp_num):
    proc = ProcessSignal(f'{exp_num}')
    all_files_list = os.listdir(proc.signal_pics_path)
    osc_files_list = [all_files_list[i] for i in range(len(all_files_list)) if 'png' in all_files_list[i] and 'all' not in all_files_list[i]]
    print('files:', osc_files_list)
    density_vals = pl_densities_from_excel(exp_num)
    for i, file in enumerate(osc_files_list):
        new_file_name = file.replace(f'{file}', f'{file[4:7]}_{density_vals[i]}.png')
        old_file_path = proc.signal_pics_path / file
        new_file_path = proc.signal_pics_path / new_file_name
        os.rename(old_file_path, new_file_path)

#rename_osc_files(210414)


