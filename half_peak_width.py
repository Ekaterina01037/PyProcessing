from ProcessClass_10 import ProcessSignal
import openpyxl as xl
import numpy as np
from scipy.stats import linregress
import matplotlib.pyplot as plt


def peak_width(exp_num):
    exp = ProcessSignal(str(exp_num))
    types = exp.read_type_file()
    csv_signals = types['signal_files']
    csv_signal_nums = types['signal_nums']
    excel_results = exp.read_excel(csv_signal_nums)['numbers']
    m_nums = excel_results['magnetron']
    #m_nums = [excel_results['magnetron'][i] for i in range(len(excel_results['magnetron'])) if int(excel_results['magnetron'][i]) >= 125]
    nums = np.zeros(len(m_nums))
    widths = np.zeros(len(m_nums))
    np_s = np.zeros(len(m_nums))
    k = 0
    for i, csv_signal in enumerate(csv_signals):
        signal_num = csv_signal[3:6]
        if signal_num in m_nums:
            file = exp.open_file(csv_signal, reduced=True)
            use_t = file['time']
            use_u = file['voltage']
            dt = file['time_resolution']
            fft_results = exp.fft_amplitude(use_t, use_u)
            freqs = fft_results['frequency']
            amps = fft_results['amplitude']

            np_s[k] = exp.read_excel(csv_signal_nums)['dicts'][signal_num]['Ток плазмы, А']
            nums[k] = signal_num

            half_peak = max(amps) / 2
            mean_freq = freqs[np.argmax(amps)]

            #half_peak_points
            inds = np.arange(0, len(amps))
            r_bool_inds = freqs >= mean_freq
            r_inds = inds[r_bool_inds]
            w = 0
            while amps[r_inds[w]] >= half_peak:
                w += 1
            under_hp_r_ind = r_inds[w]
            over_hp_r_ind = r_inds[w-1]
            under_hp_r_amp, under_hp_r_freq = amps[under_hp_r_ind], freqs[under_hp_r_ind]
            over_hp_r_amp, over_hp_r_freq = amps[over_hp_r_ind], freqs[over_hp_r_ind]
            print('under_ind:', under_hp_r_ind)
            # right_part
            x_r = np.array([over_hp_r_freq, under_hp_r_freq])
            y_r = np.array([over_hp_r_amp, under_hp_r_amp])
            k_r, b_r, r_value_r, p_value_r, std_err_r = linregress(x_r, y_r)
            right_x_interp = np.linspace(over_hp_r_freq, under_hp_r_freq, num=200)
            right_interp = k_r * right_x_interp + b_r
            right_inds = right_interp <= half_peak
            linreg_right_freq = right_x_interp[right_inds][0]

            #left_part
            l_bool_inds = freqs <= mean_freq
            l_inds = inds[l_bool_inds][::-1]
            t = 0
            while amps[l_inds[t]] >= half_peak:
                t += 1
            under_hp_l_ind = l_inds[t]
            over_hp_l_ind = l_inds[t - 1]
            under_hp_l_amp, under_hp_l_freq = amps[under_hp_l_ind], freqs[under_hp_l_ind]
            over_hp_l_amp, over_hp_l_freq = amps[over_hp_l_ind], freqs[over_hp_l_ind]

            x_l = np.array([over_hp_l_freq, under_hp_l_freq])
            y_l = np.array([over_hp_l_amp, under_hp_l_amp])
            k_l, b_l, l_value_l, p_value_l, std_err_l = linregress(x_l, y_l)
            left_x_interp = np.linspace(under_hp_l_freq, over_hp_l_freq, num=200)
            left_interp = k_l * left_x_interp + b_l
            left_inds = left_interp <= half_peak
            linreg_left_freq = left_x_interp[left_inds][-1]
            '''
            plt.plot(freqs, amps)
            plt.plot(under_hp_l_freq, under_hp_l_amp, color='red', marker='o')
            plt.plot(over_hp_l_freq, over_hp_l_amp, color='green', marker='o')
            plt.plot(linreg_left_freq, half_peak, color='blue', marker='o')
            plt.show()
            '''

            half_peak_width = np.round(((linreg_right_freq - linreg_left_freq) / 1e6), 3)
            widths[k] = half_peak_width
            k += 1
            '''
            # high and bottom of spectrum
            h_inds = amps > half_peak
            b_inds = amps < half_peak
            b_freqs = freqs[b_inds]
            b_amps = amps[b_inds]
            right_b_inds = b_freqs > mean_freq
            left_b_inds = b_freqs < mean_freq
            print('left_inds', left_b_inds)
            left_h_amp, right_h_amp = amps[h_inds][0], amps[h_inds][-1]
            left_h_freq, right_h_freq = freqs[h_inds][0], freqs[h_inds][-1]
            left_b_amp = b_amps[left_b_inds][-1]
            left_b_freq = b_freqs[left_b_inds][-1]
            right_b_amp = b_amps[right_b_inds][0]
            right_b_freq = b_freqs[right_b_inds][0]
            #plt.plot(freq, amp)
            #plt.plot(b_freqs, b_amps, color='red')
            #plt.show()
            #print(mean_freq)
            #left_part
            x_l = np.array([left_b_freq, left_h_freq])
            y_l = np.array([left_b_amp, left_h_amp])
            k_l, b_l, r_value_l, p_value_l, std_err_l = linregress(x_l, y_l)
            left_x_interp = np.linspace(left_b_freq, left_h_freq, num=200)
            left_interp = k_l * left_x_interp + b_l
            left_inds = left_interp <= half_peak
            linreg_left_freq = left_x_interp[left_inds][-1]
            #right_part
            x_r = np.array([right_b_freq, right_h_freq])
            y_r = np.array([right_b_amp, right_h_amp])
            k_r, b_r, r_value_r, p_value_r, std_err_r = linregress(x_r, y_r)
            right_x_interp = np.linspace(right_h_freq, right_b_freq, num=200)
            right_interp = k_r * right_x_interp + b_r
            right_inds = right_interp <= half_peak
            linreg_right_freq = right_x_interp[right_inds][0]
            plt.plot(freqs, amps)
            plt.plot(right_x_interp[right_inds][0], right_interp[right_inds][0], marker='o', color='red')
            plt.plot(left_x_interp[left_inds][-1], left_interp[left_inds][-1], marker='o', color='green')
            plt.show()
            #result)
            half_peak_width = np.round(((linreg_right_freq - linreg_left_freq) / 1e6), 3)
            widths[k] = half_peak_width
            k += 1
            '''

    sorted_inds = np.argsort(np_s)
    np_s = np_s[sorted_inds]
    widths = widths[sorted_inds]
    nums = nums[sorted_inds]
    ex_table = xl.Workbook()
    ex_table.create_sheet(title='half_peak_plot', index=0)
    sheet = ex_table['half_peak_plot']
    sheet['A1'] = 'Номер файла'
    sheet['B1'] = 'Плотность плазмы, отн.ед.'
    sheet['C1'] = 'Ширина пика на полувысоте, МГц'
    for i in range(np_s.size):
        cell = sheet.cell(row=i + 2, column=1)
        cell.value = nums[i]
        cell = sheet.cell(row=i + 2, column=2)
        cell.value = np_s[i]
        cell = sheet.cell(row=i + 2, column=3)
        cell.value = widths[i]
    path = r'C:\Users\d_Nice\Documents\SignalProcessing\2020\{}\Excel\half_peak_width_{}.xlsx'.format(exp_num, exp_num)
    ex_table.save(path)


peak_width(210225)