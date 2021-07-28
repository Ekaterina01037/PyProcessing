import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal
import os
from pathlib import Path
import numpy as np
import matplotlib.gridspec as gridspec
import csv
import openpyxl as excel
from openpyxl.utils import get_column_letter


def write_2_antennas_osc_csv(exp_num, file_nums):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    central_freq = 2.714E9
    for j in range(nums_mtrx.shape[0]):
        antennas_data_dict = {}
        for i in range(nums_mtrx.shape[1]):
            file_num = f'{nums_mtrx[j, i]}'
            file_name = f'str{file_num}.csv'
            data = test.open_file(file_name, reduced=False)
            filt_freq_min, filt_freq_max = central_freq - 15e6, central_freq + 15e6
            if int(nums_mtrx[j, i]) % 2 == 0:
                t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                antennas_data_dict['t_main'], antennas_data_dict['u_main'] = t, u
                antennas_data_dict['u_filt_main'] = u_filt
                plt.plot(t, u)
            else:
                #t, u = data['time'], (data['voltage'] - np.mean(data['voltage'])) / 3.16
                t, u = data['time'], (data['voltage'] - np.mean(data['voltage']))
                u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                antennas_data_dict['t'], antennas_data_dict['u'] = t, u
                antennas_data_dict['u_filt'] = u_filt
                plt.plot(t + 4.347E-9, u)

        plt.show()
        print(f'Writing {nums_mtrx[j, i]} csv file...')
        print(antennas_data_dict.keys())

        dif_file_path = test.exp_file_path / '2_antennas_CSV'
        dif_file_path.mkdir(parents=True, exist_ok=True)
        file = open(str(dif_file_path / f'{nums_mtrx[j, 0]}_{nums_mtrx[j, 1]}.csv'), 'w', newline='')
        with file:
            writer = csv.writer(file)
            writer.writerow(['t_main(s)', 'V_main(V)', 'V_main_filt(V)', 't(s)', 'V(V)', 'V_filt(V)'])
            for i in range(1, max(len(antennas_data_dict['t_main']), len(antennas_data_dict['t']))):
                writer.writerow([antennas_data_dict['t_main'][i], antennas_data_dict['u_main'][i], antennas_data_dict['u_filt_main'][i],
                                 antennas_data_dict['t'][i], antennas_data_dict['u'][i], antennas_data_dict['u_filt'][i]])

#write_2_antennas_osc_csv(210423, [165, 166])


def find_zeros(time, voltage):
    time, voltage = np.asarray(time), np.asarray(voltage)
    voltage_signs = np.sign(voltage)
    l = 0
    for i in range(len(voltage_signs) - 1):
        if l < 1:
            if voltage_signs[i] != voltage_signs[i + 1] and voltage_signs[i + 1] > 0:
                start_ind = i
                l += 1
    zero_inds = [start_ind]
    m = 0
    for j in range(start_ind + 1, len(voltage_signs) - 1):
        if m < 2:
            if voltage_signs[j] != voltage_signs[j + 1] and voltage_signs[j + 1] != 0:
                zero_inds.append(j)
                m += 1
    print(zero_inds)
    try:
        time_part_max, volt_part_max = time[zero_inds[0]:zero_inds[1]], voltage[zero_inds[0]:zero_inds[1]]
        time_part_min, volt_part_min = time[zero_inds[1]:zero_inds[2]], voltage[zero_inds[1]:zero_inds[2]]

        ind_max, ind_min = np.argmax(volt_part_max), np.argmin(volt_part_min)
        time_max, volt_max = time_part_max[ind_max], volt_part_max[ind_max]
        time_min, volt_min = time_part_min[ind_min], volt_part_min[ind_min]
        #plt.plot(time, voltage)
        #plt.plot(time_part_min, volt_part_min)
        #plt.plot(time_part_max, volt_part_max)
        #plt.show()
        max_min_dict = {'time_max': time_max,
                        'time_min': time_min,
                        'volt_max': volt_max,
                        'volt_min': volt_min}
        return max_min_dict
    except:
        pass



def signal_envelope(t, u, time_frame=1e-8):
    dt = t[1] - t[0]
    num_of_pts = int(time_frame / dt)
    times = t[::num_of_pts]
    time_step = np.abs(times[0] - times[1])
    t_env = []
    u_env = []
    for time in times:
        time_lim = time + time_step
        ind_t = np.logical_and(t > time, t <= time_lim)
        t_frame = t[ind_t]
        u_frame = u[ind_t]
        ind_max = np.argmax(u_frame)
        t_env.append(t_frame[ind_max])
        u_env.append(u_frame[ind_max])
    envelope_dict = {'env_time': t_env,
                     'env_voltage': u_env}
    return envelope_dict


def envelope_part(t, u, max_part):
    env = signal_envelope(t, u)
    env_t = np.array(env['env_time'])
    env_u = np.array(env['env_voltage'])
    max_env_u = np.max(env_u)
    ind_lim = env_u > max_part * max_env_u
    t_lim = env_t[ind_lim]
    ind_stop = len(t_lim-1)
    l = 0
    for i in range(len(t_lim)-1):
        if l < 1:
            if t_lim[i+1] - t_lim[i] > 90e-9:
                ind_stop = i
                l += 1
    try:
        t_bond = t_lim[:ind_stop:]
        u_zeros = np.zeros(t_bond.size)
        t_min, t_max = t_bond[0], t_bond[-1]
        ind_use = np.logical_and(t >= t_min, t <= t_max)
        t_use = t[ind_use]
        u_use = u[ind_use]
        plt.plot(t, u)
        plt.plot(env_t, env_u)
        plt.plot(t_use, u_use)
        plt.plot(t_bond, u_zeros, marker='o', color='red', linestyle='')
        plt.show()
        signal_part_dict = {'signal_time': t_use,
                            'signal_voltage': u_use}
        return signal_part_dict
    except:
        pass


def two_antennas_max_table(exp_num, file_nums):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    print(nums_mtrx)
    central_freq = 2.714E9
    pts = [100, 200, 300, 400]
    compare_pts = [pts[i] * 1E-9 for i in range(len(pts))]
    col_nums = [i for i in range(3, 3 + len(compare_pts))]
    print(col_nums)

    ex_table = excel.Workbook()
    ex_table.create_sheet(title='Integral', index=0)
    sheet = ex_table['Integral']
    sheet['A1'] = exp_num
    sheet['A2'] = 'No (центр/бок)'
    sheet['B2'] = 'n, отн.ед.'
    sheet['C1'] = 'Без фильтра'
    sheet[f'{get_column_letter(3 + len(compare_pts) +1)}1'] = 'Фильтрованный'
    for n in range(len(col_nums)):
        letter = get_column_letter(col_nums[n])
        sheet[f'{letter}2'] = f't_{n + 1} = {np.round(compare_pts[n] / 1e-9, 0)}'

        filt_letter = get_column_letter(col_nums[n] + len(col_nums) + 1)
        sheet[f'{filt_letter}2'] = f't_{n + 1} = {np.round(compare_pts[n] / 1e-9, 0)} c'

    for k, pt in enumerate(compare_pts):
        for j in range(nums_mtrx.shape[0]):
            antennas_data_dict = {}
            for i in range(nums_mtrx.shape[1]):
                print(nums_mtrx[j, i], 'pt=', pt)
                file_num = f'{nums_mtrx[j, i]}'
                file_name = f'str{file_num}.csv'
                data = test.open_file(file_name, reduced=False)
                filt_freq_min, filt_freq_max = central_freq - 15e6, central_freq + 15e6
                if int(nums_mtrx[j, i]) % 2 == 0:
                    t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    pl_density = test.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    delta_main = volt_max - volt_min

                    u_filt = u_filt[ind_mask]
                    filt_dict = find_zeros(t, u_filt)
                    time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                    time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                    delta_main_filt = volt_max_filt - volt_min_filt

                    cell = sheet.cell(row=j + 3, column=2)
                    cell.value = f'{pl_density}'

                    plt.plot(t, u)
                    plt.plot(time_min, volt_min,  marker='o')
                    plt.plot(time_max, volt_max,  marker='o')
                else:
                    t_shift = 4.347E-9
                    t_shift = 0
                    t, u = data['time'] + t_shift, (data['voltage'] - np.mean(data['voltage']))
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)

                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    delta_sub = volt_max - volt_min

                    u_filt = u_filt[ind_mask]
                    filt_dict = find_zeros(t, u_filt)
                    time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                    time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                    delta_sub_filt = volt_max_filt - volt_min_filt

                    plt.plot(t, u_filt)
                    plt.plot(time_min_filt, volt_min_filt, marker='o')
                    plt.plot(time_max_filt, volt_max_filt, marker='o')
            u_relat_filt = delta_main_filt / delta_sub_filt
            u_relat = delta_main / delta_sub

            column = col_nums[k]
            column_filt = column + len(compare_pts) + 1

            cell = sheet.cell(row=j + 3, column=column)
            cell.value = f'{np.round(u_relat, 3)}'
            cell = sheet.cell(row=j + 3, column=column_filt)
            cell.value = f'{np.round(u_relat_filt, 3)}'

            print(u_relat_filt)
            plt.title(nums_mtrx[j, i])
            #plt.show()
            cell = sheet.cell(row=j + 3, column=1)
            cell.value = f'{int(nums_mtrx[j, i]) -1 } / {nums_mtrx[j, i]}'

    path = test.excel_folder_path / f'Ampl_{exp_num}_2_antennas_123_129.xlsx'
    ex_table.save(path)


exception_list = [55, 56]
#two_antennas_max_table(210607, [f'{i:03d}' for i in range(123, 129) if i not in exception_list])


def two_antennas_ampl_relate(exp_num, file_nums, pts):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    print(nums_mtrx)
    central_freq = 2.714E9
    compare_pts = [pts[i] * 1E-9for i in range(len(pts))]
    for pt in compare_pts:
        for j in range(nums_mtrx.shape[0]):
            for i in range(nums_mtrx.shape[1]):
                print(nums_mtrx[j, i], 'pt=', np.round(pt/1e-9, 1))
                file_num = f'{nums_mtrx[j, i]}'
                file_name = f'str{file_num}.csv'
                data = test.open_file(file_name, reduced=False)
                filt_freq_min, filt_freq_max = central_freq - 15e6, central_freq + 15e6
                if int(nums_mtrx[j, i]) % 2 == 0:
                    t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                    plt.plot(t, u)
                    plt.show()
                    plt.close()
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    #pl_density = test.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    delta_main = volt_max - volt_min
                    print('volt_max =', volt_max, 'time_max', time_max)

                    u_filt = u_filt[ind_mask]
                    filt_dict = find_zeros(t, u_filt)
                    time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                    time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                    delta_main_filt = volt_max_filt - volt_min_filt

                    plt.plot(t, u)
                    plt.plot(time_min, volt_min,  marker='o')
                    plt.plot(time_max, volt_max,  marker='o')
                else:
                    #t, u = data['time'] + 4.347E-9, (data['voltage'] - np.mean(data['voltage']))
                    t, u = data['time'], (data['voltage'] - np.mean(data['voltage']))
                    plt.plot(t, u)
                    plt.show()
                    plt.close()
                    u_filt = test.fft_filter(t, u, filt_freq_min, filt_freq_max)
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    delta_sub = volt_max - volt_min
                    print('volt_max =', volt_max, 'time_max', time_max)

                    u_filt = u_filt[ind_mask]
                    filt_dict = find_zeros(t, u_filt)
                    time_max_filt, volt_max_filt = filt_dict['time_max'], filt_dict['volt_max']
                    time_min_filt, volt_min_filt = filt_dict['time_min'], filt_dict['volt_min']
                    delta_sub_filt = volt_max_filt - volt_min_filt

                    #plt.plot(t, u_filt)
                    #plt.plot(time_min_filt, volt_min_filt, marker='o')
                    #plt.plot(time_max_filt, volt_max_filt, marker='o')
            u_relat_filt = delta_main_filt / delta_sub_filt
            u_relat = delta_main / delta_sub
            print('Отн_без_фильтра:', np.round(u_relat, 3))
            print('Oтн_фильтр:', np.round(u_relat_filt, 3))


two_antennas_ampl_relate(210607, ['117', '118'], [100, 200, 300, 400])


def two_antennas_time_max_dif(exp_num, file_nums, pts):
    print(f'Experiment {exp_num}')
    test = ProcessSignal(str(exp_num))
    nums_mtrx = np.reshape(np.asarray(file_nums), (int(len(file_nums) / 2), 2))
    print(nums_mtrx)
    central_freq = 2.714E9
    compare_pts = [pts[i] * 1E-9for i in range(len(pts))]
    for pt in compare_pts:
        for j in range(nums_mtrx.shape[0]):
            for i in range(nums_mtrx.shape[1]):
                print(nums_mtrx[j, i], 'pt=', np.round(pt/1e-9, 1))
                file_num = f'{nums_mtrx[j, i]}'
                file_name = f'str{file_num}.csv'
                data = test.open_file(file_name, reduced=False)
                filt_freq_min, filt_freq_max = central_freq - 15e6, central_freq + 15e6
                if int(nums_mtrx[j, i]) % 2 == 0:
                    t, u = data['time'], data['voltage'] - np.mean(data['voltage'])
                    #plt.plot(t, u)
                    #plt.show()
                    #plt.close()
                    #pl_density = test.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    time_max_ch_1 = time_max
                    print('volt_max =', volt_max, 'time_max', np.round(time_max / 1e-9, 3))

                    plt.plot(t, u)
                    plt.plot(time_max, volt_max,  marker='o')
                else:
                    t, u = data['time'], (data['voltage'] - np.mean(data['voltage']))
                    #plt.plot(t, u)
                    #plt.show()
                    #plt.close()
                    ind_mask = np.logical_and(t >= pt, t <= pt + 1E-9)
                    t, u = t[ind_mask], u[ind_mask]
                    peak_dict = find_zeros(t, u)
                    time_max, volt_max, time_min, volt_min = peak_dict['time_max'], peak_dict['volt_max'], peak_dict['time_min'], peak_dict['volt_min']
                    time_max_ch_2 = time_max
                    print('volt_max =', volt_max, 'time_max', np.round(time_max / 1e-9, 3))

                    plt.plot(t, u)
                    plt.plot(time_max, volt_max, marker='o')

                    #plt.plot(t, u_filt)
                    #plt.plot(time_min_filt, volt_min_filt, marker='o')
                    #plt.plot(time_max_filt, volt_max_filt, marker='o')
            plt.show()
            u_dif = time_max_ch_2 - time_max_ch_1
            print('raw_u_dif', u_dif / 1e-12)
            print('time_dif =', np.round(u_dif / 1e-12))

#file_list = [f'{i:03d}' for i in range(31, 33)]
#two_antennas_time_max_dif(210530, file_list, [700, 700.4])


def read_excel_average_vals(exp_num, exclusios_list):
    excel_folder_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\Excel'.format(exp_num))
    excel_path = excel_folder_path / f'Ampl_{exp_num}_2_antennas(1).xlsx'
    wb = excel.load_workbook(excel_path)
    sheet = wb['Integral']

    max_row = sheet.max_row
    max_col = sheet.max_column

    for c in range(1, max_col + 1):
        cell_name = str(sheet.cell(row=1, column=c).value)
        if 'Средн' in cell_name:
            av_column = get_column_letter(c)
        elif 'откл' in cell_name:
            dev_column = get_column_letter(c)
    n_vals = np.asarray([float(sheet[f'B{i}'].value) for i in range(3, max_row) if i not in exclusios_list])
    average_vals = np.asarray([float(sheet[f'{av_column}{i}'].value) for i in range(3, max_row) if i not in exclusios_list])
    stand_dev_vals = np.asarray([float(sheet[f'{dev_column}{i}'].value) for i in range(3, max_row) if i not in exclusios_list])
    sorted_inds = np.argsort(n_vals)

    n_vals = n_vals[sorted_inds]
    average_vals = average_vals[sorted_inds]
    stand_dev_vals = stand_dev_vals[sorted_inds]

    new_table = excel.Workbook()
    new_table.create_sheet(title='Average_vals', index=0)
    sheet = new_table['Average_vals']
    sheet['A1'] = 'n, отн.ед.'
    sheet['B1'] = 'Среднее'
    sheet['C1'] = 'Станд откл'

    for z in range(len(n_vals)):
        cell = sheet.cell(row=z + 2, column=1)
        cell.value = n_vals[z]
        cell = sheet.cell(row=z + 2, column=2)
        cell.value = average_vals[z]
        cell = sheet.cell(row=z + 2, column=3)
        cell.value = stand_dev_vals[z]

    new_table_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\Excel\average_vals_210607(1).xlsx'.format(exp_num))
    new_table.save(new_table_path)

#read_excel_average_vals(210607, [])


def signal_envelope(t, u, dt, time_frame=1e-8):
    num_of_pts = int(time_frame / dt)
    times = t[::num_of_pts]
    time_step = np.abs(times[0] - times[1])
    t_env = []
    u_env = []
    for time in times:
        time_lim = time + time_step
        ind_t = np.logical_and(t > time, t <= time_lim)
        t_frame = t[ind_t]
        u_frame = u[ind_t]
        ind_max = np.argmax(u_frame)
        t_env.append(t_frame[ind_max])
        u_env.append(u_frame[ind_max])
    envelope_dict = {'env_time': t_env,
                     'env_voltage': u_env}
    return envelope_dict


def envelope_max_part(self, t, u, dt, max_part):
    env = self.signal_envelope(t, u, dt)
    env_t = np.array(env['env_time'])
    env_u = np.array(env['env_voltage'])
    max_env_u = np.max(env_u)
    ind_lim = env_u > max_part * max_env_u
    t_lim = env_t[ind_lim]
    ind_stop = len(t_lim - 1)
    l = 0
    for i in range(len(t_lim) - 1):
        if l < 1:
            if t_lim[i + 1] - t_lim[i] > 90e-9:
                ind_stop = i
                l += 1
    try:
        t_bond = t_lim[:ind_stop:]
        u_zeros = np.zeros(t_bond.size)
        t_min, t_max = t_bond[0], t_bond[-1]
        ind_use = np.logical_and(t >= t_min, t <= t_max)
        t_use = t[ind_use]
        u_use = u[ind_use]
        signal_part_dict = {'signal_time': t_use,
                            'signal_voltage': u_use}
        return signal_part_dict
    except:
        pass


