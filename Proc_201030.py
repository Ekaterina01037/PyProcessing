import numpy as np
import openpyxl as excel
import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal

proc = ProcessSignal('201113')
csv_types = proc.read_type_file()
csv_signals = csv_types['signal_files']
csv_signal_nums = csv_types['signal_nums']
excel_dicts = proc.read_excel(csv_signal_nums)['numbers']
noise_nums_1 = excel_dicts['noise'][:12:]
noise_nums_2 = excel_dicts['noise'][12:28:]
magnetron_nums = excel_dicts['magnetron']

print('Noise nums_1:', noise_nums_1)
print('Noise nums_2:', noise_nums_2)
print('Magsnetron nums:', magnetron_nums)
nums, n_s, full_ints, peak_ints, peak_max_s = [], [], [], [], []
noise_ints_1, noise_ns_1, n_nums_1 = [], [], []
noise_ints_2, noise_ns_2, n_nums_2 = [], [], []
for signal in csv_signals:
    num = signal[3:6]
    if num in magnetron_nums:
        file = proc.open_file(signal, reduced=True)
        u = file['voltage']
        t = file['time']
        u_filt = proc.fft_filter(t, u, 2.695e9, 2.725e9)
        dt = file['time_resolution']
        dt_2 = (t[-1] - t[0]) ** 2

        pl_density = proc.read_excel(signal)['dicts'][num]['Ток плазмы, А']

        integral = np.round(proc.e_square(t, u) / 1e-8, 3)
        filt_int = np.round(proc.e_square(t, u_filt) / 1e-8, 3)
        freqs, amps = proc.fft_amplitude(t, u)['frequency'], proc.fft_amplitude(t, u)['amplitude']
        peak_inds = np.logical_and(freqs >= 2.695e9, freqs <= 2.725e9)
        p_freqs, p_amps = freqs[peak_inds], amps[peak_inds]
        max_peak = np.round(np.max(amps), 2)

        full_int = np.round(dt_2 * proc.e_square(freqs, amps) / 2e-8, 3)
        peak_int = np.round(dt_2 * proc.e_square(p_freqs, p_amps) / 2e-8, 3)

        peak_max_s.append(max_peak)
        n_s.append(pl_density)
        nums.append(num)
        full_ints.append(full_int)
        peak_ints.append(peak_int)

        print('peak_int = ', np.round(integral, 2), 'noise_int =', np.round(integral - filt_int, 2),
              'noise_fft =', np.round(full_int - peak_int, 2))
    if num in noise_nums_1 or num in noise_nums_2:
        file = proc.open_file(signal, reduced=True)
        u = file['voltage']
        t = file['time']
        dt = file['time_resolution']
        dt_2 = (t[-1] - t[0]) ** 2

        pl_density = proc.read_excel(signal)['dicts'][num]['Ток плазмы, А']

        integral = np.round(proc.e_square(t, u) / 1e-8, 3)
        freqs, amps = proc.fft_amplitude(t, u)['frequency'], proc.fft_amplitude(t, u)['amplitude']
        noise_int = np.round(dt_2 * proc.e_square(freqs, amps) / 2e-8, 3)
        if num in noise_nums_1:
            noise_ints_1.append(noise_int)
            noise_ns_1.append(pl_density)
            n_nums_1.append(num)
        else:
            noise_ints_2.append(noise_int)
            noise_ns_2.append(pl_density)
            n_nums_2.append(num)
        
ind_sort = np.argsort(np.asarray(n_s))
ns = np.asarray(n_s)[ind_sort]
peak_max_s = np.asarray(peak_max_s)[ind_sort]
nums = np.asarray(nums)[ind_sort]
full_ints, peak_ints = np.asarray(full_ints)[ind_sort], np.asarray(peak_ints)[ind_sort]

n_ind_sort_1 = np.argsort(np.asarray(noise_ns_1))
n_ind_sort_2 = np.argsort(np.asarray(noise_ns_2))
noise_ns_1, n_nums_1, noise_ints_1 = np.asarray(noise_ns_1)[n_ind_sort_1], np.asarray(n_nums_1)[n_ind_sort_1], np.asarray(noise_ints_1)[n_ind_sort_1]
noise_ns_2, n_nums_2, noise_ints_2 = np.asarray(noise_ns_2)[n_ind_sort_2], np.asarray(n_nums_2)[n_ind_sort_2], np.asarray(noise_ints_2)[n_ind_sort_2]
ex_table = excel.Workbook()
ex_table.create_sheet(title='Integral', index=0)
sheet = ex_table['Integral']
sheet['A1'] = 'Номер(шум)'
sheet['B1'] = 'Плотность плазмы, отн.ед.'
sheet['C1'] = 'W_2, *10-8'

sheet['E1'] = 'Номер(шум, поглотитель 3 см)'
sheet['F1'] = 'Плотность плазмы, отн.ед.'
sheet['G1'] = 'W_2, *10-8'

sheet['I1'] = 'Номер(поглотитель 3 см)'
sheet['J1'] = 'Плотность плазмы, отн.ед.'
sheet['K1'] = 'Полный интеграл, *10-8'
sheet['L1'] = 'W_f0, *10-8'
sheet['M1'] = 'W_1, *10-8'
sheet['N1'] = 'Пик'

for z in range(noise_ints_1.size):
    cell = sheet.cell(row=z + 2, column=1)
    cell.value = int(n_nums_1[z])
    cell = sheet.cell(row=z + 2, column=2)
    cell.value = noise_ns_1[z]
    cell = sheet.cell(row=z + 2, column=3)
    cell.value = noise_ints_1[z]

for w in range(noise_ints_2.size):
    cell = sheet.cell(row=w + 2, column=5)
    cell.value = int(n_nums_2[w])
    cell = sheet.cell(row=w + 2, column=6)
    cell.value = noise_ns_2[w]
    cell = sheet.cell(row=w + 2, column=7)
    cell.value = noise_ints_2[w]

for k in range(full_ints.size):
    cell = sheet.cell(row=k + 2, column=9)
    cell.value = int(nums[k])
    cell = sheet.cell(row=k + 2, column=10)
    cell.value = ns[k]
    cell = sheet.cell(row=k + 2, column=11)
    cell.value = full_ints[k]
    cell = sheet.cell(row=k + 2, column=12)
    cell.value = peak_ints[k]
    cell = sheet.cell(row=k + 2, column=13)
    cell.value = full_ints[k] - peak_ints[k]
    cell = sheet.cell(row=k + 2, column=14)
    cell.value = peak_max_s[k]

path = r'C:\Users\d_Nice\Documents\SignalProcessing\2020\201113\Excel\Integrals_201113.xlsx'
ex_table.save(path)



