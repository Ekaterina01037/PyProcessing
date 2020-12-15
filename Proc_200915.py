import numpy as np
import openpyxl as excel
import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal

test = ProcessSignal('200924')
test.files_classification()
csv_types = test.read_type_file()
csv_signals = csv_types['signal_files']
csv_signal_nums = csv_types['signal_nums']
excel_dicts = test.read_excel(csv_signal_nums)['numbers']
magnetron_nums = excel_dicts['magnetron']
plasma_nums = excel_dicts['noise']
print('Magnetron nums are:', magnetron_nums,
      'REB nums are:', plasma_nums)
pl_densities = np.zeros(len(csv_signals))
integrals = np.zeros(len(csv_signals))
nums = np.zeros(len(csv_signals))
test.reduce_fft()
for i, csv_signal in enumerate(csv_signals):
    csv_signal_num = csv_signal[3:6]
    file = test.open_file(csv_signal, reduced=True)
    pl_density = test.read_excel(csv_signal)['dicts'][csv_signal_num]['Ток плазмы, А']
    pl_densities[i] = pl_density
    time = file['time']
    voltage = file['voltage']
    t, u, dt = file['time'], file['voltage'], file['time_resolution']
    fft = test.fft_amplitude(t, u, dt)
    freqs = fft['frequency']
    amps = fft['amplitude']
    n_dt_2 = (t[-1] - t[0]) ** 2
    full_integral = 2 * n_dt_2 * test.e_square(freqs, amps) / 1e-8
    integrals[i] = full_integral
    nums[i] = csv_signal_num

ex_table = excel.Workbook()
ex_table.create_sheet(title='Integral', index=0)
sheet = ex_table['Integral']
sheet['A1'] = 'Номер'
sheet['B1'] = 'Плотность плазмы, отн.ед.'
sheet['C1'] = 'Интеграл, *10-8'
table_size = integrals.size
for k in range(table_size):
    cell = sheet.cell(row=k+2, column=1)
    cell.value = nums[k]
    cell = sheet.cell(row=k+2, column=2)
    cell.value = pl_densities[k]
    cell = sheet.cell(row=k+2, column=3)
    cell.value = integrals[k]
path = test.excel_folder_path / 'Integral.xlsx'
ex_table.save(path)
