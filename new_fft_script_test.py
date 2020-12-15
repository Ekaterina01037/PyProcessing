from ProcessClass_10 import ProcessSignal
import openpyxl as xl
import matplotlib.pyplot as plt

table_path = path / file
wb = xl.load_workbook(table_path)
ws = wb.active
max_row, max_col = ws.max_row,  ws.max_column
last_cell = ws.cell(row=max_row, column=max_col)
last_param_cell = ws.cell(row=1, column=max_col)
file_data = ws['A1':last_cell.coordinate]
param_list = ws['A1':last_param_cell.coordinate]
fun = lambda row: list(map(lambda cell: cell.value, row))
values = np.asarray(list(map(fun, file_data))).T
param_list_data = list(map(fun, param_list))
key_list, vals_list = [], []
for i in range(len(param_list_data[0])):
    param_list_val = param_list_data[0][i]
    if 'Detector' in param_list_val:
        det_num_row = i
        detector_nos = values[det_num_row, 1:].astype(int)
    if 'Sensitivity' in param_list_val:
        sens_coef_row = i
        sens_coefs = values[sens_coef_row, 1:].astype(float)
        key_list.append('sens_coef')
        vals_list.append(sens_coefs)

def fft_reb_noise(exp_num, table=False):
    proc = ProcessSignal(f'{exp_num}')
    types = proc.read_type_file()
    csv_signal_nums = types['signal_nums']
    excel_results = proc.read_excel(csv_signal_nums)['numbers']
    noise_nums = excel_results['noise']
    print(excel_results)

fft_reb_noise(201110)
def part_fft(self, csv_signals, interest_nums,
             part_nums, fft_type='part',
             block_full=False, block_part=True,
             peak=False, noise=False):
    ex_table = xl.Workbook()
    ex_table.create_sheet(title='Full FFT', index=0)
    ex_table.create_sheet(title='Part FFT', index=1)
    sheet1 = ex_table['Full FFT']
    sheet2 = ex_table['Part FFT']
    sheet2['A1'] = 'File Number'
    sheet1['A1'] = 'File Number'
    signal_dict = {}
    row_f = 1
    row_p = 1
    for j, csv_signal in enumerate(csv_signals):
        signal_num = csv_signal[3:6]
        if signal_num in interest_nums or signal_num in part_nums:
            use_signal = self.open_file(csv_signal, reduced=True)
            use_t = use_signal['time']
            use_u = use_signal['voltage']
            dt = use_signal['time_resolution']
            fig = plt.figure(num=1, dpi=300)
            ax = fig.add_subplot(111)
            line = None
            ax.set_prop_cycle(color=['mediumseagreen', 'dodgerblue', 'indigo'])
            if fft_type == 'full' or fft_type == 'both':
                if signal_num in interest_nums:
                    print('{} am in interest_nums'.format(signal_num))
                    if block_full is True:
                        filtered_u = self.fft_filter(use_t, use_u, 2.695e9, 2.725e9, filt_type='bandstop')
                        fft_results = self.fft_amplitude(use_t, filtered_u)
                    else:
                        fft_results = self.fft_amplitude(use_t, use_u)
                    pl_density = self.read_excel(interest_nums)['dicts'][signal_num]['Ток плазмы, А']
                    freq = fft_results['frequency']
                    amp = fft_results['amplitude']

                    if peak:
                        peak_inds = np.logical_and(2.67e9 < freq, freq < 2.730e9)
                        peak_freqs = freq[peak_inds]
                        peak_amps = amp[peak_inds]
                        mean_freq = self.mean_frequency(peak_freqs, peak_amps)
                        # line, = ax.plot(peak_freqs, peak_amps, linewidth=1.2)
                        line, = ax.plot(freq, amp, linewidth=1.2)
                        peak_max = np.round(np.max(peak_amps), 2)
                    else:
                        mean_freq = self.mean_frequency(freq, amp)
                        line, = ax.plot(freq, amp, linewidth=0.7)
                        line, = ax.plot(freq, amp, linewidth=0.7)
                    try:
                        spectrum_mean_freq = mean_freq['mean_freq']
                        line.set_label(r'$f = {} GHz$'.format(spectrum_mean_freq))

                        value_f = str('f, GHz')
                        cell_f = sheet1.cell(row=1, column=2)
                        cell_f.value = value_f

                        cell_pl = sheet1.cell(row=1, column=3)
                        cell_pl.value = str('n, arb.units')

                        row_f = row_f + 1
                        cell_name = 'A{}'.format(row_f)
                        sheet1[str(cell_name)] = '{}'.format(signal_num)
                        value = spectrum_mean_freq
                        cell = sheet1.cell(row=row_f, column=2)
                        cell.value = value

                        value_pl = pl_density
                        cell = sheet1.cell(row=row_f, column=3)
                        cell.value = value_pl
                    except TypeError:
                        pass

            if fft_type == 'part' or fft_type == 'both':
                if signal_num in part_nums:
                    print('{} in part_nums'.format(signal_num))
                    if block_part is True:
                        filtered_u = self.bandstop_filter(use_t, use_u, 2.709e9, 2.769e9)
                        part_signal = self.signal_parts(use_t, filtered_u, dt)
                    else:
                        part_signal = self.signal_parts(use_t, use_u, dt)
                    part_keys = part_signal.keys()
                    part_freq_dict = {}

                    row_p = row_p + 1
                    cell_name = 'A{}'.format(row_p)
                    sheet2[str(cell_name)] = '{}'.format(signal_num)
                    for i, part_key in enumerate(part_keys):
                        k = i + 1
                        col = k + 1
                        part_time = part_signal[part_key]['time']
                        part_voltage = part_signal[part_key]['voltage']
                        fft_results = self.fft_signal(part_time, part_voltage, dt)
                        freq = fft_results['frequency']
                        amp = fft_results['amplitude']
                        line, = ax.plot(freq, amp, linewidth=0.7)

                        mean_freq = self.mean_frequency(freq, amp)
                        part_freq = mean_freq['mean_freq']

                        value_f = str('f{}'.format(k))
                        cell_f = sheet2.cell(row=1, column=col)
                        cell_f.value = value_f

                        value = part_freq
                        cell = sheet2.cell(row=row_p, column=col)
                        cell.value = value
                        part_freq_dict['f{}, GHz'.format(k)] = part_freq
                        line.set_label(r'$f_{}= {} GHz$'.format(k, part_freq))
                    pl_density = self.read_excel(part_nums)['dicts'][signal_num]['Ток плазмы, А']
                    signal_dict[signal_num] = part_freq_dict
            if fft_type == 'part' or fft_type == 'both':
                if signal_num in part_nums:
                    ax.set_title(r'$File\/Number = {} (Part FFT, n={})$'.format(signal_num, pl_density))
            if fft_type == 'full' or fft_type == 'both':
                if signal_num in interest_nums:
                    if noise:
                        ax.set_title(r'$File\/Number = {},\/ (n={}) $'.format(signal_num, pl_density))
                    else:
                        # absorbers = self.read_excel(part_nums)['dicts'][signal_num]['Поглотители в тракте магнетрона']
                        ax.set_title(r'$File\/Number = {},\/ \/ n={} $'.format(signal_num, pl_density))

            if line is not None:
                ax.set_ylim(bottom=0)
                if peak:
                    ax.set_xlim(left=2.68e9, right=2.74e9)
                else:
                    ax.set_xlim(left=0, right=4e9)
                ax.grid(which='both', axis='both')
                ax.set_xlabel(r'$Frequency, GHz$')
                ax.set_ylabel(r'$Amplitude$')
                ax.legend()
                if fft_type == 'part' or fft_type == 'both':
                    if signal_num in part_nums:
                        png_name = self.fft_part_path / signal_num
                if fft_type == 'full' or fft_type == 'both':
                    if signal_num in interest_nums:
                        if noise:
                            png_name = self.fft_reb / 'reb_{}'.format(signal_num)
                            table_path = self.fft_reb / 'fft_reb_{}.xlsx'.format(signal_num)
                        else:
                            if block_full:
                                png_name = self.fft_noise_base / 'noise_base_{}'.format(signal_num)
                                table_path = self.fft_noise_base / 'fft_noise_base_{}.xlsx'.format(signal_num)
                            else:
                                png_name = self.fft_magnetron / 'magnetron_{}'.format(signal_num)
                                table_path = self.fft_magnetron / 'fft_magnetron_{}.xlsx'.format(signal_num)
                        if peak:
                            png_name = self.fft_peak / 'peak_{}_1'.format(signal_num)
                            table_path = self.fft_peak / 'peak_{}.xlsx'.format(signal_num)
                fig.savefig(png_name)
                plt.close(fig)
        print('Circle {} complete'.format(j))
    ex_table.save(table_path)
