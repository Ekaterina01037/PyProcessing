import numpy as np
import openpyxl as excel
import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal

proc = ProcessSignal('201113')
csv_types = proc.read_type_file()
csv_signals = csv_types['signal_files']
csv_signal_nums = csv_types['signal_nums']
excel_dicts = proc.read_excel(csv_signal_nums)['numbers']
noise_nums = excel_dicts['noise'][:12:]
magnetron_nums = excel_dicts['magnetron'][0:0:]
nums, n_s, l_ints, r_ints = [], [], [], []
l_noise_ints, r_noise_ints, noise_ns, n_nums = [], [], [], []
for signal in csv_signals:
    num = signal[3:6]
    file = proc.open_file(signal, reduced=True)
    u = file['voltage']
    t = file['time']
    dt = file['time_resolution']
    dt_2 = (t[-1] - t[0]) ** 2
    fft_data = proc.fft_amplitude(t, u)
    freqs, amps = fft_data['frequency'][1::], fft_data['amplitude'][1::]
    left_inds = freqs <= 2.71e9 - 50e6
    l_freqs = freqs[left_inds]
    l_amps = amps[left_inds]
    l_int = np.round(dt_2 * proc.e_square(l_freqs, l_amps) / 2e-8, 3)
    r_inds = freqs >= 2.71e9 + 50e6
    r_freqs = freqs[r_inds]
    r_amps = amps[r_inds]
    r_int = np.round(dt_2 * proc.e_square(r_freqs, r_amps) / 2e-8, 3)
    '''
    plt.plot(freqs, amps)
    plt.plot(l_freqs, l_amps)
    plt.plot(r_freqs, r_amps)
    plt.title(f'l_int = {l_int}, r_int = {r_int}')
    plt.xlim(right=6e9)
    plt.show()
    '''
    pl_density = proc.read_excel(signal)['dicts'][num]['Ток плазмы, А']
    if num in magnetron_nums:
        nums.append(num)
        n_s.append(pl_density)
        l_ints.append(l_int)
        r_ints.append(r_int)
    if num in noise_nums:
        print(num)
        n_nums.append(num)
        noise_ns.append(pl_density)
        l_noise_ints.append(l_int)
        r_noise_ints.append(r_int)

ind_sort = np.argsort(np.asarray(n_s))
n_s = np.asarray(n_s)[ind_sort]
nums = np.asarray(nums)[ind_sort]
l_ints = np.asarray(l_ints)[ind_sort]
r_ints = np.asarray(r_ints)[ind_sort]

n_ind_sort = np.argsort(np.asarray(noise_ns))
noise_ns = np.asarray(noise_ns)[n_ind_sort]
n_nums = np.asarray(n_nums)[n_ind_sort]
l_noise_ints = np.asarray(l_noise_ints)[n_ind_sort]
r_noise_ints = np.asarray(r_noise_ints)[n_ind_sort]

ex_table = excel.Workbook()
ex_table.create_sheet(title='Left_right', index=0)
sheet = ex_table['Left_right']
sheet['A1'] = 'Номер'
sheet['B1'] = 'Плотность плазмы, отн.ед.'
sheet['C1'] = 'W_1_left, *10-8'
sheet['D1'] = 'W_1_right, *10-8'

sheet['F1'] = 'Номер(шум)'
sheet['G1'] = 'Плотность плазмы, отн.ед.'
sheet['H1'] = 'W_2_left, *10-8'
sheet['I1'] = 'W_2_right, *10-8'

for z in range(nums.size):
    cell = sheet.cell(row=z + 2, column=1)
    cell.value = int(nums[z])
    cell = sheet.cell(row=z + 2, column=2)
    cell.value = n_s[z]
    cell = sheet.cell(row=z + 2, column=3)
    cell.value = l_ints[z]
    cell = sheet.cell(row=z + 2, column=4)
    cell.value = r_ints[z]

for w in range(n_nums.size):
    cell = sheet.cell(row=w + 2, column=6)
    cell.value = int(n_nums[w])
    cell = sheet.cell(row=w + 2, column=7)
    cell.value = noise_ns[w]
    cell = sheet.cell(row=w + 2, column=8)
    cell.value = l_noise_ints[w]
    cell = sheet.cell(row=w + 2, column=9)
    cell.value = r_noise_ints[w]

path = r'C:\Users\d_Nice\Documents\SignalProcessing\2020\201113\Excel\left_right_201113.xlsx'
ex_table.save(path)