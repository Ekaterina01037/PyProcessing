import numpy as np
import openpyxl as excel
import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal


nums, n_s, mean_freqs = [], [], []
noise_freqs, noise_ns, n_nums = [], [], []
for signal in csv_signals:proc = ProcessSignal('201110')
csv_types = proc.read_type_file()
csv_signals = csv_types['signal_files']
csv_signal_nums = csv_types['signal_nums']
excel_dicts = proc.read_excel(csv_signal_nums)['numbers']
noise_nums = excel_dicts['noise']
magnetron_nums = excel_dicts['magnetron'][:18:]
    num = signal[3:6]
    if num in magnetron_nums:
        file = proc.open_file(signal, reduced=True)
        u = file['voltage']
        t = file['time']
        u_filt = proc.fft_filter(t, u, 2.695e9, 2.725e9, filt_type='bandstop')
        dt = file['time_resolution']
        dt_2 = (t[-1] - t[0]) ** 2

        pl_density = proc.read_excel(signal)['dicts'][num]['Ток плазмы, А']

        fft_data = proc.fft_amplitude(t, u_filt)
        freqs, amps = fft_data['frequency'], fft_data['amplitude']
        mean_freq = proc.mean_frequency(freqs, amps)['mean_freq']

        nums.append(num)
        n_s.append(pl_density)
        mean_freqs.append(mean_freq)
    if num in noise_nums:
        file = proc.open_file(signal, reduced=True)
        u = file['voltage']
        t = file['time']
        dt = file['time_resolution']
        dt_2 = (t[-1] - t[0]) ** 2

        pl_density = proc.read_excel(signal)['dicts'][num]['Ток плазмы, А']

        fft_data = proc.fft_amplitude(t, u)
        freqs, amps = fft_data['frequency'], fft_data['amplitude']
        mean_freq = proc.mean_frequency(freqs, amps)['mean_freq']

        noise_nums.append(num)
        noise_ns.append(pl_density)
        noise_freqs.append(mean_freq)

ind_sort = np.argsort(np.asarray(n_s))
ns = np.asarray(n_s)[ind_sort]
nums = np.asarray(nums)[ind_sort]
freqs = np.asarray(mean_freqs)[ind_sort]

n_ind_sort = np.argsort(np.asarray(noise_ns))
noise_ns = np.asarray(noise_ns)[n_ind_sort]
noise_nums = np.asarray(nums)[n_ind_sort]
noise_freqs = np.asarray(freqs)[n_ind_sort]

ex_table = excel.Workbook()
ex_table.create_sheet(title='Mean_frequency', index=0)
sheet = ex_table['Mean_frequency']
sheet['A1'] = 'Номер'
sheet['B1'] = 'Плотность плазмы, отн.ед.'
sheet['C1'] = 'f_ср, ГГц'

sheet['E1'] = 'Номер(шум)'
sheet['F1'] = 'Плотность плазмы, отн.ед.'
sheet['G1'] = 'f_ср, ГГц'

for z in range(nums.size):
    cell = sheet.cell(row=z + 2, column=1)
    cell.value = int(nums[z])
    cell = sheet.cell(row=z + 2, column=2)
    cell.value = n_s[z]
    cell = sheet.cell(row=z + 2, column=3)
    cell.value = mean_freqs[z]

for w in range(noise_nums.size):
    cell = sheet.cell(row=w + 2, column=5)
    cell.value = int(noise_nums[w])
    cell = sheet.cell(row=w + 2, column=6)
    cell.value = noise_ns[w]
    cell = sheet.cell(row=w + 2, column=7)
    cell.value = noise_freqs[w]

path = r'C:\Users\d_Nice\Documents\SignalProcessing\2020\201110\Excel\Mean_freq_201110.xlsx'
ex_table.save(path)

