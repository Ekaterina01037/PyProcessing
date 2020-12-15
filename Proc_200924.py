import numpy as np
import openpyxl as excel
import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal
import os
from pathlib import Path
from scipy.fftpack import rfft, irfft, rfftfreq
'''
test = ProcessSignal('200924')
#test.files_classification()
csv_types = test.read_type_file()
csv_signals = csv_types['signal_files']
csv_signal_nums = csv_types['signal_nums']
pl_densities = np.zeros(len(csv_signals))
integrals = np.zeros(len(csv_signals))
nums = np.zeros(len(csv_signals))
csv_path = test.csv_files_path
csv_dir_files = os.listdir(csv_path)
#test.reduce_fft()
for i, csv_signal in enumerate(csv_signals):
    csv_signal_num = csv_signal[3:6]
    if csv_signal in csv_dir_files:
        file = test.open_file(csv_signal, reduced=True)
        pl_density = test.read_excel(csv_signal)['dicts'][csv_signal_num]['Ток плазмы, А']
        pl_densities[i] = pl_density
        t, u, dt = file['time'], file['voltage'], file['time_resolution']
        plt.plot(t, u)
        integrals[i] = np.round(test.e_square(t, u) / 1e-7, 2)
        print(csv_signal_num, integrals[i])
        plt.title("№ {}, e^2 = {}".format(csv_signal_num, integrals[i]))
        plt.show()
'''
def two_exp_integrals(exp_num, table=False, data=True):
    test = ProcessSignal('{}'.format(exp_num))
    csv_types = test.read_type_file()
    csv_signals = csv_types['signal_files']
    pl_densities = np.zeros(len(csv_signals))
    integrals = np.zeros(len(csv_signals))
    nums = np.zeros(len(csv_signals))
    csv_path = test.csv_files_path
    csv_dir_files = os.listdir(csv_path)
    test.reduce_fft()
    for i, csv_signal in enumerate(csv_signals):
        csv_signal_num = csv_signal[3:6]
        if csv_signal in csv_dir_files:
            file = test.open_file(csv_signal, reduced=True)
            pl_density = test.read_excel(csv_signal)['dicts'][csv_signal_num]['Ток плазмы, А']
            pl_densities[i] = pl_density
            nums[i] = csv_signal_num
            t, u, dt = file['time'], file['voltage'], file['time_resolution']
            plt.plot(t, u)
            integrals[i] = np.round(test.e_square(t, u) / 1e-7, 2)
            print(csv_signal_num, integrals[i])
            plt.title("№ {}, e^2 = {}".format(csv_signal_num, integrals[i]))
            #plt.show()

    sorted_inds = np.argsort(pl_densities)
    pl_densities = pl_densities[sorted_inds]
    integrals = integrals[sorted_inds]
    nums = nums[sorted_inds]
    if table:
        ex_table = excel.Workbook()
        ex_table.create_sheet(title='Integral', index=0)
        sheet = ex_table['Integral']
        sheet['A1'] = 'Номер'
        sheet['B1'] = 'Плотность плазмы, отн.ед.'
        sheet['C1'] = 'Интеграл, *10-8'
        table_size = integrals.size
        for k in range(table_size):
            cell = sheet.cell(row=k+2, column=1)
            cell.value = nums[k]
            cell = sheet.cell(row=k+2, column=2)
            cell.value = pl_densities[k]
            cell = sheet.cell(row=k+2, column=3)
            cell.value = integrals[k]
        path = test.excel_folder_path / 'Integral_{}.xlsx'.format(exp_num)
        ex_table.save(path)

    if data:
        integral_dict = {'pl_ds': pl_densities,
                         'integral_vals': integrals,
                         'shot_nums': nums}
        return integral_dict

#dict_22 = two_exp_integrals(200922)
#dict_24 = two_exp_integrals(200924)

def m_signals_classification(magnetron_in_vals, magnetron_nums):
    test = ProcessSignal('200925')
    m_classif_dict = {}
    for val in magnetron_in_vals:
        file_numbers = []
        for m_num in magnetron_nums:
            magnetron_in = test.read_excel(m_num)['dicts'][m_num]['Входное напряжение магнетрона, В']
            if magnetron_in == val:
                file_numbers.append(m_num)
            else:
                pass
        m_classif_dict[val] = file_numbers
    return m_classif_dict

def absorber_classification(exp_number, magnetron_dicts):
    test = ProcessSignal(str(exp_number))
    types = test.read_type_file()
    csv_signal_nums = types['signal_nums']
    magnetron_nums = test.read_excel(csv_signal_nums)['numbers']['magnetron']
    dicts = test.read_excel(csv_signal_nums)['dicts']
    absorbers_list = []*(len(magnetron_nums))
    for num in magnetron_nums:
        absorbers_list.append(dicts[num]['Поглотители в тракте магнетрона'])
    uniq_abs = set(absorbers_list)
    print(uniq_abs)

    abs_dict = {}
    for element in uniq_abs:
        nums_list = []
        for dict_num in magnetron_dicts:
            absorbers = magnetron_dicts[dict_num]['abs']
            if absorbers == element:
                nums_list.append(magnetron_dicts[dict_num])
        abs_dict[element] = nums_list
    return abs_dict


def intgrals_calc_25():
    #magnetron_in_vals = [8.7, 10.7, 13]
    test = ProcessSignal('201008')
    csv_files_list = os.listdir(test.csv_files_path)
    csv_types = test.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = test.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_dicts['magnetron']
    reb_nums = excel_dicts['noise']
    #m_classif_dict = m_signals_classification(magnetron_in_vals, magnetron_nums)
    magnetron_dict = {}
    reb_dict = {}
    for csv_signal in csv_signals:
        if csv_signal in csv_files_list:
            num = csv_signal[3:6]
            file = test.open_file(csv_signal, reduced=True)
            t, u = file['time'], file['voltage']
            pl_density = test.read_excel(csv_signal)['dicts'][num]['Ток плазмы, А']
            full_integral = np.round(test.e_square(t, u) / 1e-8, 3)
            if num in reb_nums:
                reb_dict[num] = {'n': pl_density,
                                 'full_integral': full_integral,
                                 'num': num}
            else:
                len_t = len(t)
                dt = np.abs(t[1] - t[0])
                fft_u = rfft(u)
                freqs_fft = rfftfreq(len_t, dt)

                ind_mask = np.logical_and(2.695e9 < freqs_fft, freqs_fft < 2.725e9)
                u_empty = np.zeros(len(ind_mask))
                u_empty[ind_mask] = fft_u[ind_mask]
                fft_filtered_u = irfft(u_empty)
                #plt.plot(t, u, color='k')
                #plt.plot(t, fft_filtered_u)
                #plt.show()
                absorbers = test.read_excel(csv_signal)['dicts'][num]['Поглотители в тракте магнетрона']
                filtered_integral = np.round(test.e_square(t, fft_filtered_u) / 1e-8, 3)
                pedestal_integral = full_integral - filtered_integral
                magnetron_dict[num] = {'n': pl_density,
                                       'full_integral': full_integral,
                                       'peak_integral': filtered_integral,
                                       'pedestal_integral': pedestal_integral,
                                       'abs': absorbers,
                                       'num': num}
                plt.plot(pl_density, full_integral, color='blue', marker='o')
                plt.plot(pl_density, filtered_integral, color='red', marker='o', linestyle='-')
    return reb_dict, magnetron_dict

ex_table = excel.Workbook()
reb_dicts, magnetron_dicts = intgrals_calc_25()
print('Calculating integrals...')
nums, reb_integrals, n_s = np.zeros(len(reb_dicts.keys())), np.zeros(len(reb_dicts.keys())), np.zeros(len(reb_dicts.keys()))
for j, key in enumerate(reb_dicts.keys()):
    nums[j] = reb_dicts[key]['num']
    reb_integrals[j] = reb_dicts[key]['full_integral']
    n_s[j] = reb_dicts[key]['n']
ind_mask = np.argsort(n_s)
nums = nums[ind_mask]
reb_integrals = reb_integrals[ind_mask]
n_s = n_s[ind_mask]
print('Creating REB integrals table...')
ex_table.create_sheet(title='Integral', index=0)
sheet = ex_table['Integral']
sheet['A1'] = 'Номер'
sheet['B1'] = 'Плотность плазмы, отн.ед.'
sheet['C1'] = 'Интеграл, *10-8'
reb_table_size = reb_integrals.size
for k in range(reb_table_size):
    cell = sheet.cell(row=k+2, column=1)
    cell.value = nums[k]
    cell = sheet.cell(row=k+2, column=2)
    cell.value = n_s[k]
    cell = sheet.cell(row=k+2, column=3)
    cell.value = reb_integrals[k]
classif_mag_dict = absorber_classification(201008, magnetron_dicts)
print('Sorting magnetron signals...')
i_0 = 5
for key in classif_mag_dict.keys():
    dicts_list = classif_mag_dict[key]
    nums, absorbers, n_s = np.zeros(len(dicts_list)), [0]*(len(dicts_list)), np.zeros(len(dicts_list)),
    full_ints, peak_ints, p_ints = np.zeros(len(dicts_list)), np.zeros(len(dicts_list)), np.zeros(len(dicts_list))
    for z, dict in enumerate(dicts_list):
        nums[z], absorbers[z], n_s[z] = dict['num'], dict['abs'], dict['n']
        full_ints[z], peak_ints[z], p_ints[z] = dict['full_integral'], dict['peak_integral'], dict['pedestal_integral']
    ind_mask_m = np.argsort(n_s)
    nums, absorbers, n_s = nums[ind_mask_m], np.asarray(absorbers)[ind_mask_m], n_s[ind_mask_m]
    full_ints, peak_ints, p_ints = full_ints[ind_mask_m], peak_ints[ind_mask_m], p_ints[ind_mask_m]
    sheet.cell(1, i_0).value = 'Номер'
    sheet.cell(1, i_0+1).value = 'Поглотители'
    sheet.cell(1, i_0+2).value = 'Плотность плазмы, отн.ед.'
    sheet.cell(1, i_0+3).value = 'Полный интеграл, *10-8'
    sheet.cell(1, i_0+4).value = 'Энергия в пике, *10-8'
    sheet.cell(1, i_0+5).value = 'Энергия в шумах, *10-8'
    magnetron_table_size = nums.size
    print(f'Creating {key} absorbers table...')
    for k in range(magnetron_table_size):
        cell = sheet.cell(row=k + 2, column=i_0)
        cell.value = nums[k]
        cell = sheet.cell(row=k + 2, column=i_0 + 1)
        cell.value = absorbers[k]
        cell = sheet.cell(row=k + 2, column=i_0 + 2)
        cell.value = n_s[k]
        cell = sheet.cell(row=k + 2, column=i_0 + 3)
        cell.value = full_ints[k]
        cell = sheet.cell(row=k + 2, column=i_0 + 4)
        cell.value = peak_ints[k]
        cell = sheet.cell(row=k + 2, column=i_0 + 5)
        cell.value = p_ints[k]
    i_0 += 8

path = r'C:\Users\Public\Documents\201008\Excel\Integrals_201008.xlsx'
ex_table.save(path)

