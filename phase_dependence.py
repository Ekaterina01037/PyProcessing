import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal
import os
from pathlib import Path
import numpy as np
import csv
from scipy.stats import linregress
import openpyxl as xl
import matplotlib.gridspec as gridspec
import datetime
from ProcessClass_10 import ProcessSignal

folder_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\210326\CSV')

all_files = os.listdir(folder_path)
files = [all_files[i] for i in range(len(all_files)) if 'csv' in all_files[i]]


def open_file(file_name):
    file_path = folder_path / '{}'.format(file_name)
    t = []
    u = []
    with open(file_path) as File:
        reader = csv.reader(File)
        for row in reader:
            t_val = float(row[3])
            t.append(t_val)
            u_val = float(row[4])
            u.append(u_val)
    t = np.asarray(t)
    u = np.asarray(u)
    file_dict = {'time': t,
                 'voltage': u}
    return file_dict


def cut_files():
    for file in files:
        if '1211' in file:
            file_data = open_file(file)
            t, u = file_data['time'], file_data['voltage']
            plt.plot(t, u)
            plt.show()
            mask_inds = np.logical_and(t >= 120e-9, t <= 570e-9)
            t_cut, u_cut = t[mask_inds], u[mask_inds]
            name_str = file[0:6] + '_cut' + file[6:]
            file_path = folder_path / name_str
            file = open(str(file_path), 'w', newline='')
            with file:
                writer = csv.writer(file)
                for i in range(0, len(t_cut)):
                    writer.writerow([0, 0, 0, t_cut[i], u_cut[i]])

#cut_files()

def signal_envelope(t, u, time_frame=1e-8):
    dt = t[1] - t[0]
    num_of_pts = int(time_frame / dt)
    times = t[::num_of_pts]
    time_step = np.abs(times[0] - times[1])
    t_env = []
    u_env = []
    for time in times:
        time_lim = time + time_step
        ind_t = np.logical_and(t > time, t <= time_lim)
        t_frame = t[ind_t]
        u_frame = u[ind_t]
        ind_max = np.argmax(u_frame)
        t_env.append(t_frame[ind_max])
        u_env.append(u_frame[ind_max])
    envelope_dict = {'env_time': t_env,
                     'env_voltage': u_env}
    return envelope_dict


def enveloped_part(t, u, max_part=0.2):
    env = signal_envelope(t, u)
    env_t = np.array(env['env_time'])
    env_u = np.array(env['env_voltage'])
    max_env_u = np.max(env_u)
    ind_lim = env_u > max_part * max_env_u
    t_lim = env_t[ind_lim]
    ind_stop = len(t_lim-1)
    l = 0
    for i in range(len(t_lim)-1):
        if l < 1:
            if t_lim[i+1] - t_lim[i] > 90e-9:
                ind_stop = i
                l += 1
    try:
        t_bond = t_lim[:ind_stop:]
        u_zeros = np.zeros(t_bond.size)
        t_min, t_max = t_bond[0], t_bond[-1]
        ind_use = np.logical_and(t >= t_min, t <= t_max)
        t_use = t[ind_use]
        u_use = u[ind_use]
        #plt.plot(t, u)
        #plt.plot(env_t, env_u)
        #plt.plot(t_use, u_use)
        #plt.plot(t_bond, u_zeros, marker='o', color='red', linestyle='')
        #plt.show()
        signal_part_dict = {'time': t_use,
                            'voltage': u_use}
        return signal_part_dict
    except:
        pass


def signal_periods(num, time, voltage, table=True):
    '''
    file_data = open_file(file)
    if 'in000' in file:
        time, voltage = file_data['time'], file_data['voltage']
    else:
        time, voltage = file_data['time'], file_data['voltage']
    #plt.plot(time, voltage)
    #plt.show()
    '''
    voltage_signs = np.sign(voltage)
    m = 0
    l = 0
    start_inds = []
    end_inds = []
    for i in range(len(voltage_signs) - 1):
        if l < 1:
            if voltage_signs[i] != voltage_signs[i + 1] and voltage_signs[i + 1] != 0:
                m += 1
                if m % 2 != 0:
                   start_inds.append(i)
                   end_inds.append(i+1)
    start_volt = voltage[start_inds]
    start_time = time[start_inds]
    end_volt = voltage[end_inds]
    end_time = time[end_inds]
    zero_xs = []
    for i in range(len(start_time)):
        x = np.array([start_time[i], end_time[i]])
        y = np.array([start_volt[i], end_volt[i]])
        k, b, r_value, p_value, std_err = linregress(x, y)
        new_y = 0
        new_x = (new_y - b) / k
        zero_xs.append(new_x)
    periods = np.diff(zero_xs)
    zeros = zero_xs[1::]
    '''
    #prelim_view
    plt.plot(time, voltage)
    plt.plot(zeros, np.zeros(len(zeros)), marker='o', linestyle=' ')
    plt.show()
    '''
    if table:
        ex_table = xl.Workbook()
        ex_table.create_sheet(title='Период', index=0)
        sheet = ex_table['Период']
        sheet['A1'] = 'Время'
        sheet['B1'] = 'Период'
        for z in range(periods.size):
            cell = sheet.cell(row=z + 2, column=1)
            cell.value = zeros[z]
            cell = sheet.cell(row=z + 2, column=2)
            cell.value = periods[z]
        path = folder_path / f'{num}.xlsx'
        ex_table.save(path)
    else:
        periods_dict = {'zero_times': zeros, 'periods': periods}
        return periods_dict

#signal_periods()


def phase_pics(full_time_plot=False):
    files = os.listdir(folder_path)
    period_difference = []
    for file in files:
        if 'cut' in file:
            periods = signal_periods(file)
            zero_times, delta_t = np.asarray(periods['zero_times']), periods['periods']
            period_difference.append(delta_t)
    t_bounds = [(120 + 40 * i) * 1e-9 for i in range(12)]
    sign_magnetron_diff = period_difference[1] - period_difference[0][:1205:]
    if full_time_plot:
        plt.plot(np.asarray(zero_times) / 1e-9, sign_magnetron_diff / 1e-9)
        plt.xlim(left=120, right=560)
        plt.ylabel('Разность периодов, нс')
        plt.xlabel('Время, нс')
        path_full_time_diff = folder_path / 'full_time.png'
        plt.savefig(path_full_time_diff)
    else:
        for j in range(11):
            t_small_graphs = [t_bounds[j] + 5e-9*k for k in range(9)]
            gs = gridspec.GridSpec(8, 2)

            main_y_mask = np.logical_and(zero_times >= t_bounds[j], zero_times <= t_bounds[j+1])
            main_y = delta_t[main_y_mask]
            main_x = zero_times[main_y_mask]

            fig = plt.figure(num=1, dpi=300, figsize=[11.69, 12.27])
            ax = fig.add_subplot(gs[0:4, :])
            ax.plot(main_x / 1e-9, main_y / 1e-9, color='darkmagenta')
            ax.set_xlim(left=min(main_x / 1e-9), right=max(main_x / 1e-9))
            ax.set_ylabel('Разность периодов, нс')
            ax.set_xlabel('t, нс')

            file_data = open_file('str039_cut_1211.csv')
            t, u = file_data['time'], file_data['voltage']
            file_data_m = open_file('magnet_cutron_in000.csv')
            t_m, u_m = file_data_m['time'], file_data_m['voltage']
            graph_num = 0
            for k in range(2):
                for l in range(4):
                    ax = fig.add_subplot(gs[4 + l, k])
                    print(graph_num)
                    local_mask = np.logical_and(t >= t_small_graphs[graph_num], t <= t_small_graphs[graph_num+1])
                    x, y = t[local_mask], u[local_mask]
                    local_mask_m = np.logical_and(t_m >= t_small_graphs[graph_num], t_m <= t_small_graphs[graph_num + 1])
                    x_m, y_m = t_m[local_mask_m], u_m[local_mask_m]
                    line_m, = ax.plot(x_m / 1e-9, y_m, color='orange')
                    line, = ax.plot(x / 1e-9, y)
                    ax.set_ylabel('V, В')
                    ax.set_xlabel('t, нс')
                    ax.set_xlim(left=min(min(x), min(x_m)) / 1e-9, right=max(max(x), max(x_m)) / 1e-9)
                    graph_num += 1
            small_graphs_path = folder_path / f'{t_small_graphs[graph_num]}_1218.png'
            fig.savefig(small_graphs_path)
            plt.close(fig)

#phase_pics(full_time_plot=True)


def phase_pisc_by_nums(file_nums):
    magnetron_file = 'str097_m210302.csv'
    m_file_data = open_file(magnetron_file)
    m_t, m_v = m_file_data['time'], m_file_data['voltage'] - np.mean(m_file_data['voltage'])
    m_inds = m_t >= 400e-9
    #m_inds = m_t >= -100e-9
    m_t_plato, m_v_plato = m_t[m_inds], m_v[m_inds]
    for num in file_nums:
        file_name = f'str{num:03d}.csv'
        file_data = open_file(file_name)
        t, v = file_data['time'], file_data['voltage']
        t_env, v_env = t, v
        #env_part_data = enveloped_part(t, v, max_part=0.45)
        #t_env, v_env = env_part_data['time'], env_part_data['voltage'] - np.mean(env_part_data['voltage'])
        m_t_plato, m_v_plato_norm = m_t_plato - (m_t_plato[0] - t_env[0]), m_v_plato * np.max(v_env) / np.max(m_v_plato)
        periods_data = signal_periods(num, t_env, v_env)
        m_periods_data = signal_periods(97, m_t_plato, m_v_plato)
        signal_zeros, m_zeros = periods_data['zero_times'], m_periods_data['zero_times']
        sign_periods = periods_data['periods']
        t_env_shifted = t_env - np.mean(sign_periods)


        #prelim_plots
        plt.plot(m_t_plato, m_v_plato_norm)
        #plt.plot(t, v)
        plt.plot(t_env_shifted, v_env)
        #plt.plot(signal_zeros, np.zeros(len(signal_zeros)), marker='o', linestyle=' ')
        #plt.plot(m_zeros, np.zeros(len(m_zeros)), marker='o', linestyle=' ')
        plt.hlines(y=0, xmin=t_env[0], xmax=t_env[-1])
        plt.title(f'{num}')
        plt.show()


#phase_pisc_by_nums([56, 59, 62, 112, 114, 86, 95, 90])
phase_pisc_by_nums([116])

def average_phase_diff(delay_array=True, pic_type='opposite'):
    today_data = datetime.date.today().strftime("%Y%m%d")[2::]
    files = ['magnet_cutron_in000.csv', 'str039_cut_1211.csv']
    period_difference = []
    first_zeros = []
    u_maxs, t_maxs = [], []
    u_mins, t_mins = [], []
    for file in files:
        u_max_vals, t_max_vals = [], []
        u_min_vals, t_min_vals = [], []
        file_data = open_file(file)
        t, u = file_data['time'], file_data['voltage']
        periods = signal_periods(file)
        zero_times, delta_t = np.asarray(periods['zero_times']), periods['periods']
        first_zeros.append(zero_times[0])
        period_difference.append(delta_t)
        for i in range(len(zero_times)-1):
            period_inds = np.logical_and(t >= zero_times[i], t <= zero_times[i+1])
            t_period, u_period = t[period_inds], u[period_inds]
            ind_max, ind_min = np.argmax(u_period), np.argmin(u_period)
            u_max_vals.append(u_period[ind_max])
            t_max_vals.append(t_period[ind_max])
            u_min_vals.append(u_period[ind_min])
            t_min_vals.append(t_period[ind_min])
        u_maxs.append(u_max_vals)
        t_maxs.append(t_max_vals)
        u_mins.append(u_min_vals)
        t_mins.append(t_min_vals)
    sign_magnetron_diff = period_difference[1] - period_difference[0][:1205:]
    average_diff = np.mean(sign_magnetron_diff)
    first_zero_dif = first_zeros[1] - first_zeros[0]
    average_max_dif = np.mean(np.asarray(t_maxs[1] - first_zero_dif) - np.asarray(t_maxs[0][:1204:]))
    average_min_dif = np.mean(np.asarray(t_mins[1] - first_zero_dif) - np.asarray(t_mins[0][:1204:]))
    average_peak_dif = np.mean([average_min_dif, average_max_dif])

    magnetron_data, signal_data = open_file('magnet_cutron_in000.csv'), open_file('str039_cut_1211.csv')
    m_t, m_u = magnetron_data['time'], magnetron_data['voltage']

    if pic_type is 'in':
        inphase_delay = first_zero_dif + average_max_dif / 2
        t, u = signal_data['time'] - inphase_delay, signal_data['voltage']
        print('Inphase_delay:', np.round(inphase_delay / 1e-9, 3))
        sign_sum_inds = t >= m_t[0]
        t_sum_sign, u_sum_sign = t[sign_sum_inds], u[sign_sum_inds]
        magnetron_sum_inds = m_t <= t[-1]
        t_sum_m, u_sum_m = m_t[magnetron_sum_inds], m_u[magnetron_sum_inds]
        u_same_phase = u_sum_sign + u_sum_m

        plt.plot(t_sum_sign / 1e-9, u_same_phase, color='mediumvioletred')
        left_x, right_x = t_sum_m[0] / 1e-9, t_sum_m[-1] / 1e-9
        title = f'Сигналы совпадают по фазе, задержка {inphase_delay / 1e-9} нс'
        pic_path = folder_path / f'op_phase_full_{today_data}_{inphase_delay / 1e-9}.png'

    #u_zero_lvl = np.mean(signal_data['voltage'])
    elif pic_type is 'opposite':
        if delay_array:
            k = np.arange(0.7, 1.2, 0.05)
            delays = np.asarray([k[i] * first_zero_dif + average_min_dif for i in range(len(k))])
        else:
            delays = 1.234e-9
        for delay in delays:
            t_1, u_1 = signal_data['time'] - delay, signal_data['voltage']
            delay_round = np.round(delay / 1e-9, 3)
            print('Opposite phase delay:', delay_round)
            sign_dif_inds = t_1 >= m_t[0]
            t_dif_sign, u_dif_sign = t_1[sign_dif_inds], u_1[sign_dif_inds]
            magnetron_dif_inds = m_t <= t_1[-1]
            t_dif_m, u_dif_m = m_t[magnetron_dif_inds], m_u[magnetron_dif_inds]
            u_contra_phase = u_dif_m + u_dif_sign

            plt.plot(t_dif_m / 1e-9, u_contra_phase, color='mediumvioletred')
            left_x, right_x = t_dif_m[0] / 1e-9, t_dif_m[-1] / 1e-9
            title = f'Сигналы в противофазе, задержка {delay / 1e-9} нс'
            pic_path = folder_path / f'op_phase_full_{today_data}_{delay / 1e-9}.png'
        '''
        print('dif_delay:', first_zero_dif + average_diff)
        plt.plot(t_dif_sign, u_contra_phase, color='mediumvioletred')
        plt.hlines(y=0, xmin=min(m_t[0], t_1[0]), xmax=max(m_t[-1], t_1[-1]))
        plt.show()
        '''
        #plt.plot(t / 1e-9, u, color='blue')
        #plt.plot(m_t / 1e-9, m_u, color='orange')

        #plt.plot(t_1 / 1e-9, u_1, color='darkblue')
        #plt.plot(t_sum_sign / 1e-9, u_same_phase, color='mediumvioletred')
        plt.ylim(bottom=-6, top=6)
        plt.xlim(left=left_x, right=right_x)
        plt.ylabel('V, В')
        plt.xlabel('t, нс')
        plt.title(title)
        pic_phase_full_path = pic_path
        plt.savefig(pic_phase_full_path)
        plt.close()

#average_phase_diff()