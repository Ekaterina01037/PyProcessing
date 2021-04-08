import numpy as np
import openpyxl as xl
from pathlib import Path
import os
from ProcessClass_10 import ProcessSignal
import matplotlib.pyplot as plt


def e_square_spectra(exp_num, signal_type='magnetron', scale='log'):
    proc = ProcessSignal(f'{exp_num}')
    types = proc.read_type_file()
    csv_signals, csv_signal_nums = types['signal_files'], types['signal_nums']
    excel_results = proc.read_excel(csv_signal_nums)['numbers']
    if signal_type == 'magnetron':
        nums_for_proc = excel_results['magnetron']
    elif signal_type == 'noise':
        nums_for_proc = excel_results['noise']
    for num in nums_for_proc:
        file_name = f'str{num}.csv'
        file_data = proc.open_file(file_name, True)
        t, u = file_data['time'], file_data['voltage']
        fft_data = proc.fft_amplitude(t, u * u)
        freqs, amps = fft_data['frequency'][2::], fft_data['amplitude'][2::]
        inds = freqs <= 4e9
        freqs_4, amps_4 = freqs[inds], amps[inds]
        pl_dens = proc.read_excel(file_name)['dicts'][num]['Ток плазмы, А']

        fig = plt.figure(num=1, dpi=300)
        ax = fig.add_subplot(111)

        line, = ax.plot(freqs_4, amps_4)
        ax.set_xlim(left=1e9, right=4e9)
        if scale == 'log':
            ax.set_yscale('log')
            ax.set_ylim(bottom=10**(-5))
            pic_name = r'Спектр квадрата напряжения\Логарифмический\ u_2_log_{}'.format(num)
        else:
            ax.set_ylim(bottom=0)
            pic_name = r'Спектр квадрата напряжения\Обычный\ u_2_{}'.format(num)
        ax.grid(which='both', axis='both')
        ax.set_title(r'$№={}, n={}$'.format(num, pl_dens))
        png_name = proc.fft_pics_path / pic_name
        fig.savefig(png_name)
        plt.close(fig)
        plt.show()


e_square_spectra(210322, scale='ordinary')


def read_excel_spectra_data(exp_num):
    excel_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\Excel'.format(exp_num))
    noise_base_folder_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\Pictures\FFT\Спектр шумового пъедестала'.format(exp_num))
    reb_noise_folder_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\Pictures\FFT\Спектр шумов РЭП'.format(exp_num))
    path_list = [noise_base_folder_path, reb_noise_folder_path]
    for path in path_list:
        file_list = os.listdir(path)
        for file in file_list:
            if 'xlsx' in file:
                wb = xl.load_workbook(path/file)
                sheet = wb['Full FFT']
                rows = sheet.max_row

                if path == noise_base_folder_path:
                    nums_base = np.asarray([int(sheet[f'A{i}'].value) for i in range(2, rows+1)])
                    n_s_base, freqs_base = np.asarray([float(sheet[f'C{i}'].value) for i in range(2, rows+1)]), np.asarray([float(sheet[f'B{i}'].value) for i in range(2, rows+1)])
                    sorted_inds = np.argsort(n_s_base)
                    nums_base_sorted, n_s_base_sorted, freqs_base_sorted = nums_base[sorted_inds], n_s_base[sorted_inds], freqs_base[sorted_inds]
                else:
                    nums_reb = np.asarray([int(sheet[f'A{i}'].value) for i in range(2, rows+1)])
                    n_s_reb, freqs_reb = np.asarray([float(sheet[f'C{i}'].value) for i in range(2, rows+1)]), np.asarray([float(sheet[f'B{i}'].value) for i in range(2, rows+1)])
                    sorted_inds = np.argsort(n_s_reb)
                    nums_reb_sorted, n_s_reb_sorted, freqs_reb_sorted = nums_reb[sorted_inds], n_s_reb[sorted_inds], freqs_reb[sorted_inds]

    new_table = xl.Workbook()
    new_table.create_sheet(title='Mean_freq_noise', index=0)
    sheet = new_table['Mean_freq_noise']
    sheet['A1'] = 'Номер(шум)'
    sheet['B1'] = 'Плотность плазмы, отн.ед.'
    sheet['C1'] = 'Средняя частота, ГГц'

    sheet['E1'] = 'Номер'
    sheet['F1'] = 'Плотность плазмы, отн.ед.'
    sheet['G1'] = 'Средняя частота, ГГц'
    for z in range(len(nums_reb_sorted)):
        cell = sheet.cell(row=z + 2, column=1)
        cell.value = int(nums_reb_sorted[z])
        cell = sheet.cell(row=z + 2, column=2)
        cell.value = n_s_reb_sorted[z]
        cell = sheet.cell(row=z + 2, column=3)
        cell.value = freqs_reb_sorted[z]

    for v in range(len(nums_base_sorted)):
        cell = sheet.cell(row=v + 2, column=5)
        cell.value = int(nums_base_sorted[v])
        cell = sheet.cell(row=v + 2, column=6)
        cell.value = n_s_base_sorted[v]
        cell = sheet.cell(row=v + 2, column=7)
        cell.value = freqs_base_sorted[v]

    new_table_path = excel_path/f'mean_freq_{exp_num}.xlsx'
    new_table.save(new_table_path)


#read_excel_spectra_data(210322)


def noise_mean_freq_calc(exp_num, central_freq=2.714e9, band_half_width=50e6):
    proc = ProcessSignal(f'{exp_num}')
    types = proc.read_type_file()
    csv_signals, csv_signal_nums = types['signal_files'], types['signal_nums']
    excel_results = proc.read_excel(csv_signal_nums)['numbers']
    noise_nums = excel_results['noise']
    magnetron_nums = excel_results['magnetron']
    nums_for_proc = [noise_nums, magnetron_nums]

    n_nums, n_plasma_dens, n_mean_freqs = [], [], []
    m_nums, m_plasma_dens, m_mean_freqs = [], [], []

    for list in nums_for_proc:
        for file_num in list:
            file_name = f'str{file_num}.csv'
            file_data = proc.open_file(file_name, reduced=True)
            t, u, dt = file_data['time'], file_data['voltage'], file_data['time_resolution']
            pl_density = proc.read_excel(file_name)['dicts'][file_num]['Ток плазмы, А']
            fft_results = proc.fft_amplitude(t, u)
            freqs, amps = fft_results['frequency'], fft_results['amplitude']
            if file_num in noise_nums:
                mean_freq = proc.mean_frequency(freqs, amps)
                spectrum_mean_freq = mean_freq['mean_freq']
                n_nums.append(file_num), n_plasma_dens.append(pl_density), n_mean_freqs.append(spectrum_mean_freq)
            else:
                left_boundary, right_boundary = central_freq - band_half_width, central_freq + band_half_width
                noise_base_inds = np.logical_or(freqs <= left_boundary, freqs >= right_boundary)
                noise_base_freqs, noise_base_amps = freqs[noise_base_inds], amps[noise_base_inds]
                mean_freq = proc.mean_frequency(noise_base_freqs, noise_base_amps)
                spectrum_mean_freq = mean_freq['mean_freq']
                m_nums.append(file_num), m_plasma_dens.append(pl_density), m_mean_freqs.append(spectrum_mean_freq)
    n_sorted_inds = np.argsort(np.asarray(n_plasma_dens))
    n_nums_sort = np.asarray(n_nums)[n_sorted_inds]
    n_plasma_dens_sort = np.asarray(n_plasma_dens)[n_sorted_inds]
    n_mean_freqs_sort = np.asarray(n_mean_freqs)[n_sorted_inds]

    m_sorted_inds = np.argsort(np.asarray(m_plasma_dens))
    m_nums_sort = np.asarray(m_nums)[m_sorted_inds]
    m_plasma_dens_sort = np.asarray(m_plasma_dens)[m_sorted_inds]
    m_mean_freqs_sort = np.asarray(m_mean_freqs)[m_sorted_inds]

    ex_table = xl.Workbook()
    ex_table.create_sheet(title='Mean_freq', index=0)
    sheet = ex_table['Mean_freq']
    sheet['A1'] = 'Номер(шум)'
    sheet['B1'] = 'Плотность плазмы, отн.ед.'
    sheet['C1'] = 'Средняя частота, ГГц'

    sheet['E1'] = 'Номер'
    sheet['F1'] = 'Плотность плазмы, отн.ед.'
    sheet['I1'] = 'Средняя частота, ГГц'

    for z in range(n_nums_sort.size):
        cell = sheet.cell(row=z + 2, column=1)
        cell.value = int(n_nums_sort[z])
        cell = sheet.cell(row=z + 2, column=2)
        cell.value = n_plasma_dens_sort[z]
        cell = sheet.cell(row=z + 2, column=3)
        cell.value = n_mean_freqs_sort[z]

    for k in range(m_nums_sort.size):
        cell = sheet.cell(row=k + 2, column=5)
        cell.value = int(m_nums_sort[k])
        cell = sheet.cell(row=k + 2, column=6)
        cell.value = m_plasma_dens_sort[k]
        cell = sheet.cell(row=k + 2, column=7)
        cell.value = m_mean_freqs_sort[k]


    path = proc.excel_folder_path / f'Mean_freq_{exp_num}_test.xlsx'
    ex_table.save(path)

#noise_mean_freq_calc(210304)


def peak_mean_freq_calc(exp_num, central_freq=2.714e9, band_half_width=50e6):
    proc = ProcessSignal(f'{exp_num}')
    types = proc.read_type_file()
    csv_signals, csv_signal_nums = types['signal_files'], types['signal_nums']
    excel_results = proc.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_results['magnetron']

    m_nums, m_plasma_dens, m_mean_freqs = [], [], []

    for num in magnetron_nums:
        file_name = f'str{num}.csv'
        file_data = proc.open_file(file_name, reduced=True)
        t, u, dt = file_data['time'], file_data['voltage'], file_data['time_resolution']
        pl_density = proc.read_excel(file_name)['dicts'][num]['Ток плазмы, А']
        fft_results = proc.fft_amplitude(t, u)
        freqs, amps = fft_results['frequency'], fft_results['amplitude']

        left_boundary, right_boundary = central_freq - band_half_width, central_freq + band_half_width
        noise_base_inds = np.logical_and(freqs >= left_boundary, freqs <= right_boundary)
        noise_base_freqs, noise_base_amps = freqs[noise_base_inds], amps[noise_base_inds]
        mean_freq = proc.mean_frequency(noise_base_freqs, noise_base_amps)
        spectrum_mean_freq = mean_freq['mean_freq']
        m_nums.append(num), m_plasma_dens.append(pl_density), m_mean_freqs.append(spectrum_mean_freq)

    m_sorted_inds = np.argsort(np.asarray(m_plasma_dens))
    m_nums_sort = np.asarray(m_nums)[m_sorted_inds]
    m_plasma_dens_sort = np.asarray(m_plasma_dens)[m_sorted_inds]
    m_mean_freqs_sort = np.asarray(m_mean_freqs)[m_sorted_inds]

    ex_table = xl.Workbook()
    ex_table.create_sheet(title='Mean_freq', index=0)
    sheet = ex_table['Mean_freq']
    sheet['A1'] = 'Номер'
    sheet['B1'] = 'Плотность плазмы, отн.ед.'
    sheet['C1'] = 'Средняя частота, ГГц'

    for k in range(m_nums_sort.size):
        cell = sheet.cell(row=k + 2, column=1)
        cell.value = int(m_nums_sort[k])
        cell = sheet.cell(row=k + 2, column=2)
        cell.value = m_plasma_dens_sort[k]
        cell = sheet.cell(row=k + 2, column=3)
        cell.value = m_mean_freqs_sort[k]

    path = proc.excel_folder_path / f'Mean_freq_{exp_num}_test.xlsx'
    ex_table.save(path)


#peak_mean_freq_calc(210322)