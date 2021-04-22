from pathlib import Path
import csv
import os
import numpy as np
from scipy.fftpack import rfft, rfftfreq, irfft
import matplotlib.pyplot as plt
from scipy.signal import iirfilter, sosfilt, zpk2sos
from scipy.stats import linregress
#from obspy.signal.filter import bandstop
import openpyxl as xl
default_work_folder_path = r"C:\Users\d_Nice\Documents\SignalProcessing\2021"

class ProcessSignal:
    def __init__(self, experiment_number,
                 work_folder_path=default_work_folder_path,
                 series_meas=True):
        self.work_folder_path = Path(work_folder_path)
        self.exp_file_path = self.work_folder_path / experiment_number
        self.excel_folder_path = self.exp_file_path / 'Excel'
        self.excel_folder_path.mkdir(parents=True, exist_ok=True)
        self.main_excel_path = self.exp_file_path / '{}.xlsx'.format(experiment_number)
        self.fft_excel_file_path = self.excel_folder_path / 'fft_{}.xlsx'.format(experiment_number)
        self.integral_excel_path = self.excel_folder_path / 'integral_{}.xlsx'.format(experiment_number)
        if series_meas:
            self.pics_path = self.exp_file_path / "Pictures"
            self.pics_path.mkdir(parents=True, exist_ok=True)
            self.csv_files_path = self.exp_file_path / "CSV"
            self.csv_files_path.mkdir(parents=True, exist_ok=True)
            self.fft_pics_path = self.pics_path / "FFT"
            self.fft_pics_path.mkdir(parents=True, exist_ok=True)
            self.fft_part_path = self.fft_pics_path / "Part FFT"
            self.fft_part_path.mkdir(parents=True, exist_ok=True)
            self.fft_noise_base = self.fft_pics_path / "Спектр шумового пъедестала"
            self.fft_noise_base.mkdir(parents=True, exist_ok=True)
            self.fft_magnetron = self.fft_pics_path / "Спектр усиленного сигнала"
            self.fft_magnetron.mkdir(parents=True, exist_ok=True)
            self.fft_peak = self.fft_pics_path / "Пик"
            self.fft_peak.mkdir(parents=True, exist_ok=True)
            self.fft_reb = self.fft_pics_path / "Спектр шумов РЭП"
            self.fft_reb.mkdir(parents=True, exist_ok=True)
            self.signal_pics_path = self.pics_path / "Oscillogramms"
            self.signal_pics_path.mkdir(parents=True, exist_ok=True)
        else:
            self.fft_full_path = self.exp_file_path
            self.fft_part_path = self.exp_file_path
            self.signal_pics_path = self.exp_file_path

    def csv_files(self):
        file_list = os.listdir(self.exp_file_path)
        num_of_files = len(file_list)
        csv_files = []
        for i in range(num_of_files):
            if '.csv' in file_list[i]:
                csv_files.append(file_list[i])
        return csv_files

    def open_file(self, file_name, reduced=False, red_type=262):
        if reduced:
            if red_type == 262:
                file_path = self.csv_files_path / '{}'.format(file_name)
            else:
                folder_path = self.exp_file_path / f"CSV_{red_type}.0"
                file_path = folder_path / '{}'.format(file_name)
        else:
            file_path = self.exp_file_path / '{}'.format(file_name)
        t = []
        u = []
        with open(str(file_path)) as File:
            reader = csv.reader(File)
            for row in reader:
                t.append(float(row[3]))
        with open(str(file_path)) as File:
            reader = csv.reader(File)
            for row in reader:
                u.append(float(row[4]))
        dt = t[1] - t[0]
        t = np.array(t)
        u = np.array(u)
        file_dict = {'time': t,
                     'voltage': u,
                     'time_resolution': dt}
        return file_dict

    def files_classification(self):
        csv_files = self.csv_files()
        doc_path = self.excel_folder_path / 'types.xlsx'
        wb = xl.Workbook()
        wb.create_sheet(title='Лист1', index=0)
        sheet = wb['Лист1']
        headings_list = ['signal_files', 'signal_nums', 'voltage_files', 'voltage_nums']
        for i, heading in enumerate(headings_list):
            cell_k = sheet.cell(row=1, column=i + 1)
            cell_k.value = '{}'.format(heading)
        j_v = 2
        j_s = 2
        for csv_file in csv_files:
            csv_num = csv_file[3:6]
            file_dict = self.open_file(csv_file)
            voltage = file_dict['voltage']
            crit = np.mean(voltage)
            if crit < -0.45:
                file_dict.update({'file_type': 'Voltage'})
                cell_f = sheet.cell(row=j_v, column=3)
                cell_f.value = csv_file
                cell_n = sheet.cell(row=j_v, column=4)
                cell_n.value = csv_num
                j_v += 1
            else:
                file_dict.update({'file_type': 'Signal'})
                cell_f = sheet.cell(row=j_s, column=1)
                cell_f.value = csv_file
                cell_n = sheet.cell(row=j_s, column=2)
                cell_n.value = csv_num
                j_s += 1
        wb.save(str(doc_path))

    def reduce_files(self):
        types = self.read_type_file()
        csv_signals = types['signal_files']
        print('CSV files {}-{} obtained...'.format(csv_signals[0][3:6], csv_signals[-1][3:6]))
        for csv_signal in csv_signals:
            time = self.open_file(csv_signal)['time']
            voltage = self.open_file(csv_signal)['voltage']
            red_inds = np.logical_and(70e-9 <= time, time <= 332e-9)
            red_t = time[red_inds]
            red_u = voltage[red_inds]
            file_path = self.csv_files_path / csv_signal
            file = open(str(file_path), 'w', newline='')
            with file:
                writer = csv.writer(file)
                for i in range(0, len(red_t)):
                    writer.writerow([0, 0, 0, red_t[i], red_u[i]])

    def reduce_fft(self, time_0=100e-9, time_interval=262e-9, prelim_view=False):
        csv_types = self.read_type_file()
        csv_signals = csv_types['signal_files']
        #csv_signals = ['str097.csv']
        csv_gin_files = csv_types['voltage_files']
        time_1 = time_0 + time_interval + 13e-9
        for j, csv_signal in enumerate(csv_signals):
            file = self.open_file(csv_signal)
            raw_time, raw_voltage = file['time'], file['voltage']
            work_inds = np.logical_and(time_0 <= raw_time, raw_time <= time_1)
            time = raw_time[work_inds]
            voltage = file['voltage'][work_inds]
            voltage_signs = np.sign(voltage)
            m = 0
            l = 0
            start_inds = []
            end_inds = []
            for i in range(len(voltage_signs) - 1):
                if l < 1:
                    if voltage_signs[i] != voltage_signs[i + 1] and voltage_signs[i + 1] != 0:
                        m += 1
                        if m == 1:
                            start_inds.append(i)
                            start_inds.append(i + 1)
                        else:
                            if m % 2 != 0:
                                if time[i + 1] >= time[start_inds][0] + time_interval:
                                    l += 1
                                    end_inds.append(i)
                                    end_inds.append(i + 1)
            start_volt = voltage[start_inds]
            start_time = time[start_inds]
            end_volt = voltage[end_inds]
            end_time = time[end_inds]
            if len(start_time) > 1 and len(end_time) > 1:
                if start_time[-1] != end_time[-1]:
                    x = np.array([end_time[0], end_time[-1]])
                    y = np.array([end_volt[0], end_volt[-1]])
                    k, b, r_value, p_value, std_err = linregress(x, y)
                    y_ideal = start_volt[-1]
                    x_ideal = (y_ideal - b) / k
                    ind_cut = np.logical_and(start_time[1] <= time, time <= end_time[1])
                    time_cut = time[ind_cut]
                    time_cut[-1] = x_ideal
                    volt_cut = voltage[ind_cut]
                    volt_cut[-1] = y_ideal
                if prelim_view:
                    #print("Creating ocsillogramm...")
                    gin_file = self.open_file(csv_gin_files[j])
                    gin_t, gin_v = gin_file['time'], gin_file['voltage']
                    plt.plot((gin_t - 50e-9) / 1e-9, gin_v, color='blue')
                    plt.plot(raw_time / 1e-9, raw_voltage, color='green')
                    plt.plot(time_cut / 1e-9, volt_cut, color='orange')
                    plt.vlines(start_time / 1e-9, -2.5, 2.5)
                    plt.xlabel('t, ns')
                    plt.ylabel('v')
                    plt.ylim(bottom=-2, top=2)
                    plt.xlim(left=-135, right=720)
                    plt.title(f'{csv_signal}')
                    #plt.plot(time_cut[0], volt_cut[0], marker='o', color='red')
                    #plt.plot(time_cut[-1], volt_cut[-1], marker='o', color='red')
                    plt.show()
                else:
                    if time_interval == 262e-9:
                        file_path = self.csv_files_path / csv_signal
                    else:
                        folder_path = self.exp_file_path / f"CSV_{time_interval / 1e-9}"
                        folder_path.mkdir(parents=True, exist_ok=True)
                        file_path = folder_path / csv_signal
                    file = open(str(file_path), 'w', newline='')
                    with file:
                        writer = csv.writer(file)
                        for i in range(0, len(time_cut)):
                            writer.writerow([0, 0, 0, time_cut[i], volt_cut[i]])
        else:
            pass

    def read_type_file(self):
        doc_path = self.excel_folder_path / 'types.xlsx'
        wb = xl.load_workbook(doc_path)
        sheet = wb['Лист1']
        rows = sheet.max_row
        print(doc_path)
        csv_types = {}
        for j in range(4):
            cell_k = sheet.cell(row=1, column=j + 1)
            key_name = cell_k.value
            vals = []
            for i in range(1, rows):
                cell = sheet.cell(row=i + 1, column=j + 1)
                cell_val = cell.value
                if cell_val is not None:
                    vals.append(cell_val)
            csv_types[key_name] = vals
        return csv_types

    def read_excel(self, csv_signal_nums):
        wb = xl.load_workbook(self.main_excel_path)
        sheet = wb['Лист1']
        rows = sheet.max_row
        cols = sheet.max_column
        row_min = 1
        col_max = 1
        for i in range(1, rows):
            cell = sheet.cell(row=i, column=1)
            cell_row_val = cell.value
            if cell_row_val == 'Номер файла' and row_min == 1:
                row_min = i

        for m in range(1, cols+1):
            cell = sheet.cell(row=row_min, column=m)
            cell_col_val = cell.value
            if cell_col_val == 'Комментарий':
                col_max = m

        keys = []
        for k in range(1, col_max + 1):
            cell = sheet.cell(row=row_min, column=k)
            cell_key = cell.value
            keys.append(cell_key)

        row_dicts = []
        for l in range(row_min + 1, rows+1):
            row_dict = {}
            vals = []
            for m in range(1, col_max + 1):
                cell = sheet.cell(row=l, column=m)
                cell_val = cell.value
                vals.append(cell_val)
            for i, key in enumerate(keys):
                row_dict[key] = vals[i]
            row_dicts.append(row_dict)

        plasma_dicts = []
        reb_dicts = []
        magnetron_dicts = []
        for row_dict in row_dicts:
            fnum = row_dict['Номер файла']
            gin_voltage = row_dict['Напряжение ГИНа']
            d_plasma = row_dict['Ток плазмы, А']
            heating = row_dict['Накал']
            magnetron_delay = row_dict['Задержка магнетрона, нс']
            #magnetron_in_voltage = row_dict['Входное напряжение магнетрона, В']
            comment = row_dict['Комментарий']
            if isinstance(fnum, int):
                if isinstance(d_plasma, float) or isinstance(d_plasma, int):
                    if comment != 'except':
                        if isinstance(magnetron_delay, float) or isinstance(magnetron_delay, int):
                            magnetron_dicts.append(row_dict)
                        else:
                            plasma_dicts.append(row_dict)
                elif heating is None and isinstance(gin_voltage, float):
                    reb_dicts.append(row_dict)

        useful_dicts = [plasma_dicts, reb_dicts, magnetron_dicts]

        proc_signal_dicts = {}
        use_nums = {}
        list_signals = ['noise', 'reb', 'magnetron']
        for i, use_dict in enumerate(useful_dicts):
            #use_csv_files = []
            fnums = []
            for row_dict in use_dict:
                fnum = row_dict['Номер файла']
                if fnum < 100:
                    num_1 = fnum // 10
                    num_2 = fnum % 10
                elif fnum > 10000:
                    num_1 = fnum // 1000
                    num_2 = fnum % 1000
                else:
                    num_1 = fnum // 100
                    num_2 = fnum % 100
                fnum_1 = '{:03d}'.format(num_1)
                fnum_2 = '{:03d}'.format(num_2)

                if fnum_1 in csv_signal_nums:
                    row_dict['Номер файла'] = [fnum_1]
                    fnums.append(fnum_1)
                    #use_csv_files.append(row_dict)
                    proc_signal_dicts[fnum_1] = row_dict
                elif fnum_2 in csv_signal_nums:
                    row_dict['Номер файла'] = [fnum_2]
                    fnums.append(fnum_2)
                    #use_csv_files.append(row_dict)
                    proc_signal_dicts[fnum_2] = row_dict
            #proc_signal_dicts[list_signals[i]] = use_csv_files
            use_nums[list_signals[i]] = fnums

        excel_results = {'dicts': proc_signal_dicts,
                         'numbers': use_nums}
        return excel_results

    def signal_envelope(self, t, u, dt, time_frame=1e-8):
        num_of_pts = int(time_frame / dt)
        times = t[::num_of_pts]
        time_step = np.abs(times[0] - times[1])
        t_env = []
        u_env = []
        for time in times:
            time_lim = time + time_step
            ind_t = np.logical_and(t > time, t <= time_lim)
            t_frame = t[ind_t]
            u_frame = u[ind_t]
            ind_max = np.argmax(u_frame)
            t_env.append(t_frame[ind_max])
            u_env.append(u_frame[ind_max])
        envelope_dict = {'env_time': t_env,
                         'env_voltage': u_env}
        return envelope_dict

    def useful_part(self, t, u, dt, max_part):
        env = self.signal_envelope(t, u, dt)
        env_t = np.array(env['env_time'])
        env_u = np.array(env['env_voltage'])
        max_env_u = np.max(env_u)
        ind_lim = env_u > max_part * max_env_u
        t_lim = env_t[ind_lim]
        ind_stop = len(t_lim-1)
        l = 0
        for i in range(len(t_lim)-1):
            if l < 1:
                if t_lim[i+1] - t_lim[i] > 90e-9:
                    ind_stop = i
                    l += 1
        try:
            t_bond = t_lim[:ind_stop:]
            u_zeros = np.zeros(t_bond.size)
            t_min, t_max = t_bond[0], t_bond[-1]
            ind_use = np.logical_and(t >= t_min, t <= t_max)
            t_use = t[ind_use]
            u_use = u[ind_use]
            #plt.plot(t, u)
            #plt.plot(env_t, env_u)
            #plt.plot(t_use, u_use)
            #plt.plot(t_bond, u_zeros, marker='o', color='red', linestyle='')
            #plt.show()
            signal_part_dict = {'signal_time': t_use,
                                'signal_voltage': u_use}
            return signal_part_dict
        except:
            pass

    def proc_part_130_290(self, t, u):
        t_inds = np.logical_and(t >= 130e-9, t <= 390e-9)
        part_t = t[t_inds]
        part_u = u[t_inds]
        reduced_130_dict = {'part_t': part_t,
                            'part_u': part_u}
        return reduced_130_dict


    def e_square(self, t, u, plot=False):
        u_sqre = u * u
        d_time = np.abs(t[1] - t[0])
        if plot is True:
            integ_arr = np.zeros(t.size - 1)
            time_arr = np.zeros(t.size - 1)
            for i in range(1, t.size):
                ind = t <= t[i]
                times = t[ind]
                time_el = times[-1]
                u_sqre_el = u_sqre[ind]
                integ_arr_el = np.trapz(u_sqre_el, x=times, dx=d_time)
                integ_arr[i - 1] = integ_arr_el
                time_arr[i - 1] = time_el
            e_square_plot = {'time_integ': time_arr,
                             'e_square_integ': integ_arr}
            return e_square_plot
        if plot is False:
            e_square_integral = np.trapz(u_sqre, x=t, dx=d_time)
            return e_square_integral

    def left_right_table(self, nums, full_integs, left_integs,
                         right_integs, pl_densities, file_name):
        ex_table = xl.Workbook()
        ex_table.create_sheet(title='LeftRight', index=0)
        sheet = ex_table['LeftRight']
        sheet['A1'] = 'File Number'
        sheet['B1'] = 'Plasma density, arb.units'
        sheet['C1'] = 'K*Integral A^2'
        sheet['D1'] = 'Integral, f<2.725GHz'
        sheet['E1'] = 'Integral, f>2.755GHz'
        for i in range(pl_densities.size):
            cell = sheet.cell(row=i+2, column=1)
            cell.value = nums[i]
            cell = sheet.cell(row=i+2, column=2)
            cell.value = pl_densities[i]
            cell = sheet.cell(row=i+2, column=3)
            cell.value = full_integs[i]
            cell = sheet.cell(row=i + 2, column=4)
            cell.value = left_integs[i]
            cell = sheet.cell(row=i + 2, column=5)
            cell.value = right_integs[i]
        path = self.excel_folder_path / 'left_right_{}.xlsx'.format(file_name)
        ex_table.save(path)

    def integrals_table(self, magnetron_nums, magnetron_pl_ds,
                        peak_power, magnetron_noise_power,
                        plasma_nums, pl_ds, reb_noise_power,
                        file_name):
        ex_table = xl.Workbook()
        ex_table.create_sheet(title='Power', index=0)
        sheet = ex_table['Power']
        sheet['A1'] = 'Номер файла(магнетрон)'
        sheet['B1'] = 'Плотность плазмы (магнетрон)'
        sheet['C1'] = 'Энергия в пике'
        sheet['D1'] = 'Энергия в шумовом пьедестале'
        sheet['E1'] = 'Сигнал / шум'
        sheet['F1'] = 'Номер файла(шумы)'
        sheet['G1'] = 'Плотность плазмы (шумы)'
        sheet['H1'] = 'Энергия шумов РЭП'
        size_1 = magnetron_pl_ds.size
        size_2 = pl_ds.size
        for i in range(size_1):
            cell = sheet.cell(row=i+2, column=1)
            cell.value = magnetron_nums[i]
            cell = sheet.cell(row=i+2, column=2)
            cell.value = magnetron_pl_ds[i]
            cell = sheet.cell(row=i+2, column=3)
            cell.value = peak_power[i]
            cell = sheet.cell(row=i + 2, column=4)
            cell.value = magnetron_noise_power[i]
            cell = sheet.cell(row=i + 2, column=5)
            cell.value = peak_power[i] / magnetron_noise_power[i]
        for j in range(size_2):
            cell = sheet.cell(row=j + 2, column=6)
            cell.value = plasma_nums[j]
            cell = sheet.cell(row=j + 2, column=7)
            cell.value = pl_ds[j]
            cell = sheet.cell(row=j + 2, column=8)
            cell.value = reb_noise_power[j]
        path = self.excel_folder_path / 'integral_{}.xlsx'.format(file_name)
        ex_table.save(path)

    def left_right_koef_table(self, magnetron_nums, magnetron_pl_ds,
                        magnetron_koefs,
                        plasma_nums, pl_ds, reb_koefs,
                        file_name):
        ex_table = xl.Workbook()
        ex_table.create_sheet(title='Left_right_koef', index=0)
        sheet = ex_table['Left_right_koef']
        sheet['A1'] = 'Номер файла(магнетрон)'
        sheet['B1'] = 'Плотность плазмы (магнетрон)'
        sheet['C1'] = 'W(f>2.755 ГГц)/W(f<2.725ГГц)'
        sheet['D1'] = 'Номер файла(шумы)'
        sheet['E1'] = 'Плотность плазмы (шумы)'
        sheet['F1'] = 'W(f>2.755 ГГц)/W(f<2.725ГГц)Шумы'
        size_1 = magnetron_pl_ds.size
        size_2 = pl_ds.size
        for i in range(size_1):
            cell = sheet.cell(row=i+2, column=1)
            cell.value = magnetron_nums[i]
            cell = sheet.cell(row=i+2, column=2)
            cell.value = magnetron_pl_ds[i]
            cell = sheet.cell(row=i+2, column=3)
            cell.value = magnetron_koefs[i]
        for j in range(size_2):
            cell = sheet.cell(row=j + 2, column=4)
            cell.value = plasma_nums[j]
            cell = sheet.cell(row=j + 2, column=5)
            cell.value = pl_ds[j]
            cell = sheet.cell(row=j + 2, column=6)
            cell.value = reb_koefs[j]
        path = self.excel_folder_path / 'left_right_koef_{}.xlsx'.format(file_name)
        ex_table.save(path)

    def one_side_table(self, magnetron_nums, magnetron_power,
                       magnetron_pl_ds,
                       plasma_nums, pl_ds, reb_power,
                       file_name):
        ex_table = xl.Workbook()
        ex_table.create_sheet(title='{}_power'.format(file_name), index=0)
        sheet = ex_table['{}_power'.format(file_name)]
        sheet['A1'] = 'Номер файла(магнетрон)'
        sheet['B1'] = 'Плотность плазмы (магнетрон)'
        sheet['C1'] = 'Энергия(магнетрон)'
        sheet['D1'] = 'Номер файла(шумы)'
        sheet['E1'] = 'Плотность плазмы (шумы)'
        sheet['F1'] = 'Энергия(шумы)'
        size_1 = magnetron_pl_ds.size
        size_2 = pl_ds.size
        for i in range(size_1):
            cell = sheet.cell(row=i+2, column=1)
            cell.value = magnetron_nums[i]
            cell = sheet.cell(row=i+2, column=2)
            cell.value = magnetron_pl_ds[i]
            cell = sheet.cell(row=i+2, column=3)
            cell.value = magnetron_power[i]
        for j in range(size_2):
            cell = sheet.cell(row=j + 2, column=4)
            cell.value = plasma_nums[j]
            cell = sheet.cell(row=j + 2, column=5)
            cell.value = pl_ds[j]
            cell = sheet.cell(row=j + 2, column=6)
            cell.value = reb_power[j]
        path = self.excel_folder_path / '{}_power.xlsx'.format(file_name)
        ex_table.save(path)


    def signal_parts(self, t, u):
        time_frame = t[-1] / 3
        start_time = t[0]
        part_dict = {}
        for i in range(3):
            end_time = start_time + time_frame
            frame_ind = np.logical_and(t > start_time, t < end_time)
            t_part = t[frame_ind]
            u_part = u[frame_ind]
            part_dict['Part{}'.format(i)] = {'time': t_part,
                                             'voltage': u_part}
            start_time = end_time
        return part_dict

    def fft_signal(self, t, u, dt, cutoff_frequency=6e9):
        len_t = len(t)
        u_fft = rfft(u)
        freq_fft = rfftfreq(len_t, dt)
        n = len(freq_fft)
        p_freq = freq_fft[1:n-1:2]
        p_fft = np.sqrt(u_fft[1:n-1:2] ** 2 + u_fft[2:n:2] ** 2)*dt
        ind_cutoff = p_freq <= cutoff_frequency
        cut_freq = p_freq[ind_cutoff]
        cut_fft_amp = p_fft[ind_cutoff]
        fft_dict = {'frequency': cut_freq,
                    'amplitude': cut_fft_amp}
        return fft_dict

    def fft_filter(self, t, u, low_freq=2.695e9, high_freq=2.725e9, filt_type='bandpass'):
        len_t = len(t)
        dt = np.abs(t[1] - t[0])
        fft_u = rfft(u)
        freqs_fft = rfftfreq(len_t, dt)
        ind_mask = np.logical_and(low_freq < freqs_fft, freqs_fft < high_freq)
        if filt_type == 'bandpass':
            bandpass_filter = np.zeros(len(fft_u))
            bandpass_filter[ind_mask] = 1
            b_filt_u = irfft(fft_u * bandpass_filter)
            return b_filt_u
        elif filt_type == 'bandstop':
            bandpass_filter = np.ones(len(fft_u))
            bandpass_filter[ind_mask] = 0
            b_filt_u = irfft(fft_u * bandpass_filter)
            return b_filt_u
        else:
            print('Unknown filter type')


    def fft_amplitude(self, t, u):
        dt = np.abs(t[1] - t[0])
        len_t = len(t)
        u_fft = rfft(u)
        freq_fft = rfftfreq(len_t, dt)
        n = len(u)
        p_freq = freq_fft[1:n:2]
        n_freq = len(p_freq)
        amp_fft = np.zeros(n_freq)
        amp_fft[0] = u_fft[0] / n
        if n % 2 == 0:
            amp_fft[n_freq - 1] = u_fft[-1] / n
            j = 1
            for i in range(1, n_freq - 2):
                amp_fft[i] = 2 * np.sqrt(u_fft[j] ** 2 + u_fft[j + 1] ** 2) / n
                j += 2
        if n % 2 != 0:
            j = 1
            for i in range(1, n_freq - 1):
                amp_fft[i] = 2 * np.sqrt(u_fft[j] ** 2 + u_fft[j + 1] ** 2) / n
                j += 2

        ind_fake_freqs_1 = np.logical_or(p_freq <= 1.249e9, 1.251e9 <= p_freq)
        p_freq_1 = p_freq[ind_fake_freqs_1]
        amp_fft_1 = amp_fft[ind_fake_freqs_1]

        ind_fake_freqs_2 = np.logical_or(p_freq_1 <= 2.495e9, 2.505e9 <= p_freq_1)
        p_freq_2 = p_freq_1[ind_fake_freqs_2]
        amp_fft_2 = amp_fft_1[ind_fake_freqs_2]

        #ind_cutoff = p_freq_2 <= cutoff_frequency
        #cut_freq = p_freq_2[ind_cutoff]
        #cut_fft_amp = amp_fft_2[ind_cutoff]
        fft_amp_dict = {'frequency': p_freq_2,
                        'amplitude': amp_fft_2}
        return fft_amp_dict

    def mean_frequency(self, freq, magnitude):
        m_sqre = magnitude * magnitude
        d_freq = freq[1] - freq[0]
        integ_arr = np.zeros(freq.size - 1)
        freq_arr = np.zeros(freq.size - 1)
        for i in range(1, freq.size):
            ind = freq <= freq[i]
            freqs = freq[ind]
            freq_el = freqs[-1]
            m_sqre_el = m_sqre[ind]
            integ_arr_el = np.trapz(m_sqre_el, x=freqs, dx=d_freq)
            integ_arr[i-1] = integ_arr_el
            freq_arr[i-1] = freq_el
        half_integral = 0.5 * np.max(integ_arr)
        try:
            ind_mean_left = integ_arr <= half_integral
            mean_freq_left = freq_arr[ind_mean_left][-1]
            int_left = integ_arr[ind_mean_left][-1]

            ind_mean_right = integ_arr >= half_integral
            mean_freq_right = freq_arr[ind_mean_right][0]
            int_right = integ_arr[ind_mean_right][0]

            if mean_freq_left != mean_freq_right:
                x = np.array([mean_freq_left, mean_freq_right])
                y = np.array([int_left, int_right])
                k, b, r_value, p_value, std_err = linregress(x, y)
                freq_interp = np.linspace(mean_freq_right, mean_freq_left, num=200)
                square_interp = k * freq_interp + b
                linreg_inds = square_interp <= half_integral
                linreg_mean_freq = freq_interp[linreg_inds][0]
            #plt.plot(freq_arr, integ_arr)
            #plt.plot(linreg_mean_freq, square_interp[linreg_inds][0], marker='.', color='red', ms=0.5, linestyle='')
            #plt.hlines(y=half_integral, xmin=2.72e9, xmax=2.76e9, linewidth=0.7)
            #plt.vlines(x=linreg_mean_freq, ymin=0, ymax=square_interp[0], linewidth=0.7)
            #plt.show()
            mean_freq = np.round((linreg_mean_freq / 1e9), 3)
            mean_freq_dict = {'mean_freqs': freq_arr,
                              'mean_amp_sqre': integ_arr,
                              'mean_freq': mean_freq}
            return mean_freq_dict
        except IndexError:
            pass



    def bandstop_filter(self, time, voltage, freqmin, freqmax,
                        corners=4, zerophase=False):
        dt = np.mean(np.diff(time))
        df = 1 / dt
        fe = 0.5 * df
        low = freqmin / fe
        high = freqmax / fe
        z, p, k = iirfilter(corners, [low, high],
                            btype='bandstop', ftype='butter', output='zpk')
        sos = zpk2sos(z, p, k)
        if zerophase:
            firstpass = sosfilt(sos, voltage)
            return sosfilt(sos, firstpass[::-1])[::-1]
        else:
            return sosfilt(sos, voltage)

    def bandpass_filter(self, time, voltage, freqmin, freqmax,
                        corners=4, zerophase=False):
        dt = np.mean(np.diff(time))
        df = 1 / dt
        fe = 0.5 * df
        low = freqmin / fe
        high = freqmax / fe
        z, p, k = iirfilter(corners, [low, high], btype='band',
                            ftype='butter', output='zpk')
        sos = zpk2sos(z, p, k)
        if zerophase:
            firstpass = sosfilt(sos, voltage)
            return sosfilt(sos, firstpass[::-1])[::-1]
        else:
            return sosfilt(sos, voltage)

    def lowpass(self, time, voltage, freq,
                corners=4, zerophase=False):
        dt = np.mean(np.diff(time))
        df = 1 / dt
        fe = 0.5 * df
        f = freq / fe
        z, p, k = iirfilter(corners, f, btype='lowpass', ftype='butter',
                            output='zpk')
        sos = zpk2sos(z, p, k)
        if zerophase:
            firstpass = sosfilt(sos, voltage)
            return sosfilt(sos, firstpass[::-1])[::-1]
        else:
            return sosfilt(sos, voltage)

    def highpass(self, time, voltage, freq,
                 corners=4, zerophase=False):
        dt = np.mean(np.diff(time))
        df = 1 / dt
        fe = 0.5 * df
        f = freq / fe
        z, p, k = iirfilter(corners, f, btype='highpass', ftype='butter',
                            output='zpk')
        sos = zpk2sos(z, p, k)
        if zerophase:
            firstpass = sosfilt(sos, voltage)
            return sosfilt(sos, firstpass[::-1])[::-1]
        else:
            return sosfilt(sos, voltage)


    def oscill_picture(self, num, t, u, u_filt, pl_dens, abs, save=False):
        fig = plt.figure(num=1, dpi=150)
        ax = fig.add_subplot(111)
        line1, = ax.plot(t, u, linewidth=0.7, color='k')
        line2, = ax.plot(t, u_filt, linewidth=0.7, color='red')
        line1.set_label('Signal')
        line2.set_label('Bandpass filter (2,74 +/- 0.015) GHz')
        ax.set_xlabel(r'$Time, ns$')
        ax.set_ylabel(r'$U, V$')
        ax.grid(which='both', axis='both')
        ax.legend()
        if abs != 0:
            ax.set_title('{}, n={}, поглотители:{}'.format(num, pl_dens, abs))
        else:
            ax.set_title('{}, n={}'.format(num, pl_dens))
        #plt.show()
        if save is True:
            png_name = self.signal_pics_path / num
            fig.savefig(png_name)
        plt.close(fig)

    def part_fft(self, csv_signals, interest_nums,
                 part_nums, fft_type='part',
                 block_full=False, block_part=True,
                 peak=False, noise=False):
        ex_table = xl.Workbook()
        ex_table.create_sheet(title='Full FFT', index=0)
        ex_table.create_sheet(title='Part FFT', index=1)
        sheet1 = ex_table['Full FFT']
        sheet2 = ex_table['Part FFT']
        sheet2['A1'] = 'File Number'
        sheet1['A1'] = 'File Number'
        signal_dict = {}
        row_f = 1
        row_p = 1
        for j, csv_signal in enumerate(csv_signals):
            signal_num = csv_signal[3:6]
            if signal_num in interest_nums or signal_num in part_nums:
                use_signal = self.open_file(csv_signal, reduced=False)
                use_t = use_signal['time']
                use_u = use_signal['voltage']
                dt = use_signal['time_resolution']
                fig = plt.figure(num=1, dpi=300)
                ax = fig.add_subplot(111)
                line = None
                ax.set_prop_cycle(color=['mediumseagreen', 'dodgerblue', 'indigo'])
                if fft_type == 'full' or fft_type == 'both':
                    if signal_num in interest_nums:
                        print('{} am in interest_nums'.format(signal_num))
                        if block_full is True:
                            filtered_u = self.fft_filter(use_t, use_u, 2.69e9, 2.74e9, filt_type='bandstop')
                            fft_results = self.fft_amplitude(use_t, filtered_u)
                        else:
                            fft_results = self.fft_amplitude(use_t, use_u)
                        pl_density = self.read_excel(interest_nums)['dicts'][signal_num]['Ток плазмы, А']
                        freq = fft_results['frequency']
                        amp = fft_results['amplitude']

                        if peak:
                            peak_inds = np.logical_and(2.69e9 < freq, freq < 2.74e9)
                            peak_freqs = freq[peak_inds]
                            peak_amps = amp[peak_inds]
                            mean_freq = self.mean_frequency(peak_freqs, peak_amps)
                            #line, = ax.plot(peak_freqs, peak_amps, linewidth=1.2)
                            line, = ax.plot(freq, amp, linewidth=1.2)
                            peak_max = np.round(np.max(peak_amps), 2)
                        else:
                            mean_freq = self.mean_frequency(freq, amp)
                            line, = ax.plot(freq, amp, linewidth=0.7)
                            line, = ax.plot(freq, amp, linewidth=0.7)
                        try:
                            spectrum_mean_freq = mean_freq['mean_freq']
                            line.set_label(r'$f = {} GHz$'.format(spectrum_mean_freq))

                            value_f = str('f, GHz')
                            cell_f = sheet1.cell(row=1, column=2)
                            cell_f.value = value_f

                            cell_pl = sheet1.cell(row=1, column=3)
                            cell_pl.value = str('n, arb.units')

                            row_f = row_f + 1
                            cell_name = 'A{}'.format(row_f)
                            sheet1[str(cell_name)] = '{}'.format(signal_num)
                            value = spectrum_mean_freq
                            cell = sheet1.cell(row=row_f, column=2)
                            cell.value = value

                            value_pl = pl_density
                            cell = sheet1.cell(row=row_f, column=3)
                            cell.value = value_pl
                        except TypeError:
                            pass

                if fft_type == 'part' or fft_type == 'both':
                    if signal_num in part_nums:
                        print('{} in part_nums'.format(signal_num))
                        if block_part is True:
                            filtered_u = self.bandstop_filter(use_t, use_u, 2.709e9, 2.769e9)
                            part_signal = self.signal_parts(use_t, filtered_u)
                        else:
                            part_signal = self.signal_parts(use_t, use_u)
                        part_keys = part_signal.keys()
                        part_freq_dict = {}

                        row_p = row_p + 1
                        cell_name = 'A{}'.format(row_p)
                        sheet2[str(cell_name)] = '{}'.format(signal_num)
                        for i, part_key in enumerate(part_keys):
                            k = i + 1
                            col = k + 1
                            part_time = part_signal[part_key]['time']
                            part_voltage = part_signal[part_key]['voltage']
                            fft_results = self.fft_signal(part_time, part_voltage, dt)
                            freq = fft_results['frequency']
                            amp = fft_results['amplitude']
                            line, = ax.plot(freq, amp, linewidth=1.2)

                            mean_freq = self.mean_frequency(freq, amp)
                            part_freq = mean_freq['mean_freq']

                            value_f = str('f{}'.format(k))
                            cell_f = sheet2.cell(row=1, column=col)
                            cell_f.value = value_f

                            value = part_freq
                            cell = sheet2.cell(row=row_p, column=col)
                            cell.value = value
                            part_freq_dict['f{}, GHz'.format(k)] = part_freq
                            line.set_label(r'$f_{}= {} GHz$'.format(k, part_freq))
                        pl_density = self.read_excel(part_nums)['dicts'][signal_num]['Ток плазмы, А']
                        signal_dict[signal_num] = part_freq_dict
                if fft_type == 'part' or fft_type == 'both':
                    if signal_num in part_nums:
                        ax.set_title(r'$File\/Number = {} (Part FFT, n={})$'.format(signal_num, pl_density))
                if fft_type == 'full' or fft_type == 'both':
                    if signal_num in interest_nums:
                        if noise:
                            ax.set_title(r'$File\/Number = {},\/ (n={}) $'.format(signal_num, pl_density))
                        else:
                            #absorbers = self.read_excel(part_nums)['dicts'][signal_num]['Поглотители в тракте магнетрона']
                            ax.set_title(r'$File\/Number = {},\/ \/ n={} $'.format(signal_num, pl_density))

                if line is not None:
                    ax.set_ylim(bottom=0)
                    if peak:
                        ax.set_xlim(left=2.68e9, right=2.74e9)
                    elif fft_type == 'part':
                        ax.set_xlim(left=2.5e9, right=3e9)
                    else:
                        ax.set_xlim(left=0, right=4e9)
                    ax.grid(which='both', axis='both')
                    ax.set_xlabel(r'$Frequency, GHz$')
                    ax.set_ylabel(r'$Amplitude$')
                    ax.legend()
                    if fft_type == 'part' or fft_type == 'both':
                        if signal_num in part_nums:
                            png_name = self.fft_part_path / signal_num
                    if fft_type == 'full' or fft_type == 'both':
                        if signal_num in interest_nums:
                            if noise:
                                png_name = self.fft_reb / 'reb_{}'.format(signal_num)
                                table_path = self.fft_reb / 'fft_reb_{}.xlsx'.format(signal_num)
                            else:
                                if block_full:
                                    png_name = self.fft_noise_base / 'noise_base_{}'.format(signal_num)
                                    table_path = self.fft_noise_base / 'fft_noise_base_{}.xlsx'.format(signal_num)
                                else:
                                    png_name = self.fft_magnetron / 'magnetron_{}'.format(signal_num)
                                    table_path = self.fft_magnetron / 'fft_magnetron_{}.xlsx'.format(signal_num)
                            if peak:
                                png_name = self.fft_peak / 'peak_{}_1'.format(signal_num)
                                table_path = self.fft_peak / 'peak_{}.xlsx'.format(signal_num)
                    fig.savefig(png_name)
                    plt.close(fig)
            print('Circle {} complete'.format(j))
        #ex_table.save(table_path)

    def fft_full(self, fft_type, peak_freq=2.71e9, peak_gate=50e6):
        types = self.read_type_file()
        csv_signals, csv_signal_nums = types['signal_files'], types['signal_nums']
        excel_results = self.read_excel(csv_signal_nums)['numbers']
        noise_nums = excel_results['noise']
        magnetron_nums = excel_results['magnetron']
        if fft_type == 'reb_noise_full':
            nums = noise_nums
        else:
            nums = magnetron_nums
        for num in nums:
            file_name = f'str{num}.csv'
            file_data = self.open_file(file_name, reduced=True)
            t, u, dt = file_data['time'], file_data['voltage'], file_data['time_resolution']
            pl_density = self.read_excel(file_name)['dicts'][num]['Ток плазмы, А']
            fft_results = self.fft_amplitude(t, u)
            freqs, amps = fft_results['frequency'], fft_results['amplitude']
            mean_freq = self.mean_frequency(freqs, amps)
            spectrum_mean_freq = mean_freq['mean_freq']

            fig = plt.figure(num=1, dpi=300)
            ax = fig.add_subplot(111)

            if fft_type == 'magnetron_noise_base':
                base_inds = np.logical_or(freqs < peak_freq - peak_gate, freqs > peak_freq + peak_gate)
                base_freqs, base_amps = freqs[base_inds], amps[base_inds]
                mean_freq = self.mean_frequency(base_freqs, base_amps)
                spectrum_mean_freq = mean_freq['mean_freq']
                line, = ax.plot(base_freqs, base_amps, linewidth=0.7, color='indigo')
            elif fft_type == 'peak':
                line, = ax.plot(freqs, amps, linewidth=1.2, color='mediumseagreen')
            else:
                line, = ax.plot(freqs, amps, linewidth=0.7, color='dodgerblue')
            if fft_type == 'peak':
                ax.set_xlim(left=peak_freq-30e6, right=peak_freq+30e6)
            else:
                ax.set_xlim(left=0, right=4e9)
            ax.set_ylim(bottom=0)
            ax.grid(which='both', axis='both')
            ax.set_xlabel(r'$Frequency, GHz$')
            ax.set_ylabel(r'$Amplitude$')
            ax.set_title(r'$№={}, n={}$'.format(num, pl_density))
            if fft_type != 'peak':
                line.set_label(r'$f = {} GHz$'.format(spectrum_mean_freq))
                ax.legend()
            if fft_type == 'peak':
                png_name = self.fft_peak / f'peak_{num}'
            elif fft_type == 'reb_noise_full':
                png_name = self.fft_reb / f'reb_noise_{num}'
            elif fft_type == 'magnetron_noise_base':
                png_name = self.fft_noise_base / f'noise_base_{num}'
            else:
                png_name = self.fft_magnetron / f'amplifier_{num}'
            fig.savefig(png_name)
            plt.close(fig)