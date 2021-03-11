import numpy as np
import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal
import openpyxl as xl


def write_file(exp_num):
    fft_test = ProcessSignal(f'{exp_num}')
    fft_test.files_classification()
#write_file(210304)

def load_text():
    fft_test = ProcessSignal('190925')
    txt_path = fft_test.exp_file_path / 'types.txt'
    csv_types = fft_test.read_type_file()
    print(csv_types)
#load_text()

def write_csv(exp_num):
    test = ProcessSignal(f'{exp_num}')
    test.reduce_fft(time_0=125e-9, time_interval=130e-9, prelim_view=False)
#write_csv(210211)

def load_red_csv():
    test = ProcessSignal('191001')
    types = test.read_type_file()
    csv_signals = types['signal_files']
    for signal in csv_signals:
        open_dict = test.open_file(signal, reduced=True)
        t = open_dict['time']
        u = open_dict['voltage']
        plt.plot(t, u)
        plt.show()
#load_red_csv()

def fft_amp_test():
    test = ProcessSignal('191011')
    types = test.read_type_file()
    csv_signals = types['signal_files']
    for signal in csv_signals:
        open_dict = test.open_file(signal, reduced=True)
        t = open_dict['time']
        u = open_dict['voltage']
        dt = open_dict['time_resolution']
        amp_dict = test.fft_amplitude(t, u, dt)
        freq = amp_dict['frequency']
        amp = amp_dict['amplitude']
        plt.plot(freq, amp)
        plt.show()
#fft_amp_test()

def single_signal_defpart_fft():
    fft_test = ProcessSignal('190909', series_meas=False)
    csv_file = fft_test.csv_files()
    csv_num = csv_file[0][0:-4]
    fft_test.part_fft(csv_file, interest_nums=csv_num,
                      part_nums=csv_num, fft_type='part')
#single_signal_part_fft()


def full_fft_magnetron(exp_num):
    fft_proc = ProcessSignal(f'{exp_num}')
    fft_proc.fft_full(fft_type='magnetron_full')
#full_fft_magnetron(210204)


def full_fft_noise(exp_num):
    fft_proc = ProcessSignal(f'{exp_num}')
    fft_proc.fft_full(fft_type='reb_noise_full')
#full_fft_noise(210304)


def fft_peak(exp_num):
    fft_proc = ProcessSignal(f'{exp_num}')
    fft_proc.fft_full(fft_type='peak')
#fft_peak(210304)


def fft_magnetron_noise_base(exp_num):
    fft_proc = ProcessSignal(f'{exp_num}')
    fft_proc.fft_full(fft_type='magnetron_noise_base')
#fft_magnetron_noise_base(210302)


def series_fft(exp_num):
    fft_test = ProcessSignal(f'{exp_num}')
    types = fft_test.read_type_file()
    csv_signals = types['signal_files']
    csv_signal_nums = types['signal_nums']
    excel_results = fft_test.read_excel(csv_signal_nums)['numbers']
    noise_nums = excel_results['noise']
    magnetron_nums = excel_results['magnetron']
    print('Magnetron nums are', magnetron_nums)
    print('Noise nums are:', noise_nums)
    fft_test.part_fft(csv_signals,
                      interest_nums=noise_nums,
                      part_nums=noise_nums,
                      fft_type='full', block_full=False,
                      block_part=True, peak=False, noise=True)

#series_fft(210304)



def noise_fft():
    fft_test = ProcessSignal('191001')
    types = fft_test.read_type_file()
    csv_signals = types['signal_files']
    csv_signal_nums = types['signal_nums']
    excel_results = fft_test.read_excel(csv_signal_nums)['numbers']
    noise_nums = excel_results['noise']
    doc_path = fft_test.fft_excel_file_path
    wb = xl.load_workbook(doc_path)
    wb.create_sheet(title='Noise FFT', index=2)
    sheet3 = wb['Noise FFT']
    sheet3['A1'] = 'File Number'
    sheet3['B1'] = 'f, GHz'
    sheet3['C1'] = 'n, arb.units'
    row_f = 1
    for i, csv_signal in enumerate(csv_signals):
        signal_num = csv_signal[3:6]
        if signal_num in noise_nums:
            file = fft_test.open_file(csv_signal, reduced=True)
            use_t = file['time']
            use_u = file['voltage']
            dt = file['time_resolution']
            fft_results = fft_test.fft_amplitude(use_t, use_u, dt)
            freq = fft_results['frequency']
            amp = fft_results['amplitude']
            mean_freq = fft_test.mean_frequency(freq, amp)
            spectrum_mean_freq = mean_freq['mean_freq']
            row_f = row_f + 1
            cell_name = 'A{}'.format(row_f)
            sheet3[str(cell_name)] = '{}'.format(signal_num)
            value = spectrum_mean_freq
            cell = sheet3.cell(row=row_f, column=2)
            cell.value = value
            pl_density = fft_test.read_excel(noise_nums)['dicts'][signal_num]['Ток плазмы, А']
            value_pl = pl_density
            cell = sheet3.cell(row=row_f, column=3)
            cell.value = value_pl
    new_path = fft_test.excel_folder_path / 'noise_fft.xlsx'
    wb.save(new_path)
#noise_fft()

def noise_signal_plot(exp_name, type='e_time', left_right=False):
    fft_test = ProcessSignal('{}'.format(exp_name))
    csv_types = fft_test.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = fft_test.read_excel(csv_signal_nums)['numbers']
    plasma_nums = excel_dicts['noise']
    print('Plasma nums are:', plasma_nums)
    e_pl_integs = np.zeros(len(plasma_nums))
    pl_ds = np.zeros(len(plasma_nums))
    pl_nums = np.zeros(len(plasma_nums))
    if left_right:
        power_koefs = np.zeros(len(plasma_nums))
        r_amp_ints = np.zeros(len(plasma_nums))
        l_amp_ints = np.zeros(len(plasma_nums))
    for k, p_num in enumerate(plasma_nums):
        for csv_signal in csv_signals:
            csv_num = csv_signal[3:6]
            if csv_num == p_num:
                pl_nums[k] = int(p_num)
                file = fft_test.open_file(csv_signal, reduced=True)
                plasma_time = file['time']
                plasma_voltage = file['voltage']
                dt = file['time_resolution']
                n_dt = plasma_time[-1] - plasma_time[0]

                if type == 'e_time':
                    e_pl_integ = fft_test.e_square(plasma_time, plasma_voltage)
                if type == 'amp_freq':
                    pl_fft = fft_test.fft_amplitude(plasma_time, plasma_voltage, dt)
                    pl_fr = pl_fft['frequency']
                    pl_amp = pl_fft['amplitude']
                    e_pl_integ = 2 * (n_dt**2) * fft_test.e_square(pl_fr, pl_amp)
                    if left_right:
                        left_inds = pl_fr < 2.725e9
                        l_freqs = pl_fr[left_inds]
                        l_amps = pl_amp[left_inds]
                        l_amp_integ = 2 * (n_dt**2) * fft_test.e_square(l_freqs, l_amps)

                        right_inds = pl_fr > 2.755e9
                        r_freqs = pl_fr[right_inds]
                        r_amps = pl_amp[right_inds]
                        r_amp_integ = 2 * (n_dt**2) * fft_test.e_square(r_freqs, r_amps)

                        r_amp_ints[k] = r_amp_integ
                        l_amp_ints[k] = l_amp_integ
                        power_koefs[k] = r_amp_integ / l_amp_integ
                #e_pl_integ = fft_test.e_square(red_time, red_volt)
                e_pl_integs[k] = e_pl_integ
                pl_ds[k] = fft_test.read_excel(csv_signal_nums)['dicts'][p_num]['Ток плазмы, А']
    pl_ind_sort = np.argsort(pl_ds)
    e_pl_integs = e_pl_integs[pl_ind_sort]
    pl_ds = pl_ds[pl_ind_sort]
    power_koefs = power_koefs[pl_ind_sort]
    r_amp_ints = r_amp_ints[pl_ind_sort]
    l_amp_ints = l_amp_ints[pl_ind_sort]
    plasma_nums = pl_nums[pl_ind_sort]
    pl_plot_dict = {'e_integs': e_pl_integs,
                    'pl_ds': pl_ds,
                    'koefs': power_koefs,
                    'r_power': r_amp_ints,
                    'l_power': l_amp_ints,
                    'plasma_nums': plasma_nums}
    return pl_plot_dict

def left_right_noise():
    test = ProcessSignal('191106')
    noise_dict = noise_signal_plot(191106, type='amp_freq', left_right=True)
    pl_d = noise_dict['pl_ds']
    r_power = noise_dict['r_power']
    l_power = noise_dict['l_power']
    full_power = noise_dict['e_integs']
    plasma_nums = noise_dict['plasma_nums']
    fig = plt.figure(num=1, dpi=300)
    ax = fig.add_subplot(111)
    line_r, = ax.plot(pl_d, r_power, marker='^', linewidth=0.9, color='blue', ms=3)
    line_l, = ax.plot(pl_d, l_power, marker='o', linewidth=0.9, color='red', ms=3)
    line_l.set_label(r'$f < 2.725\/GHz$')
    line_r.set_label(r'$f > 2.755\/GHz$')
    ax.set_xlabel('Plasma density, arb.units', fontsize=12)
    ax.set_ylabel(r'$K\int A^2 df $', fontsize=12)
    ax.set_ylim(bottom=0)
    ax.legend()
    ax.set_title(r'$191106. P_{in}=0 $')
    ax.grid(which='both', axis='both')
    png_name = test.pics_path / 'noise_power'
    fig.savefig(png_name)
    test.left_right_table(plasma_nums, full_power, l_power,
                          r_power, pl_d, 'noise')
#left_right_noise()

def two_noise_picks():
    dict_1106 = noise_signal_plot(191106)
    d_1106 = dict_1106['pl_ds']
    e_squares_1106 = dict_1106['e_integs']
    dict_1011 = noise_signal_plot(191011)
    d_1011 = dict_1011['pl_ds']
    e_squares_1011 = dict_1011['e_integs']
    fig = plt.figure(num=1, dpi=300)
    ax = fig.add_subplot(111)
    line_1, = ax.plot(d_1011, e_squares_1011)
    line_2, = ax.plot(d_1106, e_squares_1106)
    line_1.set_label('191011')
    line_2.set_label('191106')
    ax.legend()
    ax.grid(which='both', axis='both')
    png_path = r'C:\Users\d_Nice\Documents\SignalProcessing\2019'
    png_name = png_path + '/ 191011_191106'
    fig.savefig(png_name)
#two_noise_picks()

def integral_plots(exp, type='integrals', table=True):
    fft_test = ProcessSignal(exp)
    csv_types = fft_test.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    print(csv_signals)
    excel_dicts = fft_test.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_dicts['magnetron']
    print('All magnetron nums:', magnetron_nums)

    m_1_nums = magnetron_nums
    m_2_nums = []
    '''
    for num in magnetron_nums:
        if 11 <= int(num) < 67:
            m_1_nums.append(num)
        elif 125 <= int(num) <= 184:
            m_2_nums.append(num)
    '''
    m_lists = [m_1_nums, m_2_nums]

    labels_list = ['№2, 3 (четверть)', '№2, 3, A']

    for j, list in enumerate(m_lists):
        if len(list) != 0:
            fig = plt.figure(num=1, dpi=200, figsize=[11.69, 8.27])
            ax = fig.add_subplot(111)
            flist_len = len(list)
            print('We have {} nums'.format(flist_len))
            print('Magnetron nums are:', list)
            pl_densities = np.zeros(flist_len)
            e_noise_ints = np.zeros(flist_len)
            e_peak_ints = np.zeros(flist_len)
            csv_nums = np.zeros(flist_len)
            magnetron_nums = np.zeros(flist_len)
            e_noise_ls = []
            e_noise_rs = []

            power_koefs = np.zeros(flist_len)
            for i, num in enumerate(list):
                for csv_signal in csv_signals:
                    csv_num = csv_signal[3:6]
                    if csv_num == num:
                        magnetron_nums[i] = int(num)
                        print('I am working on {} signal'.format(csv_num))
                        pl_density = fft_test.read_excel(csv_signal_nums)['dicts'][num]['Ток плазмы, А']
                        pl_densities[i] = pl_density
                        csv_nums[i] = csv_num
                        file = fft_test.open_file(csv_signal, reduced=True)
                        time = file['time']
                        voltage = file['voltage']
                        dt = file['time_resolution']
                        n_dt = time[-1] - time[0]
                        n_dt_2 = n_dt ** 2

                        fft = fft_test.fft_amplitude(time, voltage, dt)
                        freqs = fft['frequency']
                        amps = fft['amplitude']
                        noise_inds_left = freqs < 2.725e9
                        noise_inds_right = freqs > 2.755e9
                        peak_inds = np.logical_and(freqs > 2.725e9, freqs < 2.755e9)

                        freq_noise_l = freqs[noise_inds_left]
                        amp_noise_l = amps[noise_inds_left]
                        freq_noise_r = freqs[noise_inds_right]
                        amp_noise_r = amps[noise_inds_right]
                        freq_peak = freqs[peak_inds]
                        amp_peak = amps[peak_inds]

                        amp_noise_l = 2 * n_dt_2 * fft_test.e_square(freq_noise_l, amp_noise_l)
                        amp_noise_r = 2 * n_dt_2 * fft_test.e_square(freq_noise_r, amp_noise_r)
                        amp_noise = amp_noise_l + amp_noise_r
                        amp_peak = 2 * n_dt_2 * fft_test.e_square(freq_peak, amp_peak)

                        e_noise_ls.append(amp_noise_l)
                        e_noise_rs.append(amp_noise_r)
                        e_noise_ints[i] = amp_noise
                        e_peak_ints[i] = amp_peak
                        power_koefs[i] = amp_noise_r / amp_noise_l
            print('Creating plasma plot ...')
            pl_plot = noise_signal_plot(int(exp), type='amp_freq', left_right=True)
            pl_ds = pl_plot['pl_ds']
            pl_nums = pl_plot['plasma_nums']
            pl_e_ints = pl_plot['e_integs']
            pl_koefs = pl_plot['koefs']

            ind_sort = np.argsort(pl_densities)
            pl_densities = pl_densities[ind_sort]
            e_noise_ints = e_noise_ints[ind_sort]
            e_peak_ints = e_peak_ints[ind_sort]

            x_min = min(pl_ds[0], pl_densities[0]) - 0.5
            x_max = max(pl_ds[-1], pl_densities[-1]) + 0.5

            if type == 'integrals':
                ax.set_prop_cycle(color=['orange', 'indigo', 'green'])
                line3, = ax.plot(pl_ds, pl_e_ints, marker='D', linewidth=2)
                line3.set_label('Энергия в шумовом импульсе')
                line, = ax.plot(pl_densities, e_peak_ints, marker='o', linewidth=2)
                line.set_label('Энергия в пике')
                line2, = ax.plot(pl_densities, e_noise_ints, marker='^', linewidth=2)
                line2.set_label('Энергия в шумовом пъедестале')
                ax.set_ylabel(r'$K\int A^2 df $', fontsize=16)
                path_png = fft_test.pics_path / 'integral_{}.png'.format(labels_list[j])
                if table:
                    magnetron_nums = magnetron_nums[ind_sort]
                    fft_test.integrals_table(magnetron_nums, pl_densities,
                                             e_peak_ints, e_noise_ints,
                                             pl_nums, pl_ds, pl_e_ints,
                                             labels_list[j])

            if type == 'left_right':
                e_noise_ls = e_noise_ls[ind_sort]
                e_noise_rs = e_noise_rs[ind_sort]
                ax.set_prop_cycle(color=['red', 'blue'])
                line_pl_1, = ax.plot(pl_densities, e_noise_ls, marker='o')
                line_pl_1.set_label(r'$f < 2.725\/GHz$')
                line_pl_2, = ax.plot(pl_densities, e_noise_rs, marker='^')
                line_pl_2.set_label(r'$f > 2.755\/GHz$')
                ax.set_ylabel(r'$K\int  A^2 df $', fontsize=16)
                path_png = fft_test.pics_path / 'left_right_power{}.png'.format(labels_list[j])
                if table:
                    magnetron_nums = magnetron_nums[ind_sort]
                    fft_test.left_right_table(magnetron_nums, e_noise_ints, e_noise_ls,
                                              e_noise_rs, pl_densities, labels_list[j])

            if type == 'left_right_koef':
                power_koefs = power_koefs[ind_sort]
                line, = ax.plot(pl_densities, power_koefs, marker='o', linewidth=2)
                line.set_label(r'$ P_{in} \ne 0 $')
                line1, = ax.plot(pl_ds, pl_koefs, marker='^', linewidth=2)
                line1.set_label(r'$P_{in} = 0$')
                ax.set_ylabel(r'$\frac{W(f>2.755\/ГГц)}{W(f<2.725\/ГГц)} $', fontsize=22)
                path_png = fft_test.pics_path / 'power_koef_{}.png'.format(labels_list[j])

            ax.set_ylim(bottom=0)
            ax.set_xlim(left=x_min, right=x_max)
            ax.legend(loc='best', fontsize=12)
            ax.grid(which='both', axis='both')
            ax.set_xlabel(r'$Plasma\/density, arb.units $', fontsize=16)
            ax.set_title('{}, {}'.format(exp, labels_list[j]), fontsize=20)
            fig.savefig(path_png)
            plt.close(fig)

#integral_plots(exp='191120', type='integrals', table=False)

def left_right_power():
    test = ProcessSignal('191011')
    csv_types = test.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = test.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_dicts['magnetron']
    print('All magnetron nums:', magnetron_nums)

    m_1_nums = []
    m_2_nums = []

    for num in magnetron_nums:
        if 66 <= int(num) < 112:
            m_1_nums.append(num)
        elif 112 <= int(num) < 152:
            m_2_nums.append(num)
    m_lists = [m_1_nums, m_2_nums]
    labels_list = ['№2, 3', '№2, 3, A']

    for i, m_list in enumerate(m_lists):
        e_left = []
        e_right = []
        pl_d = []
        fig = plt.figure(num=1, dpi=200, figsize=[11.69, 8.27])
        ax = fig.add_subplot(111)
        ax.set_prop_cycle(color=['orange', 'indigo'])
        for signal in csv_signals:
            num = signal[3:6]
            if num in m_list:
                file = test.open_file(signal, reduced=True)
                u = file['voltage']
                t = file['time']
                dt = file['time_resolution']
                u_highpass = test.highpass(t, u, 2.769e9)
                u_lowpass = test.lowpass(t, u, 2.706e9)

                fft_pass = test.fft_signal(t, u, dt)
                freqs = fft_pass['frequency']
                amps = fft_pass['amplitude']
                high_inds = freqs > 2.769e9
                low_inds = freqs < 2.709e9

                high_freqs = freqs[high_inds]
                low_freqs = freqs[low_inds]

                high_amps = amps[high_inds]
                low_amps = amps[low_inds]

                e_high = test.e_square(high_freqs, high_amps)
                e_low = test.e_square(low_freqs, low_amps)
                '''
                e_high = test.e_square(t, u_highpass)
                e_low = test.e_square(t, u_lowpass)
                '''
                pl_density = test.read_excel(csv_signal_nums)['dicts'][num]['Ток плазмы, А']

                #plt.plot(fft_high['frequency'], fft_high['amplitude'], color='blue')
                #plt.plot(high_freqs, high_amps, color='red')
                #plt.plot(low_freqs, low_amps, color='k')
                e_left.append(e_low)
                e_right.append(e_high)
                pl_d.append(pl_density)
                #plt.show()
        pl_d = np.array(pl_d)
        inds = np.argsort(pl_d)
        pl_d = pl_d[inds]
        e_left = np.array(e_left)[inds]

        line1, = ax.plot(pl_d, e_left, marker='o')
        line1.set_label('f < 2,72 GHz')
        line2, = ax.plot(pl_d, e_right, marker='o')
        line2.set_label('f > 2,76 GHz')

        ax.set_ylim(bottom=0)
        ax.set_xlim(left=5)
        ax.set_xlabel(r'$Plasma\/density, arb.units $', fontsize=16)
        ax.set_ylabel(r'$\int E^2 dt $', fontsize=16)
        ax.legend(loc='best', fontsize=12)
        ax.grid(which='both', axis='both')
        ax.set_title('191011, {}'.format(labels_list[i]), fontsize=20)
        mean_png = test.pics_path / 'left_right_power{}.png'.format(i)
        fig.savefig(mean_png)
        plt.close(fig)
        plt.show()

#left_right_power()
'''   
    ax.legend(loc='best', fontsize=16)
    ax.set_title('190925', fontsize=20)
    png_name = fft_test.pics_path / 'e_square.png'
    fig.savefig(png_name)
    plt.show()
    fig2 = plt.figure(num=2, dpi=200, figsize=[11.69, 8.27])
    ax2 = fig2.add_subplot(111)
    ax2.plot(absorbers, e_mean_list, marker='o')
    ax2.set_xlim(left=1)
    ax2.set_xticks([1, 2, 3, 4])
    ax2.set_xlabel(r'$Absorbers$', fontsize=14)
    ax2.set_ylabel(r'$<\int E^2 dt >$', fontsize=14)
    ax2.grid(which='both', axis='both')
    ax2.set_title('190925', fontsize=20)
    mean_png = fft_test.pics_path / 'e_square_mean.png'
    fig2.savefig(mean_png)
    plt.show()
    '''


def density_stats():
    exp_nums = ['191106', '191120']
    n_pl = []
    mean_fr = []
    for exp_num in exp_nums:
        print('Experiment number {}'.format(exp_num))
        test = ProcessSignal(exp_num)
        if 'reb':
            path_dir = test.fft_reb
        if 'noise_base':
            path_dir = test.fft_noise_base
        wb = xl.load_workbook(doc)
        sheet1 = wb['Full FFT']
        rows = sheet1.max_row
        for i in range(1, rows):
            cell_f = sheet1.cell(row=i+1, column=2)
            cell_n = sheet1.cell(row=i+1, column=3)
            mean_fr.append(cell_f.value)
            n_pl.append(cell_n.value)
    n_pl = np.array(n_pl)
    mean_fr = np.array(mean_fr)

    sorted_inds = np.argsort(n_pl)
    n_pl = n_pl[sorted_inds]
    mean_fr = mean_fr[sorted_inds]

    n_uniq = np.unique(n_pl)
    n_stats = []
    fr_stats = []
    for n_u in n_uniq:
        fr_stat_els = []
        for i, n in enumerate(n_pl):
            if n == n_u:
                fr_stat_els.append(mean_fr[i])
            fr_els_mean = np.round(np.mean(np.array(fr_stat_els)), 2)
        n_stats.append(n_u)
        fr_stats.append(fr_els_mean)
    print('n_pl:', n_stats)
    print('f_stats:', fr_stats)

    test_part = ProcessSignal('190627')
    doc_part = test_part.fft_excel_file_path
    wb = xl.load_workbook(doc_part)
    sheet_part = wb['Part FFT']
    rows = sheet_part.max_row
    fr_1 = []
    fr_2 = []
    fr_3 = []
    for i in range(1, rows):
        cell_f1 = sheet_part.cell(row=i + 1, column=2)
        cell_f2 = sheet_part.cell(row=i + 1, column=3)
        cell_f3 = sheet_part.cell(row=i + 1, column=4)
        fr_1.append(cell_f1.value)
        fr_2.append(cell_f2.value)
        fr_3.append(cell_f3.value)
    fr_1 = np.array(fr_1)
    print(fr_1)
    fr_2 = np.array(fr_2)
    fr_3 = np.array(fr_3)
    df = 0.02
    n_1 = np.zeros(fr_1.size)
    for i, fr in enumerate(fr_1):
        for j, fr_mean in enumerate(fr_stats):
            if fr >= (fr_mean - df) and fr <= (fr_mean + df):
                print('fr_mean = ({}: {})'.format(fr_mean - df, fr_mean + df))
                print('f_1 =', fr, 'fr_mean =', fr_mean)
                n_1[i] = n_pl[j]
    print('n_1:', n_1)
#density_stats()

def mean_freq_plasma():
    test = ProcessSignal('191001')
    doc = test.excel_folder_path / 'noise_fft.xlsx'
    wb = xl.load_workbook(doc)
    sheet1 = wb['Full FFT']
    sheet3 = wb['Noise FFT']
    rows = sheet1.max_row
    rows_noise = sheet3.max_row

    n_pl_noise = []
    mean_fr_noise = []
    for j in range(1, rows_noise):
        cell_f = sheet3.cell(row=j + 1, column=2)
        cell_n = sheet3.cell(row=j + 1, column=3)
        mean_fr_noise.append(cell_f.value)
        n_pl_noise.append(cell_n.value)

    n_pl_1 = []
    mean_fr_1 = []
    n_pl_2 = []
    mean_fr_2 = []
    n_pl_3 = []
    mean_fr_3 = []
    labels_list = ['№2, 3', '№2, 3, A, Б, В', '№2, 3, A', 'Шум']

    for i in range(1, rows):
        cell_num = sheet1.cell(row=i + 1, column=1)
        num = cell_num.value
        cell_f = sheet1.cell(row=i + 1, column=2)
        cell_n = sheet1.cell(row=i + 1, column=3)
        if 35 <= int(num) < 79:
            mean_fr_1.append(cell_f.value)
            n_pl_1.append(cell_n.value)
        elif 79 <= int(num) < 111:
            mean_fr_2.append(cell_f.value)
            n_pl_2.append(cell_n.value)
        elif 127 <= int(num) <= 150:
            mean_fr_3.append(cell_f.value)
            n_pl_3.append(cell_n.value)

    n_pl = [n_pl_1, n_pl_2, n_pl_3, n_pl_noise]
    mean_fr = [mean_fr_1, mean_fr_2, mean_fr_3, mean_fr_noise]
    fig = plt.figure(num=1, dpi=200, figsize=[11.69, 8.27])
    ax = fig.add_subplot(111)
    ax.set_prop_cycle(color=['blue', 'green', 'deepskyblue', 'orange'])
    for i in range(4):
        n_pl[i] = np.array(n_pl[i])
        mean_fr[i] = np.array(mean_fr[i])
        sorted_inds = np.argsort(n_pl[i])
        n_pl[i] = n_pl[i][sorted_inds]
        mean_fr[i] = mean_fr[i][sorted_inds]
        line,  = plt.plot(n_pl[i], mean_fr[i], marker='o', linewidth=2)
        line.set_label('{}'.format(labels_list[i]))
    ax.hlines(2.739, xmin=2, xmax=13)
    ax.set_ylim(bottom=1.5)
    ax.set_xlim(left=2, right=11)
    ax.set_xlabel(r'$Plasma\/density, arb.units $', fontsize=14)
    ax.set_ylabel(r'$Mean\/frequency, GHz$', fontsize=14)
    ax.grid(which='both', axis='both')
    ax.legend(loc='best', fontsize=16)
    ax.set_title('Mean frequency', fontsize=20)
    png_name = test.pics_path / 'mean_freq.png'
    plt.savefig(png_name)
    plt.show()
#mean_freq_plasma()

def amp_max():
    test = ProcessSignal('191001')
    csv_types = test.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = test.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_dicts['magnetron']
    m_1_nums = []
    m_2_nums = []
    m_3_nums = []
    for num in magnetron_nums:
        if 35 <= int(num) < 79:
            m_1_nums.append(num)
        elif 79 <= int(num) < 111:
            m_2_nums.append(num)
        elif 127 <= int(num) <= 150:
            m_3_nums.append(num)
    m_lists = [m_1_nums, m_2_nums, m_3_nums]
    labels_list = ['№2, 3', '№2, 3, A, Б, В', '№2, 3, A', 'Шум']
    fig = plt.figure(num=1, dpi=200, figsize=[11.69, 8.27])
    ax = fig.add_subplot(111)
    ax.set_prop_cycle(color=['blue', 'green', 'deepskyblue'])
    for i, list in enumerate(m_lists):
        u_maxs = []
        pl_ds = []
        for csv_signal in csv_signals:
            csv_num = csv_signal[3:6]
            for num in list:
                if csv_num == num:
                    file = test.open_file(csv_signal)
                    u = file['voltage']
                    t = file['time']
                    red_dict = test.proc_part_130_290(t, u)
                    red_u = red_dict['part_u']
                    red_t = red_dict['part_t']
                    u_filt = test.bandpass_filter(red_t, red_u, 2.709e9, 2.769e9)
                    u_max = np.max(np.array(u_filt))
                    pl_d = test.read_excel(csv_signal_nums)['dicts'][num]['Ток плазмы, А']
                    u_maxs.append(u_max)
                    pl_ds.append(pl_d)
        sorted_inds = np.argsort(np.array(pl_ds))
        pl_ds = np.array(pl_ds)
        pl_ds = pl_ds[sorted_inds]
        u_maxs = np.array(u_maxs)
        u_maxs = u_maxs[sorted_inds]
        line, = plt.plot(pl_ds, u_maxs, marker='o', linewidth=2)
        line.set_label('{}'.format(labels_list[i]))
    ax.set_ylim(bottom=0)
    ax.set_xlim(left=4, right=9)
    ax.set_xlabel(r'$Plasma\/density, arb.units $', fontsize=14)
    ax.set_ylabel(r'$Max\/amplitude$', fontsize=14)
    ax.grid(which='both', axis='both')
    ax.legend(loc='best', fontsize=16)
    ax.set_title('Max amplitude', fontsize=20)
    png_name = test.pics_path / 'max_amp.png'
    plt.savefig(png_name)
    plt.show()
#amp_max()

def oscillograms():
    test = ProcessSignal('191106')
    csv_types = test.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = test.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_dicts['magnetron']
    for signal in csv_signals:
        file = test.open_file(signal, reduced=True)
        u = file['voltage']
        t = file['time']
        dt = file['time_resolution']
        u_filt = test.bandpass_filter(t, u, 2.725e9, 2.755e9)
        num = signal[3:6]
        if int(num) < 66:
            absorbers = '№2, 3'
        elif 66 <= int(num) < 125 and num in magnetron_nums:
            absorbers = 0
        elif 125 <= int(num) <= 184 and num in magnetron_nums:
            absorbers = '№2, 3, A'
        else:
            absorbers = 0
        pl_d_nums = test.read_excel(csv_signal_nums)['dicts'].keys()
        if num in pl_d_nums:
            pl_density = test.read_excel(csv_signal_nums)['dicts'][num]['Ток плазмы, А']
            test.oscill_picture(num, t, u, u_filt, pl_density, absorbers, save=True)
#oscillograms()

def pulse_duration():
    test = ProcessSignal('190925')
    csv_types = test.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    pl_d_nums = test.read_excel(csv_signal_nums)['dicts'].keys()
    t_difs = []
    pl_ds = []
    for signal in csv_signals:
        file = test.open_file(signal)
        u = file['voltage']
        t = file['time']
        dt = file['time_resolution']
        envelope = test.useful_part(t, u, dt)
        t_use, u_use = envelope['signal_time'], envelope['signal_voltage']
        t_dif = t_use[-1] - t_use[0]
        t_difs.append(t_dif)
        num = signal[3:6]
        if num in pl_d_nums:
            pl_density = test.read_excel(csv_signal_nums)['dicts'][num]['Ток плазмы, А']
            pl_ds.append(pl_density)
    pl_ds = np.array(pl_ds)
    t_difs = np.array(t_difs)
    sorted_inds = np.argsort(pl_ds)
    pl_ds_sorted = pl_ds[sorted_inds]
    t_difs_sorted = t_difs[sorted_inds]
    plt.plot(pl_ds_sorted, t_difs_sorted)
    #plt.plot(t_use, u_use)
    plt.show()
#pulse_duration()


