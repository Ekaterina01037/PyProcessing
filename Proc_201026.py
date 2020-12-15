import numpy as np
from scipy.fftpack import rfft, irfft, rfftfreq
from ProcessClass_10 import ProcessSignal
import matplotlib.pyplot as plt
import os
import openpyxl as excel

proc = ProcessSignal('201026')
short_delay_signals = np.arange(29, 67, 2)
long_delay_signals = np.arange(85, 109, 2)
signal_list = [short_delay_signals, long_delay_signals]
print(long_delay_signals)

full_ints = []
peak_ints = []
ns = []
m_delays = []
nums = []

for signal in signal_list[1]:
    signal_file = f'str{signal:03d}.csv'
    file = proc.open_file(signal_file, reduced=False)
    t, u = file['time'], file['voltage']
    if signal < 95:
        t_min, t_max = 960e-9, 1222e-9
    else:
        t_min, t_max = 60e-9, 322e-9
    cut_inds = np.logical_and(t >= t_min, t <= t_max)
    t_cut = t[cut_inds]
    u_cut = u[cut_inds]
    print(t[0])
    dt_2 = (t_cut[-1] - t_cut[0]) ** 2
    u_filt = proc.bandpass_filter(t, u, 2.68e9, 2.74e9)
    '''
    plt.plot(t, u)
    plt.plot(t_cut, u_cut)
    plt.plot(t, u_filt)
    plt.title(signal)
    plt.show()
    '''
    integral = np.round(proc.e_square(t_cut, u_cut) / 1e-8, 3)
    freqs, amps = proc.fft_amplitude(t_cut, u_cut)['frequency'], proc.fft_amplitude(t_cut, u_cut)['amplitude']
    peak_inds = np.logical_and(freqs >= 2.695e9, freqs <= 2.725e9)
    p_freqs, p_amps = freqs[peak_inds], amps[peak_inds]

    plt.plot(freqs, amps)
    plt.title(signal)
    plt.xlim(left=0e9, right=4e9)
    plt.show()

    full_int = np.round(dt_2 * proc.e_square(freqs, amps) / 2e-8, 3)
    peak_int = np.round(dt_2 * proc.e_square(p_freqs, p_amps) / 2e-8, 3)
    print('num =', signal, integral, full_int)
    n = proc.read_excel(f'{signal:03d}')['dicts'][f'{signal:03d}']['Ток плазмы, А']
    m_delay = proc.read_excel(f'{signal:03d}')['dicts'][f'{signal:03d}']['Задержка магнетрона, нс']

    full_ints.append(full_int)
    peak_ints.append(peak_int)
    ns.append(n)
    m_delays.append(m_delay)
    nums.append(signal)

ind_sort = np.argsort(np.asarray(ns))
ns = np.asarray(ns)[ind_sort]
m_delays = np.asarray(m_delays)[ind_sort]
nums = np.asarray(nums)[ind_sort]
full_ints, peak_ints = np.asarray(full_ints)[ind_sort], np.asarray(peak_ints)[ind_sort]
ex_table = excel.Workbook()
ex_table.create_sheet(title='Integral', index=0)
sheet = ex_table['Integral']
sheet['A1'] = 'Номер'
sheet['B1'] = 'Плотность плазмы, отн.ед.'
sheet['C1'] = 'Полный интеграл, *10-7'
sheet['D1'] = 'Интеграл в пике, *10-7'
table_size = full_ints.size
for k in range(table_size):
    cell = sheet.cell(row=k + 2, column=1)
    cell.value = nums[k]
    cell = sheet.cell(row=k + 2, column=2)
    cell.value = ns[k]
    cell = sheet.cell(row=k + 2, column=3)
    cell.value = full_ints[k]
    cell = sheet.cell(row=k + 2, column=4)
    cell.value = peak_ints[k]
path = r'C:\Users\d_Nice\Documents\SignalProcessing\2020\201026\Excel\Integrals_201026_2.xlsx'
ex_table.save(path)
