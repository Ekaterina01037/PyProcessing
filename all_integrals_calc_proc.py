import numpy as np
import openpyxl as excel
import matplotlib.pyplot as plt
from ProcessClass_10 import ProcessSignal
from pathlib import Path
from openpyxl.utils import get_column_letter
from scipy.stats import tstd


def magnetron_integrals_bloc(exp_num, file_nums, central_freq=2.714, band_half_width=0.05):
    proc = ProcessSignal(str(exp_num))
    shot_nums, density_vals, full_ints, peak_ints, exp_nums = [], [], [], [], []
    for num in file_nums:
        file_name = f'str{num}.csv'
        file = proc.open_file(file_name, reduced=True)
        u = file['voltage']
        t = file['time']
        f_low = (central_freq - band_half_width) * 1e9
        f_high = (central_freq + band_half_width) * 1e9
        dt_2 = (t[-1] - t[0]) ** 2
        #pl_num = f'{(int(num) - 1):03d}'
        pl_density = proc.read_excel(file_name)['dicts'][num]['Ток плазмы, А']

        freqs, amps = proc.fft_amplitude(t, u)['frequency'], proc.fft_amplitude(t, u)['amplitude']
        peak_inds = np.logical_and(freqs >= f_low, freqs <= f_high)
        p_freqs, p_amps = freqs[peak_inds], amps[peak_inds]

        full_int = np.round(dt_2 * proc.e_square(freqs, amps) / 2e-8, 3)
        peak_int = np.round(dt_2 * proc.e_square(p_freqs, p_amps) / 2e-8, 3)

        density_vals.append(pl_density)
        shot_nums.append(num)
        full_ints.append(full_int)
        peak_ints.append(peak_int)
        exp_nums.append(exp_num)

    sorted_inds = np.argsort(density_vals)
    density_vals = np.asarray(density_vals)[sorted_inds]
    shot_nums = np.asarray(shot_nums)[sorted_inds]
    full_ints, peak_ints = np.asarray(full_ints)[sorted_inds], np.asarray(peak_ints)[sorted_inds]

    integrals_dict = {'nums': shot_nums,
                      'density': density_vals,
                      'full_ints': full_ints,
                      'peak_ints': peak_ints,
                      'exp_nums': exp_nums}
    return integrals_dict


def noise_integrals_block(exp_num, file_nums):
    proc = ProcessSignal(str(exp_num))
    shot_nums, density_vals, full_ints = [], [], []
    for num in file_nums:
        file_name = f'str{num}.csv'
        file = proc.open_file(file_name, reduced=True)
        u = file['voltage']
        t = file['time']
        dt_2 = (t[-1] - t[0]) ** 2

        pl_density = proc.read_excel(file_name)['dicts'][num]['Ток плазмы, А']

        freqs, amps = proc.fft_amplitude(t, u)['frequency'], proc.fft_amplitude(t, u)['amplitude']
        full_int = np.round(dt_2 * proc.e_square(freqs, amps) / 2e-8, 3)

        density_vals.append(pl_density)
        shot_nums.append(num)
        full_ints.append(full_int)

    sorted_inds = np.argsort(density_vals)
    density_vals = np.asarray(density_vals)[sorted_inds]
    shot_nums = np.asarray(shot_nums)[sorted_inds]
    full_ints = np.asarray(full_ints)[sorted_inds]
    integrals_dict = {'nums': shot_nums,
                      'density': density_vals,
                      'full_ints': full_ints}
    return integrals_dict


def fill_excel_table(exp_num, dict_list, proc):
    ex_table = excel.Workbook()
    ex_table.create_sheet(title='Integral', index=0)
    sheet = ex_table['Integral']
    letter_ind = 1
    for dict in dict_list:
        nums, density_vals, full_integrals, peak_integrals = dict['nums'], dict['density'], dict['full_ints'], dict['peak_ints']
        sheet[f'{get_column_letter(letter_ind)}1'] = 'Номер'
        sheet[f'{get_column_letter(letter_ind+1)}1'] = 'Плотность плазмы, отн.ед.'
        sheet[f'{get_column_letter(letter_ind+2)}1'] = 'W_f0, *10-8'
        sheet[f'{get_column_letter(letter_ind+3)}1'] = 'W_1, *10-8'

        for k in range(full_integrals.size):
            cell = sheet.cell(row=k + 2, column=letter_ind)
            cell.value = int(nums[k])
            cell = sheet.cell(row=k + 2, column=letter_ind+1)
            cell.value = density_vals[k]
            cell = sheet.cell(row=k + 2, column=letter_ind+2)
            cell.value = peak_integrals[k]
            cell = sheet.cell(row=k + 2, column=letter_ind+3)
            cell.value = full_integrals[k] - peak_integrals[k]
        letter_ind += 5
        path = proc.excel_folder_path / f'Integrals_{exp_num}.xlsx'
        ex_table.save(path)


def energy_4_experiments():
    exps = [210119, 210204, 210211, 210304]
    l_col = [19.5, 26.5, 34.5, 20.5]
    n_low, n_max = 9, 11
    fig = plt.figure(num=1, dpi=300)
    ax1 = fig.add_subplot(211)
    ax2 = fig.add_subplot(212)
    colors = ['orange', 'mediumblue', 'limegreen', 'red']
    for i, exp in enumerate(exps):
        proc = ProcessSignal(str(exp))
        csv_types = proc.read_type_file()
        csv_signals = csv_types['signal_files']
        csv_signal_nums = csv_types['signal_nums']
        excel_dicts = proc.read_excel(csv_signal_nums)['numbers']
        magnetron_nums = excel_dicts['magnetron']
        magnetron_dict = magnetron_integrals_bloc(exp, magnetron_nums)
        nums = magnetron_dict['nums']
        dens_vals = magnetron_dict['density']
        noise = magnetron_dict['full_ints'] - magnetron_dict['peak_ints']
        peak = magnetron_dict['peak_ints']
        noise_to_full = noise / magnetron_dict['full_ints']
        exp_nums = magnetron_dict['exp_nums']

        ind_mask = np.logical_and(dens_vals >= n_low, dens_vals <= n_max)
        dens_vals, noise_to_full, exp_num = dens_vals[ind_mask], noise_to_full[ind_mask], exp_nums[0]
        peak = peak[ind_mask]
        print(f'Creating plot for experiment {exp}...')
        col_len = [l_col[i]] * noise_to_full.size
        print(col_len)
        ax2.plot(col_len, peak,  marker='.', linestyle=' ', color=colors[i])
        ax1.plot(col_len, noise_to_full, marker='.', linestyle=' ', color=colors[i])

    ax1.grid(which='both', axis='both')
    ax1.set_ylabel(r'$W_{1} / W$')
    ax1.set_xlim(left=17)
    ax2.grid(which='both', axis='both')
    ax2.set_xlabel(r'$L_{кол}, см$')
    ax2.set_ylabel(r'$W_{f0}$')
    ax2.set_xlim(left=17)
    ax1.set_title(r' Плотности {} - {} отн.ед.'.format(n_low, n_max))
    fig_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\210423\Pictures')
    png_name = fig_path / f'4_exp_energy(17)'
    fig.savefig(png_name)
    plt.close(fig)
    plt.show()

#energy_4_experiments()


def multi_integral_excel(exp_num):
    proc = ProcessSignal(str(exp_num))
    csv_types = proc.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = proc.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_dicts['magnetron']
    magnetron_nums = [csv_signal_nums[i] for i in range(len(csv_signal_nums))if int(csv_signal_nums[i]) % 2 == 0]
    print(magnetron_nums)
    list_1 = [magnetron_nums[i] for i in range(len(magnetron_nums)) if 6 < int(magnetron_nums[i]) < 31 ]
    list_2 = [magnetron_nums[i] for i in range(len(magnetron_nums)) if 30 < int(magnetron_nums[i]) < 49]
    list_3 = [magnetron_nums[i] for i in range(len(magnetron_nums)) if
              48 < int(magnetron_nums[i]) < 67]
    list_4 = [magnetron_nums[i] for i in range(len(magnetron_nums)) if
              66 < int(magnetron_nums[i]) < 89]
    list_5 = [magnetron_nums[i] for i in range(len(magnetron_nums)) if
              88 < int(magnetron_nums[i]) < 107]
    list_6 = [magnetron_nums[i] for i in range(len(magnetron_nums)) if
              106 < int(magnetron_nums[i]) < 123]
    list_7 = [magnetron_nums[i] for i in range(len(magnetron_nums)) if
              122 < int(magnetron_nums[i]) < 136]
    lists = [list_1, list_2, list_3, list_4, list_5, list_6, list_7]
    dict_list = [magnetron_integrals_bloc(exp_num, lists[i]) for i in range(len(lists))]
    fill_excel_table(exp_num, dict_list, proc)


#multi_integral_excel(210609)


def magnetron_exp(exp_num):
    proc = ProcessSignal(str(exp_num))
    csv_types = proc.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = proc.read_excel(csv_signal_nums)['numbers']
    magnetron_nums = excel_dicts['magnetron']
    lists = [magnetron_nums]
    dict_list = [magnetron_integrals_bloc(exp_num, lists[i]) for i in range(len(lists))]
    fill_excel_table(exp_num, dict_list, proc)

#magnetron_exp(210708)


def all_integrals_proc(exp_num, nums, table=True, central_freq=2.714, band_half_width=0.05):
    proc = ProcessSignal(str(exp_num))
    csv_types = proc.read_type_file()
    csv_signals = csv_types['signal_files']
    csv_signal_nums = csv_types['signal_nums']
    excel_dicts = proc.read_excel(csv_signal_nums)['numbers']
    noise_nums = [str(int(excel_dicts['noise'][i]) + 1) for i in range(len(excel_dicts['noise'])) if int(excel_dicts['noise'][i]) + 1 in nums]
    magnetron_nums = [str(int(excel_dicts['magnetron'][i]) + 1) for i in range(len(excel_dicts['magnetron'])) if int(excel_dicts['magnetron'][i]) + 1 in nums]
    print('Noise:', noise_nums, '\n', 'Amplifier:', magnetron_nums)

    nums, n_s, full_ints, peak_ints = [], [], [], []
    noise_ints, noise_ns, n_nums = [], [], []
    for signal in csv_signals:
        num = signal[3:6]
        if num in magnetron_nums:
            file = proc.open_file(signal, reduced=True)
            u = file['voltage']
            t = file['time']
            f_low = (central_freq - band_half_width) * 1e9
            f_high = (central_freq + band_half_width) * 1e9
            u_filt = proc.fft_filter(t, u, f_low, f_high)
            dt_2 = (t[-1] - t[0]) ** 2

            pl_density = proc.read_excel(signal)['dicts'][num]['Ток плазмы, А']

            integral = np.round(proc.e_square(t, u) / 1e-8, 3)
            filt_int = np.round(proc.e_square(t, u_filt) / 1e-8, 3)
            freqs, amps = proc.fft_amplitude(t, u)['frequency'], proc.fft_amplitude(t, u)['amplitude']
            peak_inds = np.logical_and(freqs >= f_low, freqs <= f_high)
            p_freqs, p_amps = freqs[peak_inds], amps[peak_inds]

            full_int = np.round(dt_2 * proc.e_square(freqs, amps) / 2e-8, 3)
            peak_int = np.round(dt_2 * proc.e_square(p_freqs, p_amps) / 2e-8, 3)

            n_s.append(pl_density)
            nums.append(num)
            full_ints.append(full_int)
            peak_ints.append(peak_int)

            print('peak_int = ', np.round(integral, 2), 'noise_int =', np.round(integral - filt_int, 2),
                  'noise_fft =', np.round(full_int - peak_int, 2))
        if num in noise_nums:
            file = proc.open_file(signal, reduced=True)
            u = file['voltage']
            t = file['time']
            dt = file['time_resolution']
            f_low = (central_freq - band_half_width) * 1e9
            f_high = (central_freq + band_half_width) * 1e9
            u_filt = proc.fft_filter(t, u, f_low, f_high)
            dt_2 = (t[-1] - t[0]) ** 2

            pl_density = proc.read_excel(signal)['dicts'][num]['Ток плазмы, А']

            freqs, amps = proc.fft_amplitude(t, u)['frequency'], proc.fft_amplitude(t, u)['amplitude']
            noise_int = np.round(dt_2 * proc.e_square(freqs, amps) / 2e-8, 3)
            noise_ints.append(noise_int)
            noise_ns.append(pl_density)
            n_nums.append(num)

    ind_sort = np.argsort(np.asarray(n_s))
    ns_sort = np.asarray(n_s)[ind_sort]
    nums_sort = np.asarray(nums)[ind_sort]
    full_ints_sort, peak_ints_sort = np.asarray(full_ints)[ind_sort], np.asarray(peak_ints)[ind_sort]
    print('Sorted amplifier numbers:', nums_sort)

    n_ind_sort = np.argsort(np.asarray(noise_ns))
    noise_ns_sort, n_nums_sort, noise_ints_sort = np.asarray(noise_ns)[n_ind_sort], np.asarray(n_nums)[n_ind_sort], np.asarray(noise_ints)[n_ind_sort]
    print('Sorted noise numbers:', n_nums_sort)
    if table:
        ex_table = excel.Workbook()
        ex_table.create_sheet(title='Integral', index=0)
        sheet = ex_table['Integral']
        sheet['A1'] = 'Номер(шум)'
        sheet['B1'] = 'Плотность плазмы, отн.ед.'
        sheet['C1'] = 'W_f0, *10-8'
        sheet['D1'] = 'W_1, *10-8'

        sheet['F1'] = 'Номер'
        sheet['G1'] = 'Плотность плазмы, отн.ед.'
        sheet['H1'] = 'W_2, *10-8'

        for z in range(full_ints_sort.size):
            cell = sheet.cell(row=z + 2, column=1)
            cell.value = int(nums_sort[z])
            cell = sheet.cell(row=z + 2, column=2)
            cell.value = ns_sort[z]
            cell = sheet.cell(row=z + 2, column=3)
            cell.value = peak_ints_sort[z]
            cell = sheet.cell(row=z + 2, column=4)
            cell.value = full_ints_sort[z] - peak_ints_sort[z]

        for k in range(noise_ints_sort.size):
            cell = sheet.cell(row=k + 2, column=6)
            cell.value = int(n_nums_sort[k])
            cell = sheet.cell(row=k + 2, column=7)
            cell.value = noise_ns_sort[k]
            cell = sheet.cell(row=k + 2, column=8)
            cell.value = noise_ints_sort[k]

        path = proc.excel_folder_path / f'Integrals_{exp_num}_2_antennas.xlsx'
        ex_table.save(path)
    else:
        noise_dict = {'nums': n_nums_sort,
                      'density_vals': noise_ns_sort,
                      'w_2': noise_ints_sort}
        magnetron_dict = {'nums': nums_sort,
                          'density_vals': ns_sort,
                          'w_f0': peak_ints_sort,
                          'w_1': full_ints_sort - peak_ints_sort}
        return noise_dict, magnetron_dict


#all_integrals_proc(210423, nums=[i for i in range(132, 221, 2)], table=True)

def stats_bonds(density_vals):
    min_round, max_round = np.modf(density_vals[0])[1], np.modf(density_vals[-1])[1]
    if min_round + 0.5 > density_vals[0]:
        first_bond = min_round
    else:
        first_bond = min_round + 0.5
    if max_round + 0.5 > density_vals[-1]:
        last_bond = max_round + 0.5
    else:
        last_bond = max_round + 1
    bonds = np.arange(first_bond, last_bond + 0.5, 0.5)
    return bonds


def calc_stats(bonds, density_vals, integral_1, integral_2=[0]):
    av_densities, std_dens, av_int_1_list, std_int_1_list, av_int_2_list, std_int_2_list = [], [], [], [], [], []
    for i, bond in enumerate(bonds):
        mask = np.logical_and(bonds[i] <= density_vals, density_vals < bonds[i + 1])
        density_vals = density_vals[mask]
        av_int_1 = integral_1[mask]
        av_int_2 = integral_2[mask]
        av_density, std_density = np.mean(density_vals), tstd(density_vals)
        av_int_1, std_int_1 = np.mean(av_int_1), tstd(av_int_1)
        av_int_2, std_int_2 = np.mean(av_int_2), tstd(av_int_2)
        av_densities.append(av_density)
        std_dens.append(std_density)
        av_int_1_list.append(av_int_1_list)
        std_int_1_list.append(std_int_1)
        av_int_2_list.append(av_int_2_list)
        std_int_2_list.append(std_int_2)
    average_dict = {'density_vals': np.asarray(av_densities),
                    'std_dens': np.asarray(std_dens),
                    'int_1': np.asarray(av_int_1_list),
                    'std_1': np.asarray(std_int_1_list),
                    'int_2': np.asarray(av_int_2_list),
                    'std_2': np.asarray(std_int_2_list)}
    return average_dict


def all_integrals_stats():
    noise_dict, magnetron_dict = all_integrals_proc(210423, table=False, start_num=0, end_num=0)
    noise_nums, noise_density_vals, w_2 = noise_dict['nums'], noise_dict['density_vals'], noise_dict['w_2']
    m_nums, m_density_vals, w_f0, w_1 = magnetron_dict['nums'], magnetron_dict['density_vals'], magnetron_dict['w_f0'], magnetron_dict['w_1']
    noise_bonds = stats_bonds(noise_density_vals)
    m_bonds = stats_bonds(m_density_vals)
    av_noise_dict = calc_stats(noise_bonds, noise_density_vals, w_2, np.zeros(len(w_2)))
    av_noise_density, std_noise_dens, av_w_2, std_w_2 = av_noise_dict['density_vals'], av_noise_dict['std_dens'], av_noise_dict['int_1'], av_noise_dict['std_1']
    av_m_dict = calc_stats(m_bonds, m_density_vals, w_f0, w_1)
    av_m_density, std_m_dens, av_w_f0, std_w_f0, av_w_1, std_w_1 = av_m_dict['density_vals'], av_m_dict['std_dens'], av_m_dict['int_1'], av_m_dict['std_1'], av_m_dict['int_2'], av_m_dict['std_2']

    ex_table = excel.Workbook()
    ex_table.create_sheet(title='Integral', index=0)
    sheet = ex_table['Integral']
    sheet['A1'] = '<n>'
    sheet['B1'] = 'delta_n'
    sheet['C1'] = '<W_f0>, *10-8'
    sheet['D1'] = 'delta_W_f0, *10-8'
    sheet['E1'] = '<W_1>, *10-8'
    sheet['F1'] = 'delta_W_1, *10-8'

    sheet['H1'] = '<n>'
    sheet['I1'] = 'delta_n'
    sheet['J1'] = 'W_2, *10-8'
    sheet['K1'] = 'delta_W_2, *10-8'

    for z in range(av_m_density.size):
        cell = sheet.cell(row=z + 2, column=1)
        cell.value = av_m_density[z]
        cell = sheet.cell(row=z + 2, column=2)
        cell.value = std_m_dens[z]
        cell = sheet.cell(row=z + 2, column=3)
        cell.value = av_w_f0[z]
        cell = sheet.cell(row=z + 2, column=4)
        cell.value = std_w_f0[z]
        cell = sheet.cell(row=z + 2, column=5)
        cell.value = av_w_1[z]
        cell = sheet.cell(row=z + 2, column=6)
        cell.value = std_w_1[z]

    for k in range(av_noise_density.size):
        cell = sheet.cell(row=k + 2, column=8)
        cell.value = av_noise_density[k]
        cell = sheet.cell(row=k + 2, column=9)
        cell.value = std_noise_dens[k]
        cell = sheet.cell(row=k + 2, column=10)
        cell.value = av_w_2[k]
        cell = sheet.cell(row=k + 2, column=11)
        cell.value = std_w_2[k]
    path = r'C:\Users\d_Nice\Documents\SignalProcessing\2021\210423\Excel\Integrals_stats.xlsx'
    ex_table.save(path)

#all_integrals_stats()

def read_excel_integrals(exp_nums):
    nums, n_s, integrals_w_0, integrals_w_1 = [], [], [], []
    noise_nums, noise_n_s, integrals_w_2 = [], [], []
    for exp_num in exp_nums:
        excel_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\{}\Excel\Integrals_{}.xlsx'.format(exp_num, exp_num))
        wb = excel.load_workbook(excel_path)
        sheet = wb['Integral']
        rows_22, noise_rows_22 = 25, 27
        rows_26, noise_rows_26 = 29, 32
        if exp_num == 210322:
            rows, noise_rows = rows_22, noise_rows_22
            n_s_list = np.asarray([float(sheet[f'F{i}'].value) - 1.5 for i in range(2, rows)])

        elif exp_num == 210326:
            rows, noise_rows = rows_26, noise_rows_26
            n_s_list = np.asarray([float(sheet[f'F{i}'].value) for i in range(2, rows)])

        nums_list = np.asarray([int(sheet[f'E{i}'].value) for i in range(2, rows)])
        nums = np.concatenate((nums, nums_list))
        #n_s_list = np.asarray([float(sheet[f'F{i}'].value) for i in range(2, rows)])
        n_s = np.concatenate((n_s, n_s_list))
        w_0_list = np.asarray([float(sheet[f'H{i}'].value) for i in range(2, rows)])
        integrals_w_0 = np.concatenate((integrals_w_0, w_0_list))
        w_1_list = np.asarray([float(sheet[f'I{i}'].value) for i in range(2, rows)])
        integrals_w_1 = np.concatenate((integrals_w_1, w_1_list))
        noise_nums_list = np.asarray([int(sheet[f'A{i}'].value) for i in range(2, noise_rows)])
        noise_nums = np.concatenate((noise_nums, noise_nums_list))
        noise_n_s_list = np.asarray([float(sheet[f'B{i}'].value) for i in range(2, noise_rows)])
        noise_n_s = np.concatenate((noise_n_s, noise_n_s_list))
        w_2_list = np.asarray([float(sheet[f'C{i}'].value) for i in range(2, noise_rows)])
        integrals_w_2 = np.concatenate((integrals_w_2, w_2_list))

    sorted_inds = np.argsort(n_s)
    sorted_noise_inds = np.argsort(noise_n_s)

    n_s, nums, integrals_w_0, integrals_w_1 = n_s[sorted_inds], nums[sorted_inds], integrals_w_0[sorted_inds], integrals_w_1[sorted_inds]
    noise_nums, noise_n_s, integrals_w_2 = noise_nums[sorted_noise_inds], noise_n_s[sorted_noise_inds], integrals_w_2[sorted_noise_inds]

    new_table = excel.Workbook()
    new_table.create_sheet(title='Сравнение интегралов', index=0)
    sheet = new_table['Сравнение интегралов']
    sheet['A1'] = 'Номер(шум)'
    sheet['B1'] = 'Плотность плазмы, отн.ед.'
    sheet['C1'] = 'W_2, *10-8'

    sheet['E1'] = 'Номер'
    sheet['F1'] = 'Плотность плазмы, отн.ед.'
    sheet['G1'] = 'Полный интеграл, *10-8'
    sheet['H1'] = 'W_f0, *10-8'
    sheet['I1'] = 'W_1, *10-8'

    for z in range(integrals_w_2.size):
        cell = sheet.cell(row=z + 2, column=1)
        cell.value = int(noise_nums[z])
        cell = sheet.cell(row=z + 2, column=2)
        cell.value = noise_n_s[z]
        cell = sheet.cell(row=z + 2, column=3)
        cell.value = integrals_w_2[z]

    for k in range(integrals_w_0.size):
        cell = sheet.cell(row=k + 2, column=5)
        cell.value = int(nums[k])
        cell = sheet.cell(row=k + 2, column=6)
        cell.value = n_s[k]
        cell = sheet.cell(row=k + 2, column=7)
        cell.value = integrals_w_0[k]
        cell = sheet.cell(row=k + 2, column=8)
        cell.value = integrals_w_1[k]

    result_doc_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\210326\Excel\Integrals_210322_210326(смещение).xlsx')
    new_table.save(result_doc_path)

#read_excel_integrals([210322, 210326])


def integrals_for_different_conditions(exp_num):
    proc = ProcessSignal(str(exp_num))
    wb = excel.load_workbook(proc.main_excel_path)
    sheet = wb['Лист1']
    rows = sheet.max_row
    cols = sheet.max_column
    row_min = 1
    col_max = 1
    for i in range(1, rows):
        cell = sheet.cell(row=i, column=1)
        cell_row_val = cell.value
        if cell_row_val == 'Номер файла' and row_min == 1:
            row_min = i

    for m in range(1, cols + 1):
        cell = sheet.cell(row=row_min, column=m)
        cell_col_val = cell.value
        if cell_col_val == 'Комментарий':
            col_max = m
    print(row_min, col_max)


#integrals_for_different_conditions(210407)