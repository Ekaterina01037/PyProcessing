import numpy as np
import openpyxl as excel
from ProcessClass_10 import ProcessSignal

test = ProcessSignal('191120')
file_name = 'str018.csv'
file = test.open_file(file_name, reduced=True)
t, u, dt = file['time'], file['voltage'], file['time_resolution']
fft = test.fft_amplitude(t, u, dt)
freqs = fft['frequency']
amps = fft['amplitude']
n_dt_2 = (t[-1] - t[0])**2
full_integral = 2 * n_dt_2 * test.e_square(freqs, amps) / 1e-8
print(full_integral)

df_s = np.arange(5, 105, 5) * 1e6
noise_integrals = np.zeros(len(df_s))
peak_integrals = np.zeros(len(df_s))
magnetron_freq = 2.74e9
for i, df in enumerate(df_s):
    low_freq = magnetron_freq - df
    high_freq = magnetron_freq + df

    noise_inds_l = freqs < low_freq
    noise_inds_r = freqs > high_freq
    peak_inds = np.logical_and(freqs > low_freq, freqs < high_freq)

    noise_freqs_l, noise_amps_l = freqs[noise_inds_l], amps[noise_inds_l]
    noise_freqs_r, noise_amps_r = freqs[noise_inds_r], amps[noise_inds_r]
    peak_freqs, peak_amps = freqs[peak_inds], amps[peak_inds]

    noise_integral_l = 2 * n_dt_2 * test.e_square(noise_freqs_l, noise_amps_l) / 1e-8
    noise_integral_r = 2 * n_dt_2 * test.e_square(noise_freqs_r, noise_amps_r) / 1e-8
    #noise_integral = noise_integral_l + noise_integral_r
    peak_integral = 2 * n_dt_2 * test.e_square(peak_freqs, peak_amps) / 1e-8
    noise_integral = full_integral - peak_integral

    noise_integrals[i] = noise_integral
    peak_integrals[i] = peak_integral

ex_table = excel.Workbook()
ex_table.create_sheet(title='W_df', index=0)
sheet = ex_table['W_df']
sheet['A1'] = 'df_MHz'
sheet['B1'] = 'W_peak'
sheet['C1'] = 'W_noise'
table_size = df_s.size
for k in range(table_size):
    cell = sheet.cell(row=k+2, column=1)
    cell.value = df_s[k] / 1e6
    cell = sheet.cell(row=k+2, column=2)
    cell.value = peak_integrals[k]
    cell = sheet.cell(row=k+2, column=3)
    cell.value = noise_integrals[k]
path = test.excel_folder_path / 'w_df_dependence.xlsx'
ex_table.save(path)
