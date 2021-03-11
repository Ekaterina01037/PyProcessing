import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from ProcessClass_10 import ProcessSignal
import openpyxl as xl
from pathlib import Path

#path_line = r'C:\Users\d_Nice\Documents\SignalProcessing\2019\191120\Excel\integral_№2, 3 (четверть).xlsx'
main_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021')


def plots(exp_num, reb=True):
    path_line = main_path / f'{exp_num}/Excel/Integrals_{exp_num}.xlsx'
    wb = xl.load_workbook(path_line)
    sheet = wb['Integral']
    rows = sheet.max_row
    cols = sheet.max_column
    file_data = {}
    for j in range(cols):
        cell_k = sheet.cell(row=1, column=j + 1)
        key_name = cell_k.value
        vals = []
        for i in range(1, rows):
            cell = sheet.cell(row=i + 1, column=j + 1)
            cell_val = cell.value
            if cell_val is not None:
                vals.append(cell_val)
        file_data[key_name] = vals
    print('Obtained file data')
    fig = plt.figure(num=1, dpi=200, figsize=[11.69, 8.27])
    ax = fig.add_subplot(111)
    pl_densities = file_data['Плотность плазмы, отн.ед.']
    e_peak_ints = np.asarray(file_data['W_f0, *10-8'])
    mag_noise_ints = np.asarray(file_data['W_1, *10-8'])
    reb_pls = file_data['Плотность плазмы, отн.ед.(шумы)']
    pl_e_ints = np.asarray(file_data['W_2, *10-8'])

    line, = ax.plot(pl_densities, mag_noise_ints, marker='^', linewidth=2, color='green')
    line.set_label('Шумы усилителя\nсигнала магнетрона')
    if reb:
        line3, = ax.plot(reb_pls, pl_e_ints, marker='D', linewidth=2, color='red')
        line3.set_label('Шумы в отсутствие\nвходного сигнала')
        path_png = main_path / f'{exp_num}/Pictures/reb_presentation_{exp_num}'
        x_min = min(reb_pls[0], pl_densities[0]) - 0.5
        x_max = max(reb_pls[-1], pl_densities[-1]) + 0.5
        ax.set_ylim(bottom=0, top=6.5)
    else:
        line, = ax.plot(pl_densities, e_peak_ints, marker='o', linewidth=2, color='blue')
        line.set_label('Энергия в пике')
        path_png = main_path/f'{exp_num}/Pictures/peak_presentation_{exp_num}'
        #x_min = pl_densities[0] - 0.5
        #x_max = pl_densities[-1] + 0.5
        x_min, x_max = 5, 14
        #y_max = max(e_peak_ints) + 0.5
        y_max = 22
        ax.set_ylim(bottom=0, top=y_max)
    ax.set_ylabel(r'$W*10^{-8}, [B^{2}c] $', fontsize=28, fontweight='black')

    ax.set_xlim(left=x_min, right=x_max)
    print('x_min =', x_min, 'x_max=', x_max)
    ax.legend(loc='best', fontsize=18)
    ax.grid(which='both', axis='both')
    ax.set_xlabel(r'$n_p, отн.ед.$', fontsize=28, fontweight='black')
    plt.xticks(fontsize=24)
    plt.yticks(fontsize=24)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
    fig.savefig(path_png)
    plt.close(fig)


#plots()
plots(210119, reb=False)