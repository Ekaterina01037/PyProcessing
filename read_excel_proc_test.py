import openpyxl as xl
from openpyxl.utils import get_column_letter
from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np
import csv
from scipy.fftpack import rfft, rfftfreq, irfft

main_excel_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2020\210220\210220.xlsx')
type_file_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2020\210220\Excel\types.xlsx')


def read_excel():
    wb = xl.load_workbook(main_excel_path)
    sheet = wb['Лист1']
    max_row = sheet.max_row
    max_col = sheet.max_column
    min_row = 1

    for i in range(1, max_col):
        for j in range(1, max_row):
            cell = sheet.cell(row=j, column=i)
            cell_val = str(cell.value)
            if 'файл' in cell_val and min_row == 1:
                min_row, min_col = j, i
    min_col_letter, max_col_letter = get_column_letter(min_col), get_column_letter(max_col)
    #cells_data = [sheet[f'{min_col_letter}{min_row}':f'{max_col_letter}{max_row}'][0][i][j].value for i, j in range(max_col, max_row)]
    data_keys = [sheet[f'{min_col_letter}{min_row}':f'{max_col_letter}{min_row}'][0][i].value for i in range(max_col)]
    rows_data = []
    for row in range(min_row, max_row):
        row_data = [sheet[f'{min_col_letter}{row}':f'{max_col_letter}{row}'][0][i].value for i in range(max_col)]
        print(row_data)
    print(cells_data)
    keys = ['file_nums', 'plasma_source_heat', 'plasma_current', 'pressure', 'magnetic_field', 'magnetron_delay',
            'magnetron_start', 'GIN_voltage', 'lamp_detector', 'magnetron_charge_voltage', 'magnetron_absorbers',
            'magnetron_voltage', 'U_in', 'comment']
    for k in range(1, max_col + 1):
        cell = sheet.cell(row=min_row, column=k)
        cell_name = cell.value
        if 'файл' in cell_name:
            cell_key = 'num'
        keys.append(cell_key)
#read_excel()
'''
    row_dicts = []
    for l in range(min_row + 1, max_row + 1):
        row_dict = {}
        vals = []
        for m in range(1, max_col + 1):
            cell = sheet.cell(row=l, column=m)
            cell_val = cell.value
            vals.append(cell_val)
        for i, key in enumerate(keys):
            row_dict[key] = vals[i]
        row_dicts.append(row_dict)

    plasma_dicts = []
    reb_dicts = []
    magnetron_dicts = []
    for row_dict in row_dicts:
        fnum = row_dict['Номер файла']
        gin_voltage = row_dict['Напряжение ГИНа']
        d_plasma = row_dict['Ток плазмы, А']
        heating = row_dict['Накал']
        magnetron_delay = row_dict['Задержка магнетрона, нс']
        # magnetron_in_voltage = row_dict['Входное напряжение магнетрона, В']
        comment = 0
        if isinstance(fnum, int):
            if isinstance(d_plasma, float) or isinstance(d_plasma, int):
                if comment != 'except':
                    if isinstance(magnetron_delay, float) or isinstance(magnetron_delay, int):
                        magnetron_dicts.append(row_dict)
                    else:
                        plasma_dicts.append(row_dict)
            elif heating is None and isinstance(gin_voltage, float):
                reb_dicts.append(row_dict)

    useful_dicts = [plasma_dicts, reb_dicts, magnetron_dicts]

    proc_signal_dicts = {}
    use_nums = {}
    list_signals = ['noise', 'reb', 'magnetron']
    for i, use_dict in enumerate(useful_dicts):
        # use_csv_files = []
        fnums = []
        for row_dict in use_dict:
            fnum = row_dict['Номер файла']
            if fnum < 100:
                num_1 = fnum // 10
                num_2 = fnum % 10
            elif fnum > 10000:
                num_1 = fnum // 1000
                num_2 = fnum % 1000
            else:
                num_1 = fnum // 100
                num_2 = fnum % 100
            fnum_1 = '{:03d}'.format(num_1)
            fnum_2 = '{:03d}'.format(num_2)

            if fnum_1 in csv_signal_nums:
                row_dict['Номер файла'] = [fnum_1]
                fnums.append(fnum_1)
                # use_csv_files.append(row_dict)
                proc_signal_dicts[fnum_1] = row_dict
            elif fnum_2 in csv_signal_nums:
                row_dict['Номер файла'] = [fnum_2]
                fnums.append(fnum_2)
                # use_csv_files.append(row_dict)
                proc_signal_dicts[fnum_2] = row_dict
        # proc_signal_dicts[list_signals[i]] = use_csv_files
        use_nums[list_signals[i]] = fnums

    excel_results = {'dicts': proc_signal_dicts,
                     'numbers': use_nums}
    return excel_results
'''
folder_path = Path(r'C:\Users\d_Nice\Documents\SignalProcessing\2021\210302')


def open_file(file_name):
    file_path = folder_path / 'str{}.csv'.format(file_name)
    t = []
    u = []
    with open(file_path) as File:
        reader = csv.reader(File)
        for row in reader:
            t_val = float(row[3])
            t.append(t_val)
            u_val = float(row[4])
            u.append(u_val)
    t = np.asarray(t)
    u = np.asarray(u)
    file_dict = {'time': t,
                 'voltage': u}
    return file_dict


def fft_amplitude(t, u):
    dt = np.abs(t[1] - t[0])
    len_t = len(t)
    u_fft = rfft(u)
    freq_fft = rfftfreq(len_t, dt)
    n = len(u)
    p_freq = freq_fft[1:n:2]
    n_freq = len(p_freq)
    amp_fft = np.zeros(n_freq)
    amp_fft[0] = u_fft[0] / n
    if n % 2 == 0:
        amp_fft[n_freq - 1] = u_fft[-1] / n
        j = 1
        for i in range(1, n_freq - 2):
            amp_fft[i] = 2 * np.sqrt(u_fft[j] ** 2 + u_fft[j + 1] ** 2) / n
            j += 2
    if n % 2 != 0:
        j = 1
        for i in range(1, n_freq - 1):
            amp_fft[i] = 2 * np.sqrt(u_fft[j] ** 2 + u_fft[j + 1] ** 2) / n
            j += 2

    ind_fake_freqs_1 = np.logical_or(p_freq <= 1.249e9, 1.251e9 <= p_freq)
    p_freq_1 = p_freq[ind_fake_freqs_1]
    amp_fft_1 = amp_fft[ind_fake_freqs_1]

    ind_fake_freqs_2 = np.logical_or(p_freq_1 <= 2.499e9, 2.501e9 <= p_freq_1)
    p_freq_2 = p_freq_1[ind_fake_freqs_2]
    amp_fft_2 = amp_fft_1[ind_fake_freqs_2]

    #ind_cutoff = p_freq_2 <= cutoff_frequency
    #cut_freq = p_freq_2[ind_cutoff]
    #cut_fft_amp = amp_fft_2[ind_cutoff]
    fft_amp_dict = {'frequency': p_freq_2,
                    'amplitude': amp_fft_2}
    return fft_amp_dict


def single_fft_start_time(num, start_time=400e-9):
    file_data = open_file(num)
    t, v = file_data['time'], file_data['voltage']
    time_interval = 1024 * 8 * 2 * 16 * 1e-12
    print(time_interval)
    end_time = start_time + time_interval
    print(end_time)
    interval_inds = np.logical_and(t >= start_time, t <= end_time)
    t_plato, v_plato = t[interval_inds], v[interval_inds]
    #plt.plot(t_plato, v_plato)
    #plt.plot([t_plato], [v_plato])
    fft_results = fft_amplitude(t_plato, v_plato)
    freqs, amps = fft_results['frequency'][1::], fft_results['amplitude'][1::]
    peak_freq, peak = np.round(freqs[np.argmax(amps)] / 1e9, 3), np.max(amps)
    plt.plot(freqs, amps)
    #plt.xlim(left=2.7e9, right=2.74e9)
    plt.title(f'{num}, peak_freq ={peak_freq}')
    plt.show()
single_fft_start_time('097')


def fft_full(self, magnetron_full=False, magnetron_noise_base=True, peak=False, reb_noise=False, peak_freq=2.71e9,
             peak_gate=50e6):
    types = self.read_type_file()
    csv_signals, csv_signal_nums = types['signal_files'], types['signal_nums']
    excel_results = self.read_excel(csv_signal_nums)['numbers']
    noise_nums = excel_results['noise']
    magnetron_nums = excel_results['magnetron']
    if reb_noise:
        nums = noise_nums
    else:
        nums = magnetron_nums
    for num in nums:
        file_name = f'str{num}.csv'
        file_data = self.open_file(file_name, reduced=True)
        t, u, dt = file_data['time'], file_data['voltage'], file_data['time_resolution']
        pl_density = self.read_excel(file_name)['dicts'][num]['Ток плазмы, А']
        fft_results = self.fft_amplitude(t, u)
        freqs, amps = fft_results['frequency'], fft_results['amplitude']
        mean_freq = self.mean_frequency(freqs, amps)
        spectrum_mean_freq = mean_freq['mean_freq']

        fig = plt.figure(num=1, dpi=300)
        ax = fig.add_subplot(111)

        if magnetron_noise_base:
            base_inds = np.logical_or(freqs < peak_freq - peak_gate, freqs > peak_freq + peak_gate)
            base_freqs, base_amps = freqs[base_inds], amps[base_inds]
            mean_freq = self.mean_frequency(base_freqs, base_amps)
            spectrum_mean_freq = mean_freq['mean_freq']
            line, = ax.plot(base_freqs, base_amps, linewidth=0.7, color='mediumseagreen')
        else:
            line, = ax.plot(freqs, amps, linewidth=0.7, color='mediumseagreen')
        if peak:
            ax.set_xlim(left=peak_freq-30e6, right=peak_freq+30e6)
        else:
            ax.set_xlim(left=0, right=4e9)
        ax.grid(which='both', axis='both')
        ax.set_xlabel(r'$Frequency, GHz$')
        ax.set_ylabel(r'$Amplitude$')
        ax.set_title(r'$№={}, n={}$'.format(num, pl_density))
        if not peak:
            line.set_label(r'$f = {} GHz$'.format(spectrum_mean_freq))
            ax.legend()
        if peak:
            png_name = self.fft_magnetron / 'peak_{num}'
        elif reb_noise:
            png_name = self.fft_magnetron / 'reb_noise_{num}'
        elif magnetron_noise_base:
            png_name = self.fft_magnetron / 'noise_base_{num}'
        else:
            png_name = self.fft_magnetron / 'amplifier_{num}'
        fig.savefig(png_name)
